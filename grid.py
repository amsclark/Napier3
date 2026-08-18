"""Make the workbook's derived sheets reach as far as its case list.

CASE DATA takes one row per case and build_workbook has nothing stopping it.
The sheets that read CASE DATA do stop, each at a different row, because each
was filled down by hand to whatever depth looked like enough:

    SOL              rows to 150   covers 147 cases
    EXPUNGEMENT      rows to 200   covers 198 cases
    EXEMPTIONS       rows to 200   covers 197 cases
    BANKRUPTCY       rows to 201   covers 198 cases
    LICENSE-REGIS    rows to 300   covers 299 cases
    POLK R&B APPEAL  rows to 300   covers 299 cases

and the totals at the top of SOL, BANKRUPTCY and EXEMPTIONS are shorter still,
=SUM(C4:C100), which covers 97.

Past those the workbook does not complain. The case is on CASE DATA with its
money in columns J to S, and it is absent from the analysis: not time barred,
not "NO ARGUMENT", not a pending charge blocking expungement, nowhere. Built
from the captured corpus at 210 cases, 63 of them are missing from the SOL
sheet entirely and its NO ARGUMENT total reads $6,727.91 against $6,895.38 of
rows the sheet itself computed.

A staffer has already pulled 70 cases in one run in production, so the totals
threshold of 98 is the one that matters. A client with decades of traffic and
scheduled-violation cases reaches it.

Nothing here decides anything. Every formula written is the sheet author's own
last row, translated down, and every range widened keeps its start and its
function and only stops later. The analysis is unchanged; it just reaches the
rows that were already there. A workbook small enough for the grid it shipped
with comes out exactly as before.
"""

import re
from copy import copy

from openpyxl.formula.translate import Translator
from openpyxl.worksheet.formula import ArrayFormula

# The sheets that hold the case list itself, the clinic header and the static
# statute lookup. None of them is a per-case analysis of CASE DATA.
NOT_DERIVED = ('CASE DATA', 'BASIC INFO', 'CODE SECTIONS')

# CASE DATA is the case list rather than a reading of it, so it is not mirrored
# down from a template row the way the sheets above are, and its own formulas
# name no sheet so the survey below cannot see them. Two things on it are still
# derived and still stop at row 300: column T, each row's own =SUM(J4:S4), and
# the totals across row 1, =SUM(J4:J300). Columns W to AH stop there too and are
# deliberately not touched here, because statutes.write_statute_split has
# already written every row of them as plain text by the time this runs.
CASE_DATA = 'CASE DATA'
FIRST_CASE_ROW = 4

# Far past any case list, and there to stop a sheet whose styled region runs to
# the bottom of the spreadsheet from being walked cell by cell.
ROW_CAP = 5000

# 'CASE DATA'!A6 and 'CASE DATA'!$D$4 both name a row. The dollar matters: a
# relative reference moves when the formula is translated down and an absolute
# one does not, which is exactly the difference between a per-row lookup and a
# count over the whole column.
CASE_REF = re.compile(r"'CASE DATA'!(\$?)([A-Z]{1,3})(\$?)(\d+)")

# A range on this same sheet, so not one preceded by a sheet name. Only the end
# row is ever touched.
LOCAL_RANGE = re.compile(r"(?<!!)\b([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)\b")

# A range over CASE DATA with an absolute row, which is the shape a count over
# every case takes. A relative one is a per-row lookup and is left alone.
CASE_RANGE = re.compile(
    r"('CASE DATA'!)(\$?[A-Z]{1,3}\$)(\d+):(\$?[A-Z]{1,3}\$)(\d+)")


# A cell entered with ctrl-shift-enter is not a string. openpyxl hands it back
# as an ArrayFormula object carrying the text and the range it was entered over,
# and a formula read as a string is silently not one of these. The expungement
# sheet keeps three of its columns this way, including the one that asks whether
# there is a subsequent conviction, so treating them as blank drops those
# columns off every row this module adds and leaves their ranges unwidened on
# the rows that were already there.
ARRAY_REF = re.compile(r'^([A-Z]{1,3})(\d+)(?::([A-Z]{1,3})(\d+))?$')


def _formula(cell):
    value = cell.value
    if isinstance(value, ArrayFormula):
        return value.text
    return value if isinstance(value, str) and value.startswith('=') else None


def _shift_ref(ref, row_delta):
    """The range a copied array formula is entered over, moved down with it."""
    match = ARRAY_REF.match(str(ref or ''))
    if not match:
        return ref
    start_col, start_row, end_col, end_row = match.groups()
    moved = '%s%d' % (start_col, int(start_row) + row_delta)
    if end_col:
        moved += ':%s%d' % (end_col, int(end_row) + row_delta)
    return moved


def _write(cell, formula, source=None, row_delta=0):
    """Put a formula back the way it was entered.

    An array formula written back as a plain string loses its ctrl-shift-enter
    marking, and the ones here are all SUM(COUNTIF(range, range)) over a range,
    which without that marking stops summing and answers about one cell.
    """
    original = (source or cell).value
    if isinstance(original, ArrayFormula):
        cell.value = ArrayFormula(
            ref=_shift_ref(original.ref, row_delta), text=formula)
    else:
        cell.value = formula


def _case_rows(formula):
    """The CASE DATA rows a formula names relatively, so the ones that move."""
    return {int(m.group(4)) for m in CASE_REF.finditer(formula)
            if not m.group(3)}


def _survey(sheet):
    """Every row of this sheet that reads a case row, and which row it reads.

    Returns (rows, deepest, furthest) where rows maps a sheet row to the
    highest case row it reads, deepest is the last such sheet row and furthest
    is the highest case row the sheet reads anywhere.
    """
    rows = {}
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, ROW_CAP)):
        best = None
        for cell in row:
            formula = _formula(cell)
            if not formula:
                continue
            named = _case_rows(formula)
            if named:
                best = max(named) if best is None else max(best, max(named))
        if best is not None:
            rows[row[0].row] = best
    if not rows:
        return rows, None, None
    return rows, max(rows), max(rows.values())


def _widen_local_ranges(formula, first_case_row, last_row):
    """Push a range down the case rows to last_row, never up.

    SUM(C4:C100) on a sheet now carrying 210 cases has to become SUM(C4:C213).
    A range that already reaches far enough is left exactly as written, so a
    workbook that fits its grid keeps the formula it shipped with.

    Only a range that starts in the case rows is one of these. SOL's grand
    total is =SUM(C1:E1), three cells of the totals row itself, and stretching
    that to E150 adds every per-row figure to the totals that already contain
    them and doubles the sheet's headline number.
    """
    def replace(match):
        start_col, start_row, end_col, end_row = match.groups()
        if int(start_row) < first_case_row or int(end_row) >= last_row:
            return match.group(0)
        return '%s%s:%s%d' % (start_col, start_row, end_col, last_row)
    return LOCAL_RANGE.sub(replace, formula)


def _widen_case_ranges(formula, last_case_row):
    """The same for a count over every case row.

    The expungement sheet finds pending charges with
    COUNTIF('CASE DATA'!$D$4:$D$200,"") and a case at row 201 is a pending
    charge it cannot see. That count is what says whether anything is still
    hanging over the client under 901C.2, so it has to cover the whole list.
    """
    def replace(match):
        prefix, start, start_row, end, end_row = match.groups()
        if int(end_row) >= last_case_row:
            return match.group(0)
        return '%s%s%s:%s%d' % (prefix, start, start_row, end, last_case_row)
    return CASE_RANGE.sub(replace, formula)


def _extend_sheet(sheet, last_case_row):
    """Fill this sheet's last row of formulas down to cover last_case_row.

    Returns (rows added, the sheet's last formula row afterwards, and the
    first row of it that reads a case, which is where the header ends).

    The sheets do not agree on where a case lands. SOL row 4 reads case row 4;
    the expungement sheet starts a row higher, its row 3 reading case row 4;
    LICENSE-REGIS and POLK R&B APPEAL start at row 2. So the offset is measured
    from the sheet's own last row rather than assumed. Order does not matter to
    any of these sheets, because nothing on them reads a neighbouring row,
    which is why LICENSE-REGIS could ship reading the case list shuffled and
    still add up.
    """
    rows, deepest, furthest = _survey(sheet)
    if deepest is None:
        return 0, None, None

    first = min(rows)
    reached = rows[deepest]
    if reached < furthest:
        # The sheet reads a higher case row somewhere above its last row, so
        # filling down from that last row would leave a gap or a repeat. None
        # of the shipped templates is shaped this way, and guessing at one that
        # is would be worse than leaving it alone.
        return 0, deepest, first
    if reached >= last_case_row:
        return 0, deepest, first

    # _fill_down copies each cell's style along with its formula. Without that
    # the new rows lose the currency and date formats the filled-down ones
    # carry, and a dollar figure printed as a bare number reads as a different
    # kind of number.
    added = last_case_row - reached
    _fill_down(sheet, deepest, added)
    return added, deepest + added, first


def _fill_down(sheet, template_row, added):
    """Copy a row's formulas and formats down `added` rows."""
    template = [cell for cell in sheet[template_row] if _formula(cell)]
    for step in range(1, added + 1):
        for cell in template:
            target = sheet.cell(row=template_row + step, column=cell.column)
            _write(target,
                   Translator(_formula(cell),
                              origin=cell.coordinate).translate_formula(
                                  row_delta=step, col_delta=0),
                   source=cell, row_delta=step)
            target._style = copy(cell._style)


def _deepest_formula_row(sheet, first_row):
    """The last row at or below first_row holding a formula of any kind."""
    deepest = None
    for row in sheet.iter_rows(min_row=first_row,
                               max_row=min(sheet.max_row, ROW_CAP)):
        if any(_formula(cell) for cell in row):
            deepest = row[0].row
    return deepest


def _extend_case_data(sheet, last_case_row):
    """Fill CASE DATA's own derived cells down and widen its totals.

    Returns the rows added. Column T is what the expungement sheet reads for
    the debt subject to 910.7, and row 1 is the figure staff read off the case
    list itself, so a case list past row 300 loses both without saying so.
    """
    deepest = _deepest_formula_row(sheet, FIRST_CASE_ROW)
    if deepest is None:
        return 0

    added = 0
    if deepest < last_case_row:
        added = last_case_row - deepest
        _fill_down(sheet, deepest, added)

    for cell in sheet[1]:
        formula = _formula(cell)
        if not formula:
            continue
        widened = _widen_local_ranges(formula, FIRST_CASE_ROW,
                                      max(deepest + added, last_case_row))
        if widened != formula:
            _write(cell, widened)
    return added


def extend_formula_grid(workbook, case_count):
    """Widen every derived sheet to the number of cases actually written.

    Call once the CASE DATA rows are in and before the workbook is saved.
    Returns the sheets that had to change, which is empty for any workbook that
    already fits the grid the templates ship with.
    """
    if case_count <= 0:
        return {}

    last_case_row = 3 + case_count
    changed = {}

    added = _extend_case_data(workbook[CASE_DATA], last_case_row)
    if added:
        changed[CASE_DATA] = added

    for sheet in workbook.worksheets:
        if sheet.title in NOT_DERIVED:
            continue

        added, deepest, first_case_row = _extend_sheet(sheet, last_case_row)
        if deepest is None:
            continue
        touched = added

        # Ranges come after the fill so the totals can see the rows just added.
        # This runs whether or not the sheet grew: the totals on SOL, BANKRUPTCY
        # and EXEMPTIONS stop at row 100 while their per-row formulas run to 150
        # and beyond, so a workbook can sit inside the grid and still be
        # under-totalled. Widening a range on a per-case row would be a
        # different thing entirely, so only the header rows above the case rows
        # are eligible for it.
        for row in sheet.iter_rows(min_row=1,
                                   max_row=min(sheet.max_row, ROW_CAP)):
            for cell in row:
                formula = _formula(cell)
                if not formula:
                    continue
                widened = _widen_case_ranges(formula, last_case_row)
                if cell.row < first_case_row:
                    widened = _widen_local_ranges(
                        widened, first_case_row, deepest)
                if widened != formula:
                    _write(cell, widened)
                    touched = touched or 1
        if touched:
            changed[sheet.title] = added
    return changed


def _range_shortfalls(sheet, first_case_row, last_row, last_case_row):
    """Ranges above the case rows that still stop before the case list ends."""
    short = set()
    for row in sheet.iter_rows(min_row=1,
                               max_row=min(sheet.max_row, first_case_row - 1)):
        for cell in row:
            formula = _formula(cell)
            if not formula:
                continue
            if _widen_local_ranges(formula, first_case_row,
                                   last_row) != formula:
                short.add(cell.coordinate)
            if _widen_case_ranges(formula, last_case_row) != formula:
                short.add(cell.coordinate)
    return short


def shortfalls(workbook, case_count):
    """Every sheet whose formulas still stop before the last case, measured.

    extend_formula_grid is meant to leave this empty, and this is here because
    of what the alternative did. The finish page used to work the warning out
    from the case count and the depths the templates shipped with, so it went
    on telling staff the analysis sheets stop at 147 cases, and the totals at
    97, for as long as it took somebody to notice that they had not since the
    grid was extended. A run of 184 cases came back correct and captioned as
    short, next to advice to split the search in two, which is both unnecessary
    and something no screen in Napier will do.

    So this asks the workbook instead of predicting it. Anything it names is a
    real gap in the file that was saved, and a workbook nothing is wrong with
    says nothing at all.
    """
    if case_count <= 0:
        return {}

    last_case_row = 3 + case_count
    short = {}

    sheet = workbook[CASE_DATA]
    deepest = _deepest_formula_row(sheet, FIRST_CASE_ROW)
    if deepest is not None:
        # On CASE DATA a sheet row is a case row, so the row the totals have to
        # cover is the last case row itself. A total that sums further down
        # than the case list is adding up blank rows, which is not a shortfall.
        short[CASE_DATA] = _reasons(sheet, FIRST_CASE_ROW, last_case_row,
                                    last_case_row, deepest < last_case_row)

    for sheet in workbook.worksheets:
        if sheet.title in NOT_DERIVED:
            continue
        rows, deepest, furthest = _survey(sheet)
        if deepest is None:
            continue
        # The sheets do not agree on where a case lands -- SOL's row 4 reads
        # case row 4, the expungement sheet's row 3 does, LICENSE-REGIS starts
        # two rows up -- so the row the last case sits on is found by counting
        # back from the deepest case the sheet does reach, not by taking the
        # case row as a sheet row. Where that lands above the sheet's own last
        # row, the rows past it hold no case and a total that stops short of
        # them is not short of anything.
        reach = deepest + (last_case_row - furthest)
        short[sheet.title] = _reasons(sheet, min(rows), reach,
                                      last_case_row, furthest < last_case_row)
    return {name: reasons for name, reasons in short.items() if reasons}


def _reasons(sheet, first_case_row, reach, last_case_row, rows_short):
    reasons = ['rows'] if rows_short else []
    if _range_shortfalls(sheet, first_case_row, reach, last_case_row):
        reasons.append('totals')
    return reasons


def describe_shortfalls(short):
    """What to put in front of staff about the sheets shortfalls() named.

    One sentence per sheet, because the two failures read differently on a
    finished file: a missing row is a case absent from an analysis, and a short
    total is a figure that looks right and is low.
    """
    lines = []
    for name in sorted(short):
        reasons = short[name]
        if name == CASE_DATA:
            # The case list itself is never missing a case; what it can be
            # missing is column T on the rows past where the template's own
            # formulas stop, and the totals across the top.
            lines.append(
                "CASE DATA's TOTAL column and the totals across its top row "
                "cover only part of the case list, so the rows are right and "
                "those figures are low.")
        elif 'rows' in reasons and 'totals' in reasons:
            lines.append(
                "%s stops before the last case, so the cases past that point "
                "are missing from it and its totals are low." % name)
        elif 'rows' in reasons:
            lines.append(
                "%s stops before the last case, so the cases past that point "
                "are on CASE DATA and nowhere else." % name)
        else:
            lines.append(
                "%s adds up only part of the case list, so its rows are right "
                "and its totals are low." % name)
    return lines
