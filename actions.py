"""The page a clinic actually works from.

The CRS workbook answers seven different questions on seven different sheets,
one row per case on each, and nobody sitting across from a client reads seven
sheets. The arguments worth making are scattered: a twenty year old attorney
fee is on SOL, the case it sits on is on CASE DATA, whether that case is also
expungeable is on EXPUNGEMENT, and whether the licence is the reason the client
came in is on LICENSE-REGIS. Finding the three that matter means cross
referencing all of them by case number, by hand, while somebody waits.

So this collects them into one ranked list. Every line names a case, what to do
about it, the authority for doing it, and the money at stake, and the list is
sorted so the biggest number is at the top.

Two rules this module holds to.

It reads the values Napier has already written into CASE DATA rather than
recomputing them from the ICOS pages. The analysis sheets read those same
cells, so a list built from them cannot disagree with the workbook it is
bound into. Recomputing from the source would have been tidier and would have
drifted the first time either side changed.

And it only mirrors tests the workbook already makes. The twenty year cut is
the SOL sheet's 7300 days, the dismissal test is the EXPUNGEMENT sheet's list
of codes, the licence test is LICENSE-REGIS's. Nothing here is a new legal
opinion. Where the sheet is silent this is silent too.
"""

from decimal import Decimal

from openpyxl.styles import Alignment, Font, PatternFill

import crs

# Twenty years, counted the way the SOL sheet counts it. That sheet writes
# 'CASE DATA'!D4+7300, so the arithmetic here is days rather than calendar
# years on purpose: matching the workbook matters more than being right about
# leap days, and a list that disagreed with the sheet beside it would be worse
# than either.
SOL_DAYS = 7300

# The disposition codes each sheet tests for, taken from the sheets.
CONVICTION_CODES = frozenset({'GTR', 'GPL', 'DEF'})  # LICENSE-REGIS
CLEARED_CODES = frozenset({'DISM', 'ACQ', 'NOTF', 'WTHD', 'TNSF'})  # EXPUNGEMENT I
JUVENILE_CODE = 'JUV'

# Columns, by the letter they carry on CASE DATA.
FEE_COLUMNS = 'JKLMNOPQRS'         # everything owed, which is what T sums
REMISSIBLE_COLUMNS = 'JKLMNOP'     # what 910.7 can reach, per EXPUNGEMENT D
ATTORNEY_AND_COLLECTION = 'JK'     # SOL C
ROOM_AND_BOARD = 'L'               # SOL D, and the whole of POLK R&B APPEAL

# Ranking. The list is sorted by money first, so these only decide ties, but
# they also decide what a staffer reads first among the rows carrying no
# figure at all.
TIER_TIME_BARRED = 1
TIER_DISMISSED = 2
TIER_REMISSION = 3
TIER_ROOM_AND_BOARD = 4
TIER_MISDEMEANOUR = 5
TIER_LICENCE = 6
TIER_JUVENILE = 7

HEADER_FILL = PatternFill('solid', fgColor='DDDDDD')
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
CAVEAT_FONT = Font(italic=True, color='996600')


def _decimal(value):
    """A money cell as a Decimal. Blank, text and None all come back as zero."""
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    return Decimal(0)


def _sum(facts, columns):
    return sum((_decimal(facts['money'].get(column)) for column in columns),
               Decimal(0))


def row_facts(worksheet, row):
    """What Napier wrote on one CASE DATA row, as the sheets will read it.

    Dates come back as dates. Napier writes them as the strings ICOS uses and
    Excel coerces those in arithmetic, which is why the SOL sheet works, but
    nothing in Python coerces them and a string compared against a date is a
    TypeError rather than a wrong answer.
    """
    def cell(column):
        return worksheet[column + str(row)].value

    return {
        'row': row,
        'id': cell('A'),
        'county': cell('B'),
        'disposition_date': crs.parse_us_date(cell('D')),
        'description': cell('E'),
        'statute': cell('F'),
        'code': (cell('G') or '').strip().upper(),
        'vehicular': (cell('H') or '').strip().upper() == 'YES',
        'supervised': (cell('I') or '').strip().upper() == 'YES',
        'money': {column: cell(column) for column in FEE_COLUMNS},
        'icos_total': _decimal(cell('U')),
        'notes': cell('V') or '',
    }


def is_traffic(facts):
    """The EXPUNGEMENT sheet's TRAFFIC test, which is a bar to 901C.2.

    A simple misdemeanour under chapter 321, or any case whose number carries
    ST or NT in positions 8 and 9. Reproduced from the sheet rather than
    reasoned about, so that a case this calls traffic is exactly the set the
    sheet calls traffic.
    """
    case_id = facts['id'] or ''
    kind = case_id[7:9].upper()
    if kind in ('ST', 'NT'):
        return True
    return kind == 'SM' and (facts['statute'] or '').startswith('321.')


# Eight years, the way the EXPUNGEMENT sheet counts it: C3+2920.
MISDEMEANOUR_WAIT_DAYS = 2920

# Chapters the sheet rules out by prefix rather than by listing every section.
INELIGIBLE_CHAPTERS = ('717C', '719', '720', '724', '726', '728', '901A')


def code_lists(workbook):
    """The felony and ineligible-misdemeanour lists, read out of the workbook.

    CODE SECTIONS carries both, and the EXPUNGEMENT sheet's own eligibility
    formulas match against them by prefix. Reading them here rather than
    keeping a second copy in Python means a clinic that adds a section to the
    sheet gets it in the action list too, and the two cannot drift.
    """
    sheet = workbook['CODE SECTIONS']
    felonies = [str(cell.value).strip() for (cell,) in
                sheet.iter_rows(min_row=2, min_col=22, max_col=22)
                if cell.value]
    misdemeanours = [str(cell.value).strip() for (cell,) in
                     sheet.iter_rows(min_row=2, min_col=24, max_col=24)
                     if cell.value]
    return felonies, misdemeanours


def statutes(facts):
    """The statutes on a case, the way the helper columns split them.

    Napier joins a case's adjudicated statutes with semicolons into column F,
    and W through AH split them back out for the eligibility formulas to match
    against. This is that split.
    """
    return [part.strip() for part in (facts['statute'] or '').split(';')
            if part.strip()]


def _matches(statute, prefixes):
    return any(statute.upper().startswith(prefix.upper()) for prefix in prefixes)


def misdemeanour_expungement(facts, as_of, felonies, ineligible):
    """The EXPUNGEMENT sheet's 901C.3 test, columns M, N and S.

    Misdemeanour conviction, no section on the felony list, nothing on the
    ineligible list or in an excluded chapter, and eight years since the
    disposition. Returns True only when all four hold.

    What it does not check, because the sheet asks a person to: a subsequent
    conviction, a pending charge, a CDL, and whether the debt is paid off,
    which 901C.3 requires before anything gets expunged.
    """
    if facts['code'] not in ('GTR', 'GPL') or is_traffic(facts):
        return False
    found = statutes(facts)
    if not found:
        return False
    if any(_matches(statute, felonies) for statute in found):
        return False
    if any(_matches(statute, ineligible) or _matches(statute, INELIGIBLE_CHAPTERS)
           for statute in found):
        return False
    when = facts['disposition_date']
    return when is not None and (as_of - when).days > MISDEMEANOUR_WAIT_DAYS


def time_barred(facts, as_of):
    """Whether the disposition is more than twenty years behind the clinic."""
    when = facts['disposition_date']
    if when is None:
        return False
    return (as_of - when).days > SOL_DAYS


def case_actions(facts, as_of, sheets, codes=((), ())):
    """Everything worth doing about one case, as (tier, amount, do, cite, why).

    sheets is the set of sheet names the workbook has, because the Lite
    workbook ships without SOL, BANKRUPTCY and EXEMPTIONS, and pointing a
    staffer at a sheet that is not in the file they are holding is worse than
    saying nothing.
    """
    found = []
    owed = _sum(facts, FEE_COLUMNS)

    # Time barred. The oldest debt is the easiest to argue away and the least
    # likely to be noticed, because nothing about the row says how old it is
    # except a date in column D.
    if time_barred(facts, as_of):
        stale = _sum(facts, ATTORNEY_AND_COLLECTION) + _sum(facts, ROOM_AND_BOARD)
        if stale > 0:
            why = ("Disposed %s, more than 20 years before the clinic date. "
                   "%s in attorney, collection and room and board charges."
                   % (facts['disposition_date'].strftime('%m/%d/%Y'),
                      _dollars(stale)))
            if 'SOL' in sheets:
                why += " The SOL sheet counts this row."
            found.append((
                TIER_TIME_BARRED, stale,
                "Object to collection as time barred",
                "Iowa Code 614.1(6)",
                why,
            ))

    # Dismissed or acquitted. All of the debt on the case goes, not a category
    # of it, which is why this outranks everything except a bigger number.
    if facts['code'] in CLEARED_CODES and not is_traffic(facts):
        found.append((
            TIER_DISMISSED, owed,
            "Apply to expunge, and to discharge the debt with it",
            "Iowa Code 901C.2",
            "Coded %s and not a traffic case, so the whole balance on this "
            "case is dischargeable, not just part of it." % facts['code'],
        ))

    # Remission while a term is running. Column I is the gate the EXPUNGEMENT
    # sheet uses, and Napier only writes it where the sentence table says a
    # supervision term has not run out yet.
    if facts['supervised']:
        remissible = _sum(facts, REMISSIBLE_COLUMNS)
        if remissible > 0:
            found.append((
                TIER_REMISSION, remissible,
                "Apply to have the court reduce or remit this debt",
                "Iowa Code 910.7",
                "Still under supervision on the clinic date, so 910.7 is open. "
                "Covers everything but surcharges, fines and victim "
                "restitution. Column V says what term Napier read and what it "
                "could not see.",
            ))

    # Room and board. Polk has its own sheet because Polk is where these get
    # appealed, but the charge itself is a charge anywhere.
    board = _sum(facts, ROOM_AND_BOARD)
    if board > 0 and 'POLK' in (facts['county'] or '').upper():
        found.append((
            TIER_ROOM_AND_BOARD, board,
            "Challenge the jail room and board claim",
            "Iowa Code 356.7",
            "%s of room and board on a Polk County case. The POLK R&B APPEAL "
            "sheet lists this row." % _dollars(board),
        ))

    # The licence. No money comes off for this one, and it is still the reason
    # most people are in the room. The figure beside it is the balance that has
    # to be dealt with before the licence comes back, which is why it is not
    # zero: it is what is at stake, not what can be knocked off.
    # The registration hold is not here on purpose. It applies to every
    # conviction carrying a balance, so as a row it fired on most of the case
    # list and pushed the arguments that are actually worth an hour off the top
    # of the page. It is one line in the header block instead.
    if facts['vehicular'] and facts['code'] in CONVICTION_CODES and owed > 0:
        found.append((
            TIER_LICENCE, owed,
            "Ask about an installment agreement to get the licence back",
            "Iowa Code 321.210A, 321.210B",
            "Vehicular conviction carrying %s. An installment agreement lifts "
            "the suspension, needs a financial statement, and a person only "
            "gets five in a lifetime, so check how many have been used."
            % _dollars(owed),
        ))

    # 901C.3. Unlike a dismissal, this one does not wipe the debt out; the debt
    # has to be gone first. So the figure beside it is what stands between the
    # client and a clean record, and the row says which way round that is.
    if misdemeanour_expungement(facts, as_of, codes[0], codes[1]):
        found.append((
            TIER_MISDEMEANOUR, owed,
            "Check 901C.3 expungement, the debt has to be cleared first",
            "Iowa Code 901C.3",
            "Misdemeanour conviction disposed %s, more than 8 years ago, and "
            "not on the sheet's ineligible list. %s still owed, and 901C.3 "
            "needs it paid before anything is expunged. Napier has not checked "
            "for a later conviction, a pending charge or a CDL, which the "
            "EXPUNGEMENT sheet asks you to."
            % (facts['disposition_date'].strftime('%m/%d/%Y'), _dollars(owed)),
        ))

    if facts['code'] == JUVENILE_CODE:
        found.append((
            TIER_JUVENILE, owed,
            "Check whether the juvenile record can be sealed",
            "Iowa Code 232.150",
            "Juvenile case. The EXPUNGEMENT sheet works out the age and the "
            "two year wait; this only says the case is one to look at.",
        ))

    return found


def _dollars(amount):
    return "${:,.2f}".format(amount)


def registration_hold(worksheet, written):
    """Cases where the county treasurer can refuse to renew a registration.

    Every conviction carrying a balance qualifies, which is why this is a
    count rather than a row per case: as rows it was most of the list, and it
    is the same sentence every time.
    """
    cases, owed = 0, Decimal(0)
    for row in range(crs.FIRST_CASE_ROW, crs.FIRST_CASE_ROW + written):
        facts = row_facts(worksheet, row)
        if not facts['id'] or facts['code'] not in CONVICTION_CODES:
            continue
        balance = _sum(facts, FEE_COLUMNS)
        if balance > 0:
            cases += 1
            owed += balance
    return cases, owed


def collect(worksheet, written, as_of, sheets, codes=((), ())):
    """Every action across every case Napier wrote, best first."""
    found = []
    for row in range(crs.FIRST_CASE_ROW, crs.FIRST_CASE_ROW + written):
        facts = row_facts(worksheet, row)
        if not facts['id']:
            continue
        for action in case_actions(facts, as_of, sheets, codes):
            found.append((facts, action))
    # Money first, because a clinic hour spent on the biggest number is the
    # hour best spent. The tier only settles ties, and the case number after
    # that, so the same workbook always comes out in the same order.
    found.sort(key=lambda pair: (-pair[1][1], pair[1][0], pair[0]['id'] or ''))
    return found


# The list is Napier's reading of a court record, not advice, and it goes to a
# clinic where somebody will act on it inside an hour. It says so on the sheet
# rather than in a manual nobody has open.
CAVEAT = ("Napier worked these out from what Iowa Courts Online shows today. "
          "Check each one against the case before you file anything. Nothing "
          "here has been read by a lawyer.")

ACTION_HEADERS = [
    ("#", 5),
    ("CASE #", 20),
    ("COUNTY", 14),
    ("WHAT TO DO", 44),
    ("AUTHORITY", 26),
    ("AT STAKE", 13),
    ("WHY NAPIER SAYS SO", 78),
]

PAYMENT_HEADERS = [
    ("CASE #", 20),
    ("DATE", 12),
    ("AMOUNT", 12),
    ("PAID TOWARD", 40),
    ("RECEIPT", 14),
    ("HOW PAID", 11),
]

# The block above the ranked list, one line per row, named rather than counted.
# Adding a line to it used to mean finding every place that had worked out the
# row numbers by hand, and the row after the last of these is the header.
SUMMARY_ROWS = {
    'clinic': 2,
    'cases': 3,
    'owed': 4,
    'payments': 5,
    'hold': 6,
    'judgments': 7,
    'missing': 8,
    'caveat': 9,
}

FIRST_ACTION_ROW = max(SUMMARY_ROWS.values()) + 2

# The missing-case line is the one thing on this sheet that has to be read
# rather than skimmed, so it gets the colour the caveat gets.
MISSING_FONT = Font(bold=True, color='996600')


def _headers(worksheet, row, spec):
    for index, (label, width) in enumerate(spec, start=1):
        cell = worksheet.cell(row, index, label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        worksheet.column_dimensions[cell.column_letter].width = width


def _money_cell(worksheet, row, column, amount):
    cell = worksheet.cell(row, column, amount)
    cell.number_format = '"$"#,##0.00'
    return cell


def client_payments(cases, as_of):
    """One payment history for the client, across every case.

    A per case figure answers whether this case was being paid. What a court
    asks is whether the person pays, which is every case at once, and what
    they are paying now rather than what they paid in 1998.
    """
    every = []
    for case in cases:
        for payment in crs.payments(case):
            every.append(dict(payment, case=case['id']))
    if not every:
        return None
    every.sort(key=lambda payment: payment['date'])
    total = sum((payment['amount'] for payment in every), Decimal(0))
    cutoff = crs._add_term(as_of, -crs.RECENT_MONTHS, 'Month')
    recent = sum((payment['amount'] for payment in every
                  if payment['date'] > cutoff), Decimal(0))
    return {
        'payments': every,
        'count': len(every),
        'total': total,
        'first': every[0]['date'],
        'last': every[-1]['date'],
        'recent': recent,
        'recent_monthly': (recent / crs.RECENT_MONTHS).quantize(Decimal('0.01')),
        'cases': len({payment['case'] for payment in every}),
    }


def client_judgments(cases):
    """The civil judgments across every case, which are not court debt.

    They are kept off the payment history and out of the balance on purpose,
    and they still have to be said. A client being garnished on a judgment has
    less money for court debt than the same client without one, which is the
    whole of an ability-to-pay argument, and the number is usually an order of
    magnitude above anything else on the page.
    """
    found = []
    for case in cases:
        for judgment in crs.judgments(case):
            found.append(dict(judgment, case=case['id']))
    if not found:
        return None
    return {
        'count': len(found),
        'cases': len({judgment['case'] for judgment in found}),
        'total': sum((judgment['amount'] for judgment in found), Decimal(0)),
        'satisfied': sum((judgment['satisfied'] for judgment in found),
                         Decimal(0)),
    }


def build_action_sheet(workbook, cases, written, as_of, def_name, failed=()):
    """Add ACTION LIST to a workbook Napier has finished writing.

    Called after CASE DATA is filled in, because it reads CASE DATA back.
    Returns the ability-to-pay figures, which are a by-product of what it had
    to work out anyway.

    failed is the cases Iowa Courts would not give up. They get a line of their
    own, because this file outlives the page that says so.
    """
    sheets = set(workbook.sheetnames)
    found = collect(workbook['CASE DATA'], written, as_of, sheets,
                    code_lists(workbook))
    history = client_payments(cases, as_of)

    # At the back, with PAYMENTS, rather than in front of the workbook. This
    # sheet and that one are Napier's, and the seven in between are Iowa Legal
    # Aid's own criminal record summary, which is the thing staff open the file
    # to read and the thing they asked for. Nobody asked for a new sheet on top
    # of it. Position is presentation only: every formula in the template names
    # the sheet it reads ('CASE DATA'!G4, SOL!C4), so none of them care, and
    # putting this last also leaves BASIC INFO as the sheet the file opens on.
    worksheet = workbook.create_sheet('ACTION LIST')
    worksheet['A1'] = 'ACTION LIST'
    worksheet['A1'].font = TITLE_FONT
    worksheet['D1'] = def_name

    def line(name, label):
        row = SUMMARY_ROWS[name]
        worksheet.cell(row, 1, label)
        return 'B%d' % row

    worksheet[line('clinic', 'Clinic date')] = as_of
    worksheet['B%d' % SUMMARY_ROWS['clinic']].number_format = 'MM/DD/YYYY'

    worksheet[line('cases', 'Cases read')] = written

    line('owed', 'Total owed')
    total_owed = sum(
        (_sum(row_facts(workbook['CASE DATA'], row), FEE_COLUMNS)
         for row in range(crs.FIRST_CASE_ROW, crs.FIRST_CASE_ROW + written)),
        Decimal(0))
    _money_cell(worksheet, SUMMARY_ROWS['owed'], 2, total_owed)

    payments_cell = line('payments', 'Payments on record')
    if history is None:
        # Not the same as nothing paid, and the difference decides whether an
        # ability-to-pay argument leans on a payment record or has to be made
        # without one.
        worksheet[payments_cell] = "none in the ICOS itemization"
    else:
        worksheet[payments_cell] = (
            "%s across %d payment%s on %d case%s, %s to %s. Last 12 months: "
            "%s, about %s a month."
            % (_dollars(history['total']), history['count'],
               "" if history['count'] == 1 else "s",
               history['cases'], "" if history['cases'] == 1 else "s",
               history['first'].strftime('%m/%d/%Y'),
               history['last'].strftime('%m/%d/%Y'),
               _dollars(history['recent']),
               _dollars(history['recent_monthly'])))

    hold_cell = line('hold', 'Registration hold')
    held, held_owed = registration_hold(workbook['CASE DATA'], written)
    if held:
        worksheet[hold_cell] = (
            "%d convicted case%s carrying %s. The county treasurer can refuse "
            "to renew a registration over any of it (Iowa Code 321.40(6), "
            "602.8107(7))."
            % (held, "" if held == 1 else "s", _dollars(held_owed)))
    else:
        worksheet[hold_cell] = "none"

    # Not court debt, and kept out of every figure above on purpose. It is here
    # because a judgment is the largest number on the page and the one most
    # likely to be what is actually taking the client's money.
    judgment_cell = line('judgments', 'Civil judgments')
    judged = client_judgments(cases)
    if judged is None:
        worksheet[judgment_cell] = "none on these cases"
    else:
        worksheet[judgment_cell] = (
            "%d judgment%s on %d case%s, %s, %s of it shown satisfied. Not "
            "court debt, so none of it is in the figures above. A judgment "
            "being collected on is money the client does not have for court "
            "debt."
            % (judged['count'], "" if judged['count'] == 1 else "s",
               judged['cases'], "" if judged['cases'] == 1 else "s",
               _dollars(judged['total']), _dollars(judged['satisfied'])))

    # A workbook gets emailed to a colleague, saved to a shared drive and opened
    # three weeks later by somebody who never watched it being built. Up to now
    # the only place that said a run came back short was the finish page and the
    # progress log, both of which are gone within two hours, so the file itself
    # read as a complete criminal record when it was two cases shy of one.
    missing_cell = line('missing', 'Not in this file')
    failed = list(failed)
    if failed:
        worksheet[missing_cell] = (
            "%d case%s Iowa Courts would not give up: %s. %s on no sheet in "
            "this workbook. Look %s up on Iowa Courts before relying on this "
            "being the whole record."
            % (len(failed), "" if len(failed) == 1 else "s",
               ", ".join(failed),
               "It is" if len(failed) == 1 else "They are",
               "it" if len(failed) == 1 else "them"))
        worksheet[missing_cell].font = MISSING_FONT
    else:
        worksheet[missing_cell] = \
            "none. Every case the search turned up is here."

    caveat = worksheet.cell(SUMMARY_ROWS['caveat'], 1, CAVEAT)
    caveat.font = CAVEAT_FONT

    _headers(worksheet, FIRST_ACTION_ROW - 1, ACTION_HEADERS)
    row = FIRST_ACTION_ROW
    for rank, (facts, action) in enumerate(found, start=1):
        tier, amount, what, cite, why = action
        worksheet.cell(row, 1, rank)
        worksheet.cell(row, 2, facts['id'])
        worksheet.cell(row, 3, facts['county'])
        worksheet.cell(row, 4, what)
        worksheet.cell(row, 5, cite)
        _money_cell(worksheet, row, 6, amount)
        cell = worksheet.cell(row, 7, why)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

    if not found:
        cell = worksheet.cell(FIRST_ACTION_ROW, 1,
                              "Nothing on these cases matches an argument "
                              "Napier knows how to spot. That is not the same "
                              "as nothing being there.")
        cell.font = CAVEAT_FONT

    worksheet.freeze_panes = 'A' + str(FIRST_ACTION_ROW)
    build_payment_sheet(workbook, history)
    return ability_to_pay(total_owed, history)


def ability_to_pay(total_owed, history):
    """The two court-side numbers the ability-to-pay calculator asks for.

    The calculator at abilitytopay.org wants a court debt balance and what the
    client pays toward it each month, and then asks them for everything else:
    income, rent, groceries, all the things only the client can answer. These
    two are the ones a person sitting in a clinic cannot answer from memory,
    which is why the interview stalls on them.

    Both are already in the workbook. This is the same pair, taken from the
    same place, so the screen cannot disagree with the file. Returned as
    formatted text because that is all anyone does with them: read them off
    and type them in.

    monthly is None when ICOS itemized no payments at all, which is not the
    same as nothing having been paid and must not be handed over as $0.00.
    """
    return {
        'balance': _dollars(total_owed),
        'monthly': None if history is None else _dollars(history['recent_monthly']),
        'months': crs.RECENT_MONTHS,
    }


def build_payment_sheet(workbook, history):
    """Add PAYMENTS, the history behind the one line on the action list.

    Every payment ICOS records, newest first, with the receipt number to look
    it up by. This is the answer to "have you been paying anything?", which is
    the question that decides an ability to pay hearing, and until now it was
    downloaded on every run and thrown away.
    """
    worksheet = workbook.create_sheet('PAYMENTS')
    worksheet['A1'] = 'PAYMENT HISTORY'
    worksheet['A1'].font = TITLE_FONT
    if history is None:
        worksheet['A3'] = ("Iowa Courts Online shows no payments in the "
                           "itemization on any of these cases. That is what "
                           "the record says, not proof nothing was paid: a "
                           "case whose itemization ICOS does not publish "
                           "looks exactly the same.")
        worksheet['A3'].font = CAVEAT_FONT
        worksheet.column_dimensions['A'].width = 60
        return

    worksheet['A2'] = 'Total paid'
    _money_cell(worksheet, 2, 2, history['total'])

    _headers(worksheet, 4, PAYMENT_HEADERS)
    row = 5
    for payment in reversed(history['payments']):
        worksheet.cell(row, 1, payment['case'])
        cell = worksheet.cell(row, 2, payment['date'])
        cell.number_format = 'MM/DD/YYYY'
        _money_cell(worksheet, row, 3, payment['amount'])
        worksheet.cell(row, 4, payment['detail'])
        worksheet.cell(row, 5, payment['receipt'])
        worksheet.cell(row, 6, payment['tender'])
        row += 1
    worksheet.freeze_panes = 'A5'
