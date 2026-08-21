"""A case Iowa Courts listed with no date of birth says so on its own row.

The results page has a date of birth column and sometimes leaves it empty. A
row like that groups under DOB-UNKNOWN on the picking page, and ticking it puts
the case on the client's file on the strength of the name and nothing else --
which, on a common name, is somebody else's convictions and somebody else's
court debt.

The workbook said it in one place: BASIC INFO B6, which is one cell for the
whole file and reads as a blank field rather than as a warning. It is now on
the CASE DATA row too, where the cases actually get read. Iowa Legal Aid asked
for that on 2026-08-20.

Every name here is invented and every case number is 00000-shaped. The
repository is public.
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import crs
import tasks
from test_formula_grid import synthetic_cases   # noqa: E402

DATED = '1900-01-01 DOE, JANE'
UNDATED = tasks.DOB_UNKNOWN + ' DOE, JANE'

FIRST = '00000  FECR000000'
SECOND = '00000  FECR000001'


class TestWhichCasesAreUndated:
    def test_a_result_with_no_date_of_birth_groups_under_the_word(self):
        """The premise. dob_unknown_cases reads the key, so if _defendant_key
        ever spells this differently the reading has to change with it."""
        key = tasks._defendant_key({'dob': '', 'name': 'DOE, JANE'})
        assert key == UNDATED
        assert tasks._defendant_key(
            {'dob': ' ', 'name': 'DOE, JANE'}) == UNDATED

    def test_the_undated_group_is_marked(self):
        assert tasks.dob_unknown_cases(
            [UNDATED], {UNDATED: [FIRST, SECOND]}) == {FIRST, SECOND}

    def test_a_dated_group_is_not(self):
        assert tasks.dob_unknown_cases(
            [DATED], {DATED: [FIRST]}) == set()

    def test_only_the_undated_half_of_a_mixed_pick_is_marked(self):
        """The pick that makes this worth doing per row rather than per file:
        one client, one name, two groups, and only one of them checked against
        a date."""
        assert tasks.dob_unknown_cases(
            [DATED, UNDATED],
            {DATED: [FIRST], UNDATED: [SECOND]}) == {SECOND}

    def test_a_case_listed_under_both_keeps_its_date(self):
        """One result row missing the column is not a reason to caveat a case
        another row confirmed."""
        assert tasks.dob_unknown_cases(
            [DATED, UNDATED],
            {DATED: [FIRST], UNDATED: [FIRST, SECOND]}) == {SECOND}

    def test_the_order_the_groups_were_picked_in_does_not_matter(self):
        both = {DATED: [FIRST], UNDATED: [FIRST]}
        assert tasks.dob_unknown_cases([UNDATED, DATED], both) == set()
        assert tasks.dob_unknown_cases([DATED, UNDATED], both) == set()

    def test_a_key_with_no_cases_behind_it_is_harmless(self):
        assert tasks.dob_unknown_cases([UNDATED], {}) == set()

    def test_a_name_that_merely_starts_with_the_word_is_not_a_key(self):
        """The prefix is matched with its separating space on purpose. A
        surname is a surname."""
        assert tasks.dob_unknown_cases(
            ['1900-01-01 DOB-UNKNOWNS, JANE'],
            {'1900-01-01 DOB-UNKNOWNS, JANE': [FIRST]}) == set()


def _built(no_dob, count=2, **kw):
    from openpyxl import load_workbook
    path, _, _, _ = tasks.build_workbook(
        synthetic_cases(count), 'DOE, JANE', tasks.DOB_UNKNOWN, False,
        no_dob=no_dob, **kw)
    return load_workbook(path)['CASE DATA']


class TestTheNoteInTheWorkbook:
    def test_it_lands_in_the_notes_column(self):
        sheet = _built({FIRST})
        assert tasks.DOB_UNKNOWN_NOTE in sheet['V4'].value

    def test_only_the_row_it_belongs_to(self):
        sheet = _built({FIRST})
        assert tasks.DOB_UNKNOWN_NOTE not in (sheet['V5'].value or '')

    def test_nothing_is_written_without_it(self):
        # Column V is not empty on an ordinary row -- process_financials has
        # its own say about the fee columns -- so this is about the one
        # sentence, not about the cell.
        sheet = _built(set())
        assert 'DOB-Unknown' not in (sheet['V4'].value or '')

    def test_the_default_writes_nothing(self):
        """Every caller that has nothing to say about this passes nothing."""
        from openpyxl import load_workbook
        path, _, _, _ = tasks.build_workbook(
            synthetic_cases(1), 'DOE, JANE', '01/01/1900', False)
        sheet = load_workbook(path)['CASE DATA']
        assert 'DOB-Unknown' not in (sheet['V4'].value or '')

    def test_every_row_is_marked_when_every_row_is_undated(self):
        """Unlike the filed-under note, which goes quiet on a workbook where
        it would say the same thing on every row. Two spellings of a name is a
        fact about the workbook; no date of birth is a fact about the case,
        and a sheet where it is true throughout is a sheet where nothing was
        matched on a date at all."""
        sheet = _built({FIRST, SECOND})
        assert tasks.DOB_UNKNOWN_NOTE in sheet['V4'].value
        assert tasks.DOB_UNKNOWN_NOTE in sheet['V5'].value

    def test_it_sits_beside_the_filed_under_note(self):
        """Both are about which client this row is, and a case can need both:
        two spellings of the name and no date of birth under one of them."""
        sheet = _built({FIRST}, filed_as={FIRST: 'DOE, JAYNE'})
        assert 'Filed under DOE, JAYNE.' in sheet['V4'].value
        assert tasks.DOB_UNKNOWN_NOTE in sheet['V4'].value

    def test_it_does_not_displace_a_later_caveat(self):
        sheet = _built({FIRST}, count=1)
        crs.append_note(sheet, 4, 'A later caveat.')
        assert tasks.DOB_UNKNOWN_NOTE in sheet['V4'].value
        assert 'A later caveat.' in sheet['V4'].value


def test_it_survives_a_rebuild():
    """A retry rebuilds the workbook off the retry entry and never sees the
    search results again, which are the only thing that ever knew this."""
    entry = tasks._retry_entry('DOE, JANE', tasks.DOB_UNKNOWN, None,
                               [FIRST], [], [], no_dob={FIRST})
    assert list(entry['no_dob']) == [FIRST]


def test_an_older_retry_entry_has_no_such_key():
    """entry.get('no_dob') on the retry path, not entry['no_dob']: a job that
    started before this shipped is still in the dyno's memory."""
    entry = tasks._retry_entry('DOE, JANE', '01/01/1900', None, [FIRST],
                               [], [])
    assert entry['no_dob'] == []
