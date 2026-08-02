"""Column H, "Vehicular?", and the sheet that has been reading it empty.

LICENSE-REGIS asks one question per case: does unpaid debt on this case cost the
client a driver's licence, or only a vehicle registration. It answers it from
CASE DATA column H, and Napier has never written column H, so the answer has
been "Registration only" on every case of every workbook the app has produced.
That is not a formatting problem. A client is told their licence is safe when
it is not.

The first half of this file pins what Napier now writes there. The second half
pins why "YES" is the string, by reading the formula back out of the template
rather than trusting a comment: if CRS 3.6 renames the sentinel or moves the
column, the sheet goes quietly wrong again and these fail instead.
"""

import os
import re
import sys
from decimal import Decimal

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import crs

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')
LITE = os.path.join(ROOT, 'CRS Lite 3.5.5.xlsx')


# -- what Napier writes -----------------------------------------------------

VEHICULAR = [
    ('321J.2', 'OWI, the reason most of these clients lose a licence'),
    ('321.218', 'driving while barred'),
    ('321.174', 'no valid licence'),
    ('321.560', 'habitual offender'),
    ('321A.32', 'financial responsibility'),
    ('321.20A', 'a subsection, so the chapter test cannot key on length'),
    ('707.6A', 'homicide by vehicle, chapter 707 but revokes a licence'),
    ('707.6A(1)', 'the same with a subsection'),
]


@pytest.mark.parametrize('statute,why', VEHICULAR)
def test_a_vehicular_statute_says_yes(statute, why):
    assert crs.is_vehicular(statute) == "YES", why


NOT_VEHICULAR = [
    ('124.401', 'controlled substances'),
    ('714.1', 'theft'),
    ('708.2', 'assault'),
    ('707.6', 'not 707.6A. The vehicular section is the one with the A'),
    ('32.1', 'chapter 32, which is not 321 and must not match on a prefix'),
    ('3210.1', 'no such chapter, but a prefix test would call it vehicular'),
]


@pytest.mark.parametrize('statute,why', NOT_VEHICULAR)
def test_a_statute_that_is_not_vehicular_says_no(statute, why):
    assert crs.is_vehicular(statute) == "NO", why


NOTHING_TO_JUDGE = [
    ('', 'every count was dismissed, so column F is empty'),
    (None, 'no charge at all'),
    ('n/a', 'a civil case, which is what the civil branch writes into F'),
    ('  ', 'whitespace'),
]


@pytest.mark.parametrize('statutes,why', NOTHING_TO_JUDGE)
def test_it_says_nothing_when_there_is_nothing_to_say(statutes, why):
    """Blank, not "NO".

    A blank and a "NO" read the same to the sheet, so this costs nothing there
    and it stops the workbook asserting to a reader that Napier checked and
    found no vehicular charge when it had nothing to check.
    """
    assert crs.is_vehicular(statutes) is None, why


def test_one_vehicular_count_carries_the_case():
    """Column F is semicolon-joined and the licence follows any count in it.

    Plead to possession and OWI together and column G says GPL once for the
    whole case, but the licence consequence attaches to the OWI whichever count
    the ranking picked to speak for it.
    """
    assert crs.is_vehicular('124.401;321J.2') == "YES"
    assert crs.is_vehicular('321J.2;124.401') == "YES"


def test_several_counts_none_of_them_vehicular():
    assert crs.is_vehicular('124.401;714.1;708.2') == "NO"


def test_spacing_around_the_semicolon_does_not_matter():
    assert crs.is_vehicular('124.401; 321J.2') == "YES"


# -- and that it reaches the cell -------------------------------------------

# The tests above would all have passed against a version of Napier that never
# wrote column H, because they call is_vehicular directly. These are the ones
# that would not: they write a case into a worksheet the way a real run does.

def _case(statutes, disposition='GUILTY BY PLEA'):
    return {
        'id': '00000  FECR000000',
        'county': 'SYNTHETIC',
        'charges': [{
            'charge': statutes,
            'description': 'SYNTHETIC OFFENSE',
            'disposition': [disposition],
            'offenseDate': '01/01/1900',
            'dispositionDate': '02/02/1901',
        }],
        # No money on the case. The financial columns have their own tests.
        'financials': [],
        'summary_categories': [],
    }


def _written(case):
    sheet = load_workbook(FULL)['CASE DATA']
    crs.process_case(case, sheet, crs.FIRST_CASE_ROW)
    row = str(crs.FIRST_CASE_ROW)
    return sheet['F' + row].value, sheet['H' + row].value


def test_a_vehicular_case_reaches_column_h():
    assert _written(_case('321J.2')) == ('321J.2', 'YES')


def test_a_non_vehicular_case_reaches_column_h():
    assert _written(_case('124.401')) == ('124.401', 'NO')


def test_a_case_with_no_adjudicated_statute_leaves_column_h_alone():
    """Every count dismissed. Column F is empty and so is H."""
    assert _written(_case('', disposition='DISMISSED')) == ('', None)


def test_a_civil_case_leaves_column_h_alone():
    """The civil branch writes "n/a" into F, which is not a statute."""
    civil = {
        'id': '00000  SCSC000000',
        'county': 'SYNTHETIC',
        'charges': [],
        'summary_created_date': '01/01/1900',
        'summary_disposition_date': '02/02/1901',
        'summary_dispo_status': 'SYNTHETIC STATUS',
        'financials': [],
        'summary_categories': [],
    }
    assert _written(civil) == ('n/a', None)


# -- what the sheet does with it --------------------------------------------

def _license_regis_formula(path):
    """The first LICENSE-REGIS formula that reads CASE DATA column H."""
    sheet = load_workbook(path)['LICENSE-REGIS']
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "'CASE DATA'!H" in cell.value:
                return cell.value.replace('\n', '')
    return None


def test_the_full_template_asks_column_h_about_the_licence():
    formula = _license_regis_formula(FULL)
    assert formula is not None, "LICENSE-REGIS stopped reading column H"
    assert 'License & registration' in formula
    assert 'Registration only' in formula


def test_yes_is_the_string_the_sheet_is_testing_for():
    """The reason is_vehicular returns "YES" and not True or "Y".

    The sheet compares column H against a literal. Anything else Napier could
    write there is indistinguishable from an empty cell, which is exactly the
    state this whole change is fixing.
    """
    formula = _license_regis_formula(FULL)
    match = re.search(r"'CASE DATA'!H\d+\s*=\s*\"([^\"]*)\"", formula)
    assert match, formula
    assert match.group(1) == "YES"


def test_the_lite_template_asks_the_same_question():
    """Lite drops three sheets but keeps this one, and staff use it far more."""
    formula = _license_regis_formula(LITE)
    assert formula is not None
    match = re.search(r"'CASE DATA'!H\d+\s*=\s*\"([^\"]*)\"", formula)
    assert match and match.group(1) == "YES", formula


def test_the_sheet_only_asks_once_the_case_is_a_conviction_with_debt():
    """Context for the blank-when-unknown choice above.

    Column H is only reached when the case has debt and a conviction code, so a
    civil case or an all-dismissed case never gets asked. Writing "NO" into
    those rows would have been noise that changed no answer.
    """
    formula = _license_regis_formula(FULL)
    assert 'GTR' in formula and 'GPL' in formula and 'DEF' in formula


# -- charged under a city ordinance -----------------------------------------

# Nine of 90 captured cases are prosecuted under a local ordinance rather than
# the state section, and Iowa Courts print the ordinance citation in the
# adjudicated code. Every one of the nine is a motor vehicle matter and every one
# of them used to answer "NO".
#
# The citations below keep the real shapes and drop the real numbers.

ORDINANCE_WITH_THE_CHAPTER_IN_IT = [
    ('SY/32-321.285(d)(3)', 'a city adopting the state speeding section'),
    ('SY/9-7-321.285.D-C', 'the same city, a different ordinance numbering'),
    ('SY/321.218', 'no intervening ordinance number at all'),
]


@pytest.mark.parametrize('statute,why', ORDINANCE_WITH_THE_CHAPTER_IN_IT)
def test_an_ordinance_that_cites_the_state_section_is_read(statute, why):
    """No judgement call in this one. The chapter is written in the code.

    It only ever read "NO" because the test was anchored to the start of the
    string and an ordinance citation puts two letters and a slash in front.
    """
    assert crs.is_vehicular(statute) == "YES", why


BARE_ORDINANCE = [
    ('SY/62.01(120)-0198', 'a municipal code section, seat belts in that city'),
    ('SY/61.107', 'a municipal code section, parking'),
    ('SY/51.013-0063', 'a municipal code section, no registration'),
]


@pytest.mark.parametrize('statute,why', BARE_ORDINANCE)
def test_a_bare_ordinance_citation_is_not_answered_at_all(statute, why):
    """Not "NO". There is no chapter in it to disagree with.

    62.01 in one city's code has nothing to do with 62.01 in another's, and
    neither has anything to do with the state chapters. Napier saying "NO" here
    was the same confident answer it gives about a forgery statute, off a code
    it cannot read.
    """
    assert crs.is_vehicular(statute) is None, why


def test_an_ordinance_alongside_a_state_statute_still_answers():
    """The unreadable code only withholds the answer when nothing else gives one."""
    assert crs.is_vehicular('SY/62.01;321J.2') == "YES"


def test_a_state_statute_that_is_not_vehicular_is_still_a_confident_no():
    """The guard on the whole change.

    Chapter 714 is theft in the state code and stays a "NO". Withholding an
    answer is only for a citation Napier cannot read, not for every case that is
    not a traffic case, or LICENSE-REGIS loses the distinction entirely.
    """
    assert crs.is_vehicular('714.2(3);714.2(3)') == "NO"


def _ordinance_case(statute, costs):
    """A conviction under an unreadable ordinance, owing `costs`."""
    case = _case(statute, disposition='GUILTY')
    case['summary_categories'] = [
        {'label': label,
         'original': Decimal(costs) if label == 'COSTS' else Decimal('0'),
         'paid': Decimal('0'),
         'due': Decimal(costs) if label == 'COSTS' else Decimal('0')}
        for label in ('COSTS', 'FINE', 'SURCHARGE', 'RESTITUTION', 'OTHER')]
    # ICOS states the balance on every financials page it serves, and the
    # zero-balance case is the one that decides whether column V stays quiet.
    case['total_due'] = '$' + costs
    return case


def _note(case):
    sheet = load_workbook(FULL)['CASE DATA']
    crs.process_case(case, sheet, crs.FIRST_CASE_ROW)
    row = str(crs.FIRST_CASE_ROW)
    return sheet['H' + row].value, sheet['V' + row].value or ''


def test_a_blank_column_h_says_why_on_a_row_the_sheet_will_speak_about():
    """A blank is honest and invisible, which is the problem with leaving it.

    LICENSE-REGIS reads a blank exactly as it reads "NO" and prints
    "Registration only", so the row that gets that sentence printed about it
    carries the reason nobody could check it.
    """
    vehicular, note = _note(_ordinance_case('SY/62.01(120)-0198', '65.75'))
    assert vehicular is None
    assert 'SY/62.01(120)-0198' in note
    assert 'Registration only' in note


def test_a_case_that_owes_nothing_stays_quiet():
    """LICENSE-REGIS prints "Z - no debt" and never looks at column H.

    Seven of the nine captured ordinance cases are paid off. A caveat about a
    sentence the sheet is not going to print is the same noise column V was
    cleared of once already.
    """
    vehicular, note = _note(_ordinance_case('SY/62.01(120)-0198', '0'))
    assert vehicular is None
    assert 'Column H is blank' not in note, note


def test_a_dismissed_ordinance_case_stays_quiet():
    """No conviction, so the sheet says neither licence nor registration."""
    case = _ordinance_case('SY/62.01(120)-0198', '65.75')
    case['charges'][0]['disposition'] = ['DISMISSED']
    vehicular, note = _note(case)
    assert vehicular is None
    assert 'Column H is blank' not in note, note


def test_an_ordinance_napier_can_read_needs_no_caveat():
    """It answered the question, so there is nothing to warn about."""
    vehicular, note = _note(_ordinance_case('SY/32-321.285(d)(3)', '65.75'))
    assert vehicular == 'YES'
    assert 'Column H is blank' not in note, note


def test_the_caveat_does_not_displace_what_the_money_left_in_column_v():
    """Column V belongs to process_financials first. This joins, never replaces."""
    case = _ordinance_case('SY/62.01(120)-0198', '65.75')
    case['financials'] = [
        {'detail': 'SYNTHETIC UNCATEGORISED LINE', 'amount': Decimal('10.00'),
         'paid': None},
    ]
    _, note = _note(case)
    assert 'MISCELLANEOUS' in note, note
    assert 'SY/62.01(120)-0198' in note, note
