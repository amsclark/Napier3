"""The ranked action list, and the promise it makes to the workbook it sits in.

The list is only worth having if it agrees with the sheets behind it, so most
of what is pinned here is agreement rather than legal opinion: the twenty year
cut is the SOL sheet's 7300 days, the traffic test is the EXPUNGEMENT sheet's,
the conviction codes are LICENSE-REGIS's. Where those tests are reproduced in
Python, these tests hold the reproduction to the sheet.

Every case number is 00000 FECR000000 and every date is in the 1900s, because
this repository is public.
"""

import os
import sys
from datetime import date, timedelta
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import actions
import crs

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')
LITE = os.path.join(ROOT, 'CRS Lite 3.5.5.xlsx')

CLINIC = date(2026, 7, 31)
ALL_SHEETS = {'CASE DATA', 'SOL', 'EXPUNGEMENT & 910.7', 'POLK R&B APPEAL'}

# Twenty years and a day before the clinic, and a day short of it.
STALE = '01/01/1990'
FRESH = '01/01/2020'

FEES = 'JKLMNOPQRS'


def sheet_with(*rows):
    """A stand-in CASE DATA holding the columns the action list reads."""
    worksheet = Workbook().active
    for offset, values in enumerate(rows):
        row = crs.FIRST_CASE_ROW + offset
        cells = dict({'A': '00000  FECR000000', 'B': 'SYNTHETIC',
                      'D': FRESH, 'G': 'GTR'}, **values)
        for column, value in cells.items():
            worksheet[column + str(row)] = value
    return worksheet


def facts_for(**values):
    return actions.row_facts(sheet_with(values), crs.FIRST_CASE_ROW)


def cites(found):
    return [action[3] for _, action in found]


class TestRowFacts:
    def test_a_date_comes_back_as_a_date(self):
        # Napier writes dates as the strings ICOS uses. Excel coerces those in
        # arithmetic; Python does not, and a string against a date raises.
        assert facts_for(D='07/04/1976')['disposition_date'] == date(1976, 7, 4)

    def test_a_date_iowa_courts_left_blank_is_none_not_an_error(self):
        assert facts_for(D=None)['disposition_date'] is None

    def test_a_code_is_read_however_it_was_cased(self):
        assert facts_for(G=' gtr ')['code'] == 'GTR'

    def test_a_missing_code_is_empty_rather_than_none(self):
        assert facts_for(G=None)['code'] == ''

    def test_yes_columns_are_booleans(self):
        facts = facts_for(H='YES', I='YES')
        assert facts['vehicular'] and facts['supervised']

    def test_anything_but_yes_is_no(self):
        # Column I is blank on most rows and blank has to read as no, or every
        # case Napier could not answer turns into a 910.7 argument.
        facts = facts_for(H='', I=None)
        assert not facts['vehicular'] and not facts['supervised']

    def test_the_fee_columns_all_come_back(self):
        assert set(facts_for()['money']) == set(FEES)


class TestDecimalAndSum:
    def test_a_blank_fee_cell_is_zero(self):
        assert actions._decimal(None) == Decimal(0)

    def test_a_text_cell_is_zero_rather_than_an_exception(self):
        # The workbook is handed back to staff and comes back edited.
        assert actions._decimal('see notes') == Decimal(0)

    def test_a_float_does_not_pick_up_binary_error(self):
        assert actions._decimal(0.1) == Decimal('0.1')

    def test_the_fee_columns_add_up(self):
        facts = facts_for(J=100, K=25.5, R=10)
        assert actions._sum(facts, FEES) == Decimal('135.5')

    def test_910_7_cannot_reach_fines_surcharges_or_restitution(self):
        facts = facts_for(J=100, Q=1000, R=1000, S=1000)
        assert actions._sum(facts, actions.REMISSIBLE_COLUMNS) == Decimal(100)


class TestIsTraffic:
    def test_a_simple_misdemeanour_under_321_is_traffic(self):
        assert actions.is_traffic(facts_for(A='00000  SMSM000000', F='321.20'))

    def test_a_simple_misdemeanour_under_another_chapter_is_not(self):
        assert not actions.is_traffic(
            facts_for(A='00000  SMSM000000', F='123.46'))

    def test_a_scheduled_violation_is_traffic_whatever_the_statute(self):
        assert actions.is_traffic(facts_for(A='00000  STP0000000', F='123.46'))

    def test_a_non_scheduled_violation_is_traffic_too(self):
        assert actions.is_traffic(facts_for(A='00000  NTP0000000', F='123.46'))

    def test_a_felony_under_321_is_not_traffic(self):
        # Chapter 321 carries real felonies, and the sheet only calls the
        # simple misdemeanours traffic.
        assert not actions.is_traffic(facts_for(A='00000  FECR000000',
                                                F='321.261'))

    def test_a_case_with_no_number_does_not_raise(self):
        assert not actions.is_traffic(facts_for(A=None))


class TestTimeBarred:
    def test_a_day_past_twenty_years_is_barred(self):
        barred = CLINIC - timedelta(days=actions.SOL_DAYS + 1)
        assert actions.time_barred(
            facts_for(D=barred.strftime('%m/%d/%Y')), CLINIC)

    def test_exactly_twenty_years_is_not(self):
        # The SOL sheet tests D4+7300 < B3, so the day it lands is not barred,
        # and this has to break the same way round the sheet does.
        edge = CLINIC - timedelta(days=actions.SOL_DAYS)
        assert not actions.time_barred(
            facts_for(D=edge.strftime('%m/%d/%Y')), CLINIC)

    def test_an_undated_disposition_is_not_barred(self):
        assert not actions.time_barred(facts_for(D=None), CLINIC)


class TestMisdemeanourExpungement:
    def call(self, felonies=('707.2',), ineligible=('123.46',), **values):
        return actions.misdemeanour_expungement(
            facts_for(**dict({'D': STALE, 'G': 'GTR', 'F': '714.2'}, **values)),
            CLINIC, felonies, ineligible)

    def test_an_old_misdemeanour_conviction_qualifies(self):
        assert self.call()

    def test_a_dismissal_does_not(self):
        # A dismissal is 901C.2 and wipes the debt with it, which is a
        # different and better row.
        assert not self.call(G='DISM')

    def test_a_deferred_judgment_does_not(self):
        assert not self.call(G='DEF')

    def test_a_felony_on_the_sheet_s_list_does_not(self):
        assert not self.call(F='707.2')

    def test_one_felony_among_several_counts_disqualifies_the_case(self):
        # The helper columns split column F and the sheet matches every one of
        # them, so a clean count does not rescue the case.
        assert not self.call(F='714.2;707.2')

    def test_the_sheet_s_ineligible_misdemeanour_list_is_honoured(self):
        assert not self.call(F='123.46')

    def test_an_excluded_chapter_is_ruled_out_by_prefix(self):
        assert not self.call(F='719.1')

    def test_a_section_that_only_starts_the_same_is_still_matched(self):
        # 901A.2 is prefix-matched by the sheet, and so is anything under it.
        assert not self.call(F='901A.2')

    def test_eight_years_has_to_have_run(self):
        assert not self.call(D=FRESH)

    def test_a_traffic_case_is_not_this_argument(self):
        assert not self.call(A='00000  SMSM000000', F='321.20')

    def test_a_case_with_no_statute_is_left_alone(self):
        # Nothing to match against is not the same as matching nothing.
        assert not self.call(F=None)


class TestCaseActions:
    def find(self, **values):
        return actions.case_actions(facts_for(**values), CLINIC, ALL_SHEETS)

    def test_old_attorney_fees_are_a_614_1_6_objection(self):
        found = self.find(D=STALE, J=100, K=50)
        assert found[0][3] == 'Iowa Code 614.1(6)'
        assert found[0][1] == Decimal(150)

    def test_room_and_board_counts_toward_the_time_bar(self):
        assert self.find(D=STALE, L=200)[0][1] == Decimal(200)

    def test_fines_do_not_count_toward_the_time_bar(self):
        # The SOL sheet only puts attorney, collection and room and board in
        # its barred columns. A stale fine is a different argument.
        assert cites(
            [(None, action) for action in self.find(D=STALE, R=500)]) == []

    def test_an_old_case_with_nothing_owed_is_not_a_row(self):
        assert self.find(D=STALE) == []

    def test_the_sol_sheet_is_named_only_when_the_workbook_has_one(self):
        full = actions.case_actions(facts_for(D=STALE, J=100), CLINIC,
                                    ALL_SHEETS)
        lite = actions.case_actions(facts_for(D=STALE, J=100), CLINIC,
                                    {'CASE DATA'})
        assert 'SOL sheet' in full[0][4]
        assert 'SOL sheet' not in lite[0][4]

    def test_a_dismissal_discharges_the_whole_balance(self):
        found = self.find(G='DISM', J=100, R=400, S=500)
        assert found[0][3] == 'Iowa Code 901C.2'
        assert found[0][1] == Decimal(1000)

    def test_a_dismissed_traffic_case_is_not_901c_2(self):
        assert self.find(A='00000  STP0000000', G='DISM', J=100) == []

    def test_supervision_opens_910_7(self):
        found = self.find(I='YES', J=100, Q=900)
        assert found[0][3] == 'Iowa Code 910.7'
        # Surcharges are outside what 910.7 reaches.
        assert found[0][1] == Decimal(100)

    def test_supervision_with_nothing_remissible_is_not_a_row(self):
        assert self.find(I='YES', Q=900, R=900) == []

    def test_polk_room_and_board_is_a_356_7_challenge(self):
        found = self.find(B='Polk', L=300)
        assert 'Iowa Code 356.7' in cites([(None, a) for a in found])

    def test_room_and_board_outside_polk_is_not_the_polk_argument(self):
        # The appeal sheet is Polk's. The charge exists elsewhere and the list
        # says nothing about it rather than pointing at a sheet that will not
        # have the row.
        assert 'Iowa Code 356.7' not in cites(
            [(None, a) for a in self.find(B='SYNTHETIC', L=300)])

    def test_a_vehicular_conviction_with_debt_is_a_licence_row(self):
        found = self.find(H='YES', G='GTR', R=400)
        assert found[0][3] == 'Iowa Code 321.210A, 321.210B'
        assert found[0][1] == Decimal(400)

    def test_a_vehicular_case_that_was_dismissed_is_not(self):
        assert 'Iowa Code 321.210A, 321.210B' not in cites(
            [(None, a) for a in self.find(H='YES', G='DISM', R=400)])

    def test_a_vehicular_conviction_with_nothing_owed_is_not(self):
        assert self.find(H='YES', G='GTR') == []

    def test_a_juvenile_case_is_flagged(self):
        found = self.find(G='JUV', J=50)
        assert found[0][3] == 'Iowa Code 232.150'

    def test_one_case_can_carry_several_arguments(self):
        found = self.find(D=STALE, G='GTR', H='YES', I='YES', J=100, K=50)
        assert set(cites([(None, action) for action in found])) == {
            'Iowa Code 614.1(6)', 'Iowa Code 910.7',
            'Iowa Code 321.210A, 321.210B'}

    def test_the_registration_hold_is_not_a_row(self):
        # It fires on nearly every convicted case, so as rows it buried the
        # arguments worth an hour. It is one line in the header block.
        assert self.find(G='GTR', R=100) == []


class TestRegistrationHold:
    def test_convicted_cases_carrying_a_balance_are_counted(self):
        worksheet = sheet_with({'G': 'GTR', 'R': 100}, {'G': 'GPL', 'J': 50})
        assert actions.registration_hold(worksheet, 2) == (2, Decimal(150))

    def test_a_dismissal_is_not_counted(self):
        worksheet = sheet_with({'G': 'DISM', 'R': 100})
        assert actions.registration_hold(worksheet, 1) == (0, Decimal(0))

    def test_a_paid_off_conviction_is_not_counted(self):
        worksheet = sheet_with({'G': 'GTR'})
        assert actions.registration_hold(worksheet, 1) == (0, Decimal(0))

    def test_rows_past_the_ones_napier_wrote_are_not_read(self):
        worksheet = sheet_with({'G': 'GTR', 'R': 100}, {'G': 'GTR', 'R': 100})
        assert actions.registration_hold(worksheet, 1) == (1, Decimal(100))


class TestCollect:
    def test_the_biggest_number_is_first(self):
        worksheet = sheet_with(
            {'A': '00000  FECR000001', 'G': 'DISM', 'R': 100},
            {'A': '00000  FECR000002', 'G': 'DISM', 'R': 900},
        )
        found = actions.collect(worksheet, 2, CLINIC, ALL_SHEETS)
        assert [facts['id'] for facts, _ in found] == [
            '00000  FECR000002', '00000  FECR000001']

    def test_a_tie_on_money_is_broken_by_tier(self):
        # Both rows are worth the same; the time barred argument is the easier
        # one to make, so it reads first.
        worksheet = sheet_with({'D': STALE, 'G': 'DISM', 'J': 100})
        found = actions.collect(worksheet, 1, CLINIC, ALL_SHEETS)
        assert cites(found) == ['Iowa Code 614.1(6)', 'Iowa Code 901C.2']

    def test_the_same_workbook_always_comes_out_in_the_same_order(self):
        worksheet = sheet_with(
            {'A': '00000  FECR000002', 'G': 'DISM', 'R': 100},
            {'A': '00000  FECR000001', 'G': 'DISM', 'R': 100},
        )
        found = actions.collect(worksheet, 2, CLINIC, ALL_SHEETS)
        assert [facts['id'] for facts, _ in found] == [
            '00000  FECR000001', '00000  FECR000002']

    def test_an_empty_row_is_skipped_rather_than_read(self):
        worksheet = sheet_with({'G': 'DISM', 'R': 100}, {'A': None})
        assert len(actions.collect(worksheet, 2, CLINIC, ALL_SHEETS)) == 1


# -- the sheets themselves ---------------------------------------------------

def payment_row(detail, paid, when, receipt='000001', tender='CSH'):
    return {'detail': detail, 'amount': '100.00', 'paid': paid,
            'paidDate': when, 'receipt': receipt, 'tender': tender}


def synthetic_case(case_id='00000  FECR000000', rows=()):
    return {'id': case_id, 'financials': list(rows)}


def written_workbook(path, rows, cases=(), as_of=CLINIC):
    workbook = load_workbook(path)
    sheet = workbook['CASE DATA']
    for offset, values in enumerate(rows):
        row = crs.FIRST_CASE_ROW + offset
        cells = dict({'A': '00000  FECR000000', 'B': 'SYNTHETIC',
                      'D': FRESH, 'G': 'GTR'}, **values)
        for column, value in cells.items():
            sheet[column + str(row)] = value
    actions.build_action_sheet(workbook, list(cases), len(rows), as_of,
                               'SYNTHETIC CLIENT')
    return workbook


class TestClientPayments:
    def test_payments_are_pooled_across_every_case(self):
        history = actions.client_payments([
            synthetic_case('00000  FECR000001',
                           [payment_row('FINE', '10.00', '01/01/1900')]),
            synthetic_case('00000  FECR000002',
                           [payment_row('FINE', '30.00', '01/01/1901')]),
        ], CLINIC)
        assert history['count'] == 2
        assert history['cases'] == 2
        assert history['total'] == Decimal('40.00')

    def test_no_payments_anywhere_is_none(self):
        assert actions.client_payments([synthetic_case()], CLINIC) is None

    def test_the_recent_window_is_what_the_court_asks_about(self):
        history = actions.client_payments([
            synthetic_case(rows=[payment_row('FINE', '600.00', '01/01/1900'),
                                 payment_row('FINE', '120.00', '01/15/2026')]),
        ], CLINIC)
        assert history['recent'] == Decimal('120.00')
        assert history['recent_monthly'] == Decimal('10.00')

    def test_every_payment_carries_the_case_it_was_paid_on(self):
        history = actions.client_payments([
            synthetic_case('00000  FECR000007',
                           [payment_row('FINE', '10.00', '01/01/1900')]),
        ], CLINIC)
        assert history['payments'][0]['case'] == '00000  FECR000007'


class TestActionSheet:
    def test_the_action_list_is_the_first_sheet_staff_see(self):
        workbook = written_workbook(FULL, [{'G': 'DISM', 'R': 100}])
        assert workbook.sheetnames[0] == 'ACTION LIST'

    def test_the_client_and_the_clinic_date_are_at_the_top(self):
        workbook = written_workbook(FULL, [{'G': 'DISM', 'R': 100}])
        sheet = workbook['ACTION LIST']
        assert sheet['D1'].value == 'SYNTHETIC CLIENT'
        assert sheet['B2'].value == CLINIC
        assert sheet['B3'].value == 1

    def test_the_caveat_is_on_the_sheet_not_in_a_manual(self):
        workbook = written_workbook(FULL, [{'G': 'DISM', 'R': 100}])
        assert 'has been read by a lawyer' in workbook['ACTION LIST']['A8'].value

    def test_the_total_owed_adds_up_every_fee_column(self):
        workbook = written_workbook(
            FULL, [{'J': 100, 'R': 200}, {'S': 300}])
        assert workbook['ACTION LIST']['B4'].value == Decimal(600)

    def test_a_row_names_the_case_the_authority_and_the_money(self):
        workbook = written_workbook(FULL, [{'G': 'DISM', 'R': 100}])
        sheet = workbook['ACTION LIST']
        row = actions.FIRST_ACTION_ROW
        assert sheet.cell(row, 1).value == 1
        assert sheet.cell(row, 2).value == '00000  FECR000000'
        assert sheet.cell(row, 5).value == 'Iowa Code 901C.2'
        assert sheet.cell(row, 6).value == Decimal(100)

    def test_a_client_with_no_arguments_is_told_so_plainly(self):
        workbook = written_workbook(FULL, [{'G': 'GTR'}])
        cell = workbook['ACTION LIST'].cell(actions.FIRST_ACTION_ROW, 1)
        assert 'not the same as nothing being there' in cell.value

    def test_the_registration_hold_is_one_line(self):
        workbook = written_workbook(FULL, [{'G': 'GTR', 'R': 100}])
        assert '1 convicted case carrying $100.00' in \
            workbook['ACTION LIST']['B6'].value

    def test_no_hold_says_none_rather_than_nothing(self):
        workbook = written_workbook(FULL, [{'G': 'DISM', 'R': 100}])
        assert workbook['ACTION LIST']['B6'].value == 'none'

    def test_a_workbook_with_no_payments_says_which_kind_of_nothing(self):
        workbook = written_workbook(FULL, [{'G': 'DISM', 'R': 100}])
        assert workbook['ACTION LIST']['B5'].value == \
            'none in the ICOS itemization'

    def test_the_payment_line_reports_the_record(self):
        workbook = written_workbook(
            FULL, [{'G': 'DISM', 'R': 100}],
            [synthetic_case(rows=[payment_row('FINE', '25.00', '01/02/1900')])])
        line = workbook['ACTION LIST']['B5'].value
        assert '$25.00 across 1 payment on 1 case' in line
        assert '01/02/1900 to 01/02/1900' in line

    def test_the_helper_columns_are_left_alone(self):
        # W through AH carry the array formulas the eligibility tests match
        # against, under a header that says DO NOT EDIT THESE.
        def formula(sheet, column):
            value = sheet.cell(crs.FIRST_CASE_ROW, column).value
            return getattr(value, 'text', value)

        before = load_workbook(FULL)['CASE DATA']
        after = written_workbook(FULL, [{'G': 'DISM', 'R': 100}])['CASE DATA']
        for column in range(23, 35):
            assert formula(after, column) == formula(before, column)
            assert formula(after, column)

    def test_the_lite_workbook_gets_a_list_too(self):
        workbook = written_workbook(LITE, [{'D': STALE, 'J': 100}])
        assert workbook.sheetnames[0] == 'ACTION LIST'

    def test_the_lite_list_does_not_point_at_a_sheet_it_does_not_have(self):
        # SOL, BANKRUPTCY and EXEMPTIONS are not in the Lite file.
        workbook = written_workbook(LITE, [{'D': STALE, 'J': 100}])
        why = workbook['ACTION LIST'].cell(actions.FIRST_ACTION_ROW, 7).value
        assert 'SOL sheet' not in why
        assert 'Iowa Code 614.1(6)' == \
            workbook['ACTION LIST'].cell(actions.FIRST_ACTION_ROW, 5).value

    def test_the_sheet_can_be_saved_and_read_back(self, tmp_path):
        # openpyxl writes Decimals, and a workbook that will not reopen is not
        # a work product.
        workbook = written_workbook(FULL, [{'G': 'DISM', 'R': 100}])
        path = str(tmp_path / 'synthetic.xlsx')
        workbook.save(path)
        assert load_workbook(path)['ACTION LIST'].cell(
            actions.FIRST_ACTION_ROW, 5).value == 'Iowa Code 901C.2'


class TestPaymentSheet:
    def workbook(self, rows):
        return written_workbook(FULL, [{'G': 'GTR', 'R': 100}],
                                [synthetic_case(rows=rows)])

    def test_newest_first_because_that_is_the_question(self):
        sheet = self.workbook([
            payment_row('FINE', '10.00', '01/01/1900'),
            payment_row('FINE', '20.00', '01/01/1901'),
        ])['PAYMENTS']
        assert sheet.cell(5, 2).value == date(1901, 1, 1)
        assert sheet.cell(6, 2).value == date(1900, 1, 1)

    def test_the_receipt_and_the_tender_are_carried_through(self):
        sheet = self.workbook([
            payment_row('COURT COSTS', '10.00', '01/01/1900',
                        receipt='000123', tender='CHK'),
        ])['PAYMENTS']
        assert sheet.cell(5, 1).value == '00000  FECR000000'
        assert sheet.cell(5, 3).value == Decimal('10.00')
        assert sheet.cell(5, 4).value == 'COURT COSTS'
        assert sheet.cell(5, 5).value == '000123'
        assert sheet.cell(5, 6).value == 'CHK'

    def test_the_total_is_on_the_sheet(self):
        sheet = self.workbook([
            payment_row('FINE', '10.00', '01/01/1900'),
            payment_row('FINE', '20.00', '01/01/1901'),
        ])['PAYMENTS']
        assert sheet['B2'].value == Decimal('30.00')

    def test_an_empty_record_says_what_it_does_and_does_not_mean(self):
        sheet = written_workbook(FULL, [{'G': 'GTR', 'R': 100}])['PAYMENTS']
        assert 'not proof nothing was paid' in sheet['A3'].value


def test_the_code_lists_come_out_of_the_workbook():
    # Keeping a second copy in Python would drift the first time a clinic
    # added a section to CODE SECTIONS.
    felonies, ineligible = actions.code_lists(load_workbook(FULL))
    assert len(felonies) > 100
    assert ineligible
    assert all(isinstance(section, str) for section in felonies + ineligible)


@pytest.mark.parametrize('statute', ['321.20', ' 321.20 ', '714.2;321.20'])
def test_statutes_are_split_the_way_the_helper_columns_split_them(statute):
    assert '321.20' in actions.statutes(facts_for(F=statute))
