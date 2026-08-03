"""Every sheet has to reach the row holding the case it is talking about.

The licence sheet shipped reading CASE DATA in scrambled order, every case row
once but not in sequence, so a short run scattered the client's cases down the
sheet behind rows that read as blank. It was found by checking one sheet
against one other. This is that check generalised: every reference from any
sheet to any sheet, in both templates, has to land on the row holding the same
case the referring row holds.

The sheets disagree about where case n lives. CASE DATA puts it at row 4, the
expungement sheet at row 3, the licence and Polk sheets at row 2. So the
expungement sheet's row 3 reaching into SOL has to reach SOL row 4, and a
formula that reaches SOL row 3 is reading the case above.

An absolute row is exempt. COUNTIF('CASE DATA'!$D$4:$D$200,"") counts pending
charges over the whole list and belongs to no single case, which is exactly
what the dollar sign says.

This is a structural check on the shipped files. It needs no fixture, holds
whatever a client's cases happen to be, and is here because the defect it
guards against was invisible in every arithmetic total on the sheet.
"""

import os
import re

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATES = (os.path.join(ROOT, 'CRS 3.5.5.xlsx'),
             os.path.join(ROOT, 'CRS Lite 3.5.5.xlsx'))

# Where case n lands on each sheet: row = FIRST[sheet] + n. A sheet not listed
# here holds no per-case rows, and the two that do not appear in Lite are
# skipped there rather than demanded.
FIRST = {'CASE DATA': 4, 'EXPUNGEMENT & 910.7': 3, 'BANKRUPTCY': 4,
         'EXEMPTIONS': 4, 'SOL': 4, 'LICENSE-REGIS': 2, 'POLK R&B APPEAL': 2}

# 'SOL'!C4 or SOL!C4, with either part of the address possibly absolute. A
# reference with no sheet name is same-sheet and is not what this checks.
QUALIFIED = re.compile(r"(?:'([^']+)'|([A-Za-z][A-Za-z0-9 &]*))!"
                       r"(\$?)([A-Z]{1,3})(\$?)(\d+)")

# Beyond any case list. Sheets carry styling far below their formulas and
# walking to the bottom of the spreadsheet is the slow way to learn nothing.
ROW_CAP = 400


def text(cell):
    value = cell.value
    if isinstance(value, ArrayFormula):
        return value.text
    return value if isinstance(value, str) and value.startswith('=') else None


def misaligned(template):
    """Every reference that lands on the wrong case, and how many were read."""
    workbook = load_workbook(template)
    wrong = []
    checked = 0
    for sheet in workbook.worksheets:
        if sheet.title not in FIRST:
            continue
        base = FIRST[sheet.title]
        for row in sheet.iter_rows(min_row=1,
                                   max_row=min(sheet.max_row, ROW_CAP)):
            for cell in row:
                formula = text(cell)
                if not formula:
                    continue
                index = cell.row - base
                if index < 0:
                    continue          # a header or a totals row, not a case
                for quoted, bare, _, column, absolute, number in \
                        QUALIFIED.findall(formula):
                    target = quoted or bare
                    if target not in FIRST or absolute:
                        continue
                    checked += 1
                    want = FIRST[target] + index
                    if int(number) != want:
                        wrong.append(
                            '%s %s%d (case %d) reads %s %s%s, wants %s%d'
                            % (sheet.title, cell.column_letter, cell.row,
                               index, target, column, number, column, want))
    return wrong, checked


def test_every_reference_between_sheets_lands_on_its_own_case():
    for template in TEMPLATES:
        wrong, checked = misaligned(template)
        # A regex that stopped matching would report a clean sheet, so the
        # count is asserted before the result is trusted.
        assert checked > 20000, (os.path.basename(template), checked)
        assert not wrong, '%s: %d misaligned of %d, first few %s' % (
            os.path.basename(template), len(wrong), checked, wrong[:6])


def test_the_check_can_still_see_a_scrambled_sheet():
    """The guard above says nothing useful if it cannot fail. This shifts one
    sheet's idea of where a case lives and requires the check to notice."""
    original = FIRST['LICENSE-REGIS']
    try:
        FIRST['LICENSE-REGIS'] = original + 1
        wrong, _ = misaligned(TEMPLATES[0])
        assert len(wrong) > 200, len(wrong)
    finally:
        FIRST['LICENSE-REGIS'] = original
