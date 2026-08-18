"""A case list longer than the grid the templates were filled down to.

CASE DATA takes one row per case and nothing stops it. The sheets that read
CASE DATA stop, each at a different row, because each was dragged down by hand
to whatever looked like enough: SOL at 150, expungement and exemptions at 200,
bankruptcy at 201, licence and Polk room and board at 300. The totals at the
top of SOL, bankruptcy and exemptions are shorter still, =SUM(C4:C100).

Past those the workbook says nothing. The case sits on CASE DATA with its money
in columns J to S and it is simply absent from the analysis: not time barred,
not "NO ARGUMENT", not counted as a pending charge under 901C.2.

Measured on the 210 captured cases, built into the real workbook and computed
by LibreOffice. Before, 63 of them had no row at all on the SOL sheet and its
NO ARGUMENT total read $6,727.91. After, every case has a row and the total
reads $7,539.73. The same corpus cut to 40 cases comes out cell for cell
identical either way.

A staffer has already pulled 70 cases in one run in production, and the totals
give way at 98.

The cases here are synthetic. This repo is public and a real charges page is
one person's unredacted criminal record.
"""

import os
import re
import sys
from decimal import Decimal

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import case_parser
import grid
import tasks

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')
LITE = os.path.join(ROOT, 'CRS Lite 3.5.5.xlsx')

# The shallowest per-row grid in either template, so a list longer than this
# has cases the SOL sheet cannot see.
SOL_LAST_ROW = 150

# Where the expungement sheet's own per-row formulas stop, and the columns it
# keeps as ctrl-shift-enter array formulas rather than plain ones.
EXPUNGEMENT_LAST_ROW = 200
ARRAY_COLUMNS = {'CRS 3.5.5.xlsx': ('F', 'M', 'N'),
                 'CRS Lite 3.5.5.xlsx': ('F', 'L', 'M')}

CASE_REF = re.compile(r"'CASE DATA'!(\$?)[A-Z]{1,3}(\$?)(\d+)")


def _cells(*cells):
    return '<tr>%s</tr>' % ''.join(
        '<td><font size="2">%s</font></td>' % cell for cell in cells)


def charges_page():
    """One count, convicted, so the row carries a code and a date."""
    return ''.join([
        '<html><body><table>',
        _cells('Count 01', 'Original Charge'),
        _cells('Charge:', '321.285', 'Description:', 'SYNTHETIC SPEEDING'),
        _cells('Offense Date:', '01/01/1900', 'Arrest Date:', ''),
        _cells('Adjudication'),
        _cells('Charge:', '321.285', 'Description:', 'SYNTHETIC SPEEDING'),
        _cells('Adjudication:', 'GUILTY', 'Adjudication Date:', '02/02/1901'),
        '</table></body></html>']).encode('utf-8')


def synthetic_cases(count):
    """A case list of a given length, every case owing the same small sum."""
    page = charges_page()
    cases = []
    for n in range(count):
        case = {'id': '00000  FECR%06d' % n, 'county': 'SYNTHETIC',
                'financials': [], 'sentences': [],
                'summary_created_date': '01/01/1900',
                'summary_disposition_date': '', 'summary_dispo_status': ''}
        case_parser.parse_case_charges(page, case)
        case['summary_categories'] = [
            {'label': label, 'original': Decimal('10.00' if label == 'FINE' else '0'),
             'paid': Decimal('0'),
             'due': Decimal('10.00' if label == 'FINE' else '0')}
            for label in ('COSTS', 'FINE', 'SURCHARGE', 'RESTITUTION', 'OTHER')]
        case['total_due'] = '$10.00'
        cases.append(case)
    return cases


def extended(count, template=FULL):
    """A template with the case rows filled in and the grid widened to match."""
    workbook = load_workbook(template)
    sheet = workbook['CASE DATA']
    for n in range(count):
        sheet['A%d' % (4 + n)] = '00000  FECR%06d' % n
    grown = grid.extend_formula_grid(workbook, count)
    return workbook, grown


def formulas(sheet, limit):
    return {cell.coordinate: cell.value
            for row in sheet.iter_rows(min_row=1, max_row=limit)
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith('=')}


# -- the end of the case list reaches every sheet ----------------------------

def test_a_case_past_the_sol_grid_gets_a_row_on_it():
    count = SOL_LAST_ROW + 20
    workbook, grown = extended(count)
    last = 3 + count
    assert grown, 'nothing grew, so nothing below is being tested'
    formula = workbook['SOL']['A%d' % last].value
    assert formula, 'the last case has no row on the SOL sheet'
    assert "'CASE DATA'!A%d" % last in formula, formula


def test_every_case_row_is_read_exactly_once_afterwards():
    """Filling down must not repeat a case row or drop one. LICENSE-REGIS
    shipped reading them shuffled, case row 4 sitting at sheet row 57, and the
    thing that made that survivable is what this guards."""
    count = 320
    workbook, _ = extended(count)
    last = 3 + count
    for name in ('SOL', 'LICENSE-REGIS', 'BANKRUPTCY', 'EXEMPTIONS'):
        sheet = workbook[name]
        seen = []
        for row in sheet.iter_rows(min_row=1, max_row=last + 4):
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith('=')):
                    continue
                if cell.column_letter != 'A':
                    continue
                # Once per cell. The formula is
                # =IF('CASE DATA'!A4="","",'CASE DATA'!A4) and names its case
                # row twice, which is one row read once.
                seen.extend({int(match.group(3))
                             for match in CASE_REF.finditer(value)
                             if not match.group(2)})
        wanted = set(range(4, last + 1))
        assert wanted <= set(seen), (
            '%s misses case rows %s' % (name, sorted(wanted - set(seen))[:10]))
        repeats = [r for r in wanted if seen.count(r) > 1]
        assert not repeats, '%s reads case rows twice: %s' % (name, repeats[:10])


def test_the_totals_cover_the_whole_case_list():
    """SOL, bankruptcy and exemptions all total =SUM(C4:C100), which gives way
    at 98 cases while their own rows run to 150 and beyond."""
    count = 120
    workbook, _ = extended(count)
    for name in ('SOL', 'BANKRUPTCY', 'EXEMPTIONS'):
        total = workbook[name]['C1'].value
        end = int(re.search(r'C4:C(\d+)', total).group(1))
        assert end >= 3 + count, '%s totals stop at row %d' % (name, end)


def test_the_pending_charge_count_covers_the_whole_case_list():
    """The expungement sheet finds pending charges with
    COUNTIF('CASE DATA'!$D$4:$D$200,""), and a blank D past row 200 is a
    pending charge it cannot see. That count is the 901C.2 blocker."""
    count = 260
    workbook, _ = extended(count)
    sheet = workbook['EXPUNGEMENT & 910.7']
    found = False
    for row in sheet.iter_rows(min_row=1, max_row=3 + count + 4):
        for cell in row:
            value = cell.value
            if not (isinstance(value, str) and 'COUNTIF' in value):
                continue
            for end in re.findall(r"'CASE DATA'!\$[A-Z]+\$\d+:\$[A-Z]+\$(\d+)",
                                  value):
                found = True
                assert int(end) >= 3 + count, (
                    '%s counts only to row %s' % (cell.coordinate, end))
    assert found, 'no COUNTIF over CASE DATA was checked'


# -- and a workbook that fits is left exactly as it was ----------------------

def test_a_short_case_list_changes_no_formula():
    """Most clinics are well inside the grid, and those workbooks have to come
    out of this unchanged."""
    plain = load_workbook(FULL)
    workbook, grown = extended(40)
    assert not any(grown.values()), 'a 40 case list needs no new rows'
    for name in ('SOL', 'EXPUNGEMENT & 910.7', 'BANKRUPTCY', 'EXEMPTIONS',
                 'LICENSE-REGIS', 'POLK R&B APPEAL'):
        before = formulas(plain[name], 320)
        after = formulas(workbook[name], 320)
        assert before, 'no formulas read off %s, so nothing was compared' % name
        moved = {k: (before[k], after.get(k))
                 for k in before if before[k] != after.get(k)}
        # The totals are allowed to reach the sheet's own last row; nothing
        # else may move, and no total may reach past what the sheet computes.
        for coordinate, (was, now) in moved.items():
            assert coordinate in ('C1', 'D1', 'E1', 'F1'), \
                '%s!%s changed: %r -> %r' % (name, coordinate, was, now)


def test_the_grand_total_is_not_stretched_down_the_sheet():
    """SOL's F1 is =SUM(C1:E1), three cells of the totals row itself. Widening
    it to E150 adds every per-row figure to the totals that already hold them
    and doubles the sheet's headline number."""
    workbook, _ = extended(210)
    assert workbook['SOL']['F1'].value == '=SUM(C1:E1)'
    assert workbook['BANKRUPTCY']['G1'].value == '=SUM(C1:F1)'
    assert workbook['EXEMPTIONS']['E1'].value == '=SUM(C1:D1)'


def test_nothing_happens_to_an_empty_case_list():
    workbook, grown = extended(0)
    assert grown == {}
    assert workbook['SOL']['A%d' % (SOL_LAST_ROW + 1)].value is None


# -- the rows that get added look like the ones above them -------------------

def test_the_new_rows_keep_the_formats_of_the_row_they_came_from():
    """A dollar figure printed as a bare number reads as a different number."""
    workbook, _ = extended(SOL_LAST_ROW + 20)
    sheet = workbook['SOL']
    for column in ('C', 'D', 'E'):
        source = sheet['%s%d' % (column, SOL_LAST_ROW)].number_format
        added = sheet['%s%d' % (column, SOL_LAST_ROW + 5)].number_format
        assert added == source, column


def test_the_lite_template_grows_too():
    """Lite has no SOL sheet but its licence and expungement sheets stop too."""
    workbook, grown = extended(260, template=LITE)
    assert grown, 'the lite template did not grow'
    last = 3 + 260
    formula = workbook['EXPUNGEMENT & 910.7']['A%d' % (last - 1)].value
    assert formula and "'CASE DATA'!A%d" % last in formula, formula


def test_it_reports_which_sheets_it_had_to_touch():
    _, grown = extended(SOL_LAST_ROW + 20)
    assert 'SOL' in grown and grown['SOL'] > 0, grown


# -- the columns that were entered with ctrl-shift-enter ---------------------

def array_columns(template=FULL):
    return ARRAY_COLUMNS[os.path.basename(template)]


def test_an_array_formula_column_reaches_the_new_rows_too():
    """openpyxl hands a ctrl-shift-enter cell back as an object rather than a
    string, so a formula read as a string is silently not one of these. Three
    columns of the expungement sheet are that kind, and one of them is whether
    there is a subsequent conviction, which is what decides a juvenile record
    under 232.150. Dropping them leaves a row that is present and blank, which
    reads as nothing to report."""
    count = EXPUNGEMENT_LAST_ROW + 20
    workbook, grown = extended(count)
    sheet = workbook['EXPUNGEMENT & 910.7']
    assert grown.get('EXPUNGEMENT & 910.7'), 'the sheet did not grow'
    for column in array_columns():
        source = sheet['%s%d' % (column, EXPUNGEMENT_LAST_ROW)]
        assert isinstance(source.value, ArrayFormula), \
            '%s is no longer an array formula, so nothing is being tested' % column
        for row in (EXPUNGEMENT_LAST_ROW + 1, EXPUNGEMENT_LAST_ROW + 10):
            assert sheet['%s%d' % (column, row)].value is not None, \
                '%s%d is blank' % (column, row)


def test_a_copied_array_formula_is_still_an_array_formula():
    """Written back as a plain string it loses the marking, and every one of
    these is a SUM(COUNTIF(range, range)) that stops summing without it and
    answers about a single cell instead."""
    workbook, _ = extended(EXPUNGEMENT_LAST_ROW + 20)
    sheet = workbook['EXPUNGEMENT & 910.7']
    for column in array_columns():
        cell = sheet['%s%d' % (column, EXPUNGEMENT_LAST_ROW + 5)]
        assert isinstance(cell.value, ArrayFormula), (column, cell.value)


def test_a_copied_array_formula_is_entered_over_its_own_cell():
    """The ref travels with the formula. Left pointing at the row it came from,
    the copy is entered over a cell it is not in."""
    workbook, _ = extended(EXPUNGEMENT_LAST_ROW + 20)
    sheet = workbook['EXPUNGEMENT & 910.7']
    for column in array_columns():
        for row in (EXPUNGEMENT_LAST_ROW + 1, EXPUNGEMENT_LAST_ROW + 12):
            cell = sheet['%s%d' % (column, row)]
            assert isinstance(cell.value, ArrayFormula), (column, row,
                                                          cell.value)
            assert cell.value.ref == '%s%d' % (column, row), \
                (column, row, cell.value.ref)


def test_a_copied_array_formula_reads_its_own_case():
    """Filling down has to move the case reference with the row, or every new
    row answers about the last case that fitted."""
    count = EXPUNGEMENT_LAST_ROW + 20
    workbook, _ = extended(count)
    sheet = workbook['EXPUNGEMENT & 910.7']
    row = EXPUNGEMENT_LAST_ROW + 10
    cell = sheet['M%d' % row]
    assert isinstance(cell.value, ArrayFormula), cell.value
    assert "'CASE DATA'!G%d" % (row + 1) in cell.value.text, cell.value.text


def test_an_array_formula_counting_every_case_is_widened_too():
    """SUBSEQUENT CONVICTION counts convictions over 'CASE DATA'!$G$4:$G$200.
    On a longer list that answers "no subsequent conviction" from a partial
    reading of the case list, which is the direction that says a juvenile
    record is clear to expunge when it is not."""
    count = 260
    workbook, _ = extended(count)
    sheet = workbook['EXPUNGEMENT & 910.7']
    last_case_row = 3 + count
    for row in (3, EXPUNGEMENT_LAST_ROW, EXPUNGEMENT_LAST_ROW + 20):
        cell = sheet['F%d' % row]
        assert isinstance(cell.value, ArrayFormula), (row, cell.value)
        end = int(re.search(r"\$G\$4:\$G\$(\d+)", cell.value.text).group(1))
        assert end >= last_case_row, 'row %d counts only to %d' % (row, end)


def test_the_lite_template_keeps_its_array_formulas_too():
    workbook, _ = extended(EXPUNGEMENT_LAST_ROW + 20, template=LITE)
    sheet = workbook['EXPUNGEMENT & 910.7']
    for column in array_columns(LITE):
        cell = sheet['%s%d' % (column, EXPUNGEMENT_LAST_ROW + 5)]
        assert isinstance(cell.value, ArrayFormula), (column, cell.value)


def test_a_short_case_list_leaves_the_array_formulas_alone():
    plain = load_workbook(FULL)['EXPUNGEMENT & 910.7']
    workbook, _ = extended(40)
    sheet = workbook['EXPUNGEMENT & 910.7']
    for column in array_columns():
        for row in (3, 100, EXPUNGEMENT_LAST_ROW):
            was = plain['%s%d' % (column, row)].value
            now = sheet['%s%d' % (column, row)].value
            assert now.text == was.text, (column, row)
            assert now.ref == was.ref, (column, row)
        assert sheet['%s%d' % (column, EXPUNGEMENT_LAST_ROW + 1)].value is None


# -- through the real build ---------------------------------------------------

def test_a_built_workbook_covers_all_of_its_own_cases():
    """The whole point, through the path a clinic actually takes."""
    count = SOL_LAST_ROW + 10
    path, _, _, _ = tasks.build_workbook(
        synthetic_cases(count), 'TEST CLIENT', '01/01/1980', False)
    try:
        workbook = load_workbook(path)
        last = 3 + count
        assert workbook['CASE DATA']['A%d' % last].value, \
            'the case list did not reach that row, so nothing is proved'
        formula = workbook['SOL']['A%d' % last].value
        assert formula, 'the last case has no row on the SOL sheet'
        assert "'CASE DATA'!A%d" % last in formula, formula
        end = int(re.search(r'C4:C(\d+)', workbook['SOL']['C1'].value).group(1))
        assert end >= last, 'the SOL totals stop at row %d of %d' % (end, last)
    finally:
        os.remove(path)
