"""A case Iowa Courts has closed, on a row three sheets read as an open charge.

Napier's production alert fired on a real run with two of these:

    unrecognised disposition on an ICOS case (disposition CLOSED)

The status ICOS prints for a case as a whole is its own vocabulary, and
case_level_code deliberately translates only the one wording that overlaps the
per-count vocabulary. Everything else returns None, which leaves column G empty
and raises the alert rather than guessing at a conviction code. That part is
working as intended and the question of what CLOSED deserves belongs to Iowa
Legal Aid.

What was not working is what the row then said about itself. Both the empty
column G and an unreadable per-count adjudication went out under one note, and
that note describes the per-count case: it tells the attorney the row is coded
OTH and then explains how each sheet reads OTH. On the case-level row none of
that is true. Column G is empty, and

    BANKRUPTCY B4  =IF('CASE DATA'!A4<>"",IF('CASE DATA'!G4=0, "open charge", 'CASE DATA'!G4),"")

is also EXEMPTIONS B4 and SOL B4, so all three answer "open charge". The one
safety net on the row was describing a different row.

Of 300 captured cases 2 take this path, one carrying $197.43 of debt. Every case
number here is synthetic. The repo is public.
"""

import os
import sys
from decimal import Decimal

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import alerts
import case_parser
import crs
import tasks
from test_multi_count import charges_page

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')

CLOSED = 'CLOSED'
TRANSFERRED = 'TRANSFERRED'


def _case(counts, status, costs='197.43', case_id='00000  FECR000000'):
    """A case the shape build_workbook takes, off parsed ICOS pages."""
    case = {'id': case_id, 'county': 'SYNTHETIC',
            'financials': [], 'sentences': [],
            'summary_created_date': '01/01/1900',
            'summary_disposition_date': '02/02/1901',
            'summary_dispo_status': status}
    case_parser.parse_case_charges(charges_page(counts), case)
    case['summary_categories'] = [
        {'label': label,
         'original': Decimal(costs) if label == 'COSTS' else Decimal('0'),
         'paid': Decimal('0'),
         'due': Decimal(costs) if label == 'COSTS' else Decimal('0')}
        for label in ('COSTS', 'FINE', 'SURCHARGE', 'RESTITUTION', 'OTHER')]
    case['total_due'] = '$' + costs
    return case


def _row(counts, status, costs='197.43'):
    """A row built the way process_case builds one, off parsed ICOS pages."""
    sheet = load_workbook(FULL)['CASE DATA']
    crs.process_case(_case(counts, status, costs), sheet, crs.FIRST_CASE_ROW)
    row = str(crs.FIRST_CASE_ROW)
    return {column: sheet[column + row].value
            for column in ('D', 'G', 'V')}


UNADJUDICATED = [('714.2(3)', 'SYNTHETIC THEFT', None, '')]
UNREADABLE = [('714.2(3)', 'SYNTHETIC THEFT',
               'SYNTHETIC WORDING NOBODY HAS SEEN', '02/02/1901')]


# -- the shape the alert is firing on ----------------------------------------

@pytest.mark.parametrize('status', [CLOSED, TRANSFERRED])
def test_an_untranslated_case_status_leaves_column_g_empty(status):
    """The guard on the deliberate half. Guessing a code here is the error this
    is avoiding, so the fix must not start filling the cell in."""
    assert not _row(UNADJUDICATED, status)['G']


@pytest.mark.parametrize('status', [CLOSED, TRANSFERRED])
def test_the_row_says_the_three_sheets_will_call_it_an_open_charge(status):
    """What the attorney has to know to use the row: it is about to appear on
    BANKRUPTCY, EXEMPTIONS and SOL as a charge still pending."""
    note = _row(UNADJUDICATED, status)['V'] or ''
    assert 'open charge' in note, note
    for sheet in ('BANKRUPTCY', 'EXEMPTIONS', 'SOL'):
        assert sheet in note, note


def test_the_row_names_the_status_icos_printed():
    """Left out, the note sends someone back to ICOS without saying what to
    look for, and the wording is the whole reason the row is uncoded."""
    assert CLOSED in (_row(UNADJUDICATED, CLOSED)['V'] or '')


def test_the_row_does_not_claim_to_be_coded_oth():
    """The defect. Column G is empty, so every sentence the OTH note spends on
    how the sheets read OTH is about some other row."""
    note = _row(UNADJUDICATED, CLOSED)['V'] or ''
    assert 'OTH' not in note, note


def test_a_paid_off_case_is_told_as_well():
    """One of the two captured cases owes nothing. The mislabelling is on the
    three sheets that sort a case, not only on the ones that sort its money, so
    this note is not conditional on the row owing anything."""
    note = _row(UNADJUDICATED, TRANSFERRED, costs='0')['V'] or ''
    assert 'open charge' in note, note


# -- and the row the old note was written for still gets it ------------------

def test_a_count_napier_could_not_read_is_still_coded_oth():
    cells = _row(UNREADABLE, '')
    assert cells['G'] == 'OTH'
    assert 'coded OTH' in (cells['V'] or ''), cells['V']


def test_that_row_is_not_told_it_will_read_as_an_open_charge():
    """Because it will not. It has a code, and the sheets read the code."""
    assert 'open charge' not in (_row(UNREADABLE, '')['V'] or '')


def test_the_case_status_is_not_reached_when_a_count_was_adjudicated():
    """ICOS prints a case-level status on cases that do have adjudications. It
    is only consulted when no count has one, so a readable conviction is not
    displaced by a status nobody translates."""
    cells = _row([('714.2(3)', 'SYNTHETIC THEFT', 'GUILTY', '02/02/1901')],
                 CLOSED)
    assert cells['G'] == 'GTR'
    assert not cells['V'] or 'open charge' not in cells['V']


# -- the vocabulary the refusal is built on ----------------------------------

# Every case level status the 300 captured pages carry. The four at the end
# appeared only after the corpus passed 90, which is why case_level_code will
# not guess: the list is still growing.
CASE_LEVEL_STATUSES = [
    'GUILTY PLEA/DEFAULT', 'VIOLATIONS HANDLED BY CLERK', 'DISMISSED',
    'BY TRIAL TO COURT', 'CLOSED', 'OTHER JUDGMENT', 'TRANSFERRED',
    'SMALL CLAIM-DISPOSED BY CLERK', 'DEFAULTED', 'DEFERRED JUDGEMENT',
    'DISCHARGE', 'CONVERTED TO SIMPLE MISDEMEANR',
]


def test_dismissed_is_the_only_wording_that_translates():
    """The claim case_level_code's docstring makes, enforced rather than
    asserted. A later change to the code map that starts translating one of
    these silently changes what five sheets compute."""
    translated = {s: crs.case_level_code(s) for s in CASE_LEVEL_STATUSES}
    assert translated.pop('DISMISSED') == 'DISM'
    assert set(translated.values()) == {None}, translated


def test_deferred_judgement_at_case_level_is_still_not_a_code():
    """The one most likely to be answered first, since it names a CRS code
    outright and both the licence and expungement sheets test for DEF. Until
    Iowa Legal Aid says so it stays uncoded, and this fails if that changes
    without the open question being closed."""
    assert crs.case_level_code('DEFERRED JUDGEMENT') is None


# -- the wording, checked against the template it describes ------------------

def test_the_three_sheets_really_do_read_an_empty_column_g_that_way():
    """The note makes a claim about the workbook. If a later template changes
    the formula, this fails rather than the note quietly going wrong."""
    book = load_workbook(FULL)
    row = str(crs.FIRST_CASE_ROW)
    for sheet in ('BANKRUPTCY', 'EXEMPTIONS', 'SOL'):
        formula = book[sheet]['B' + row].value or ''
        assert 'open charge' in formula, (sheet, formula)
        assert "'CASE DATA'!G" + row in formula, (sheet, formula)


# -- what the run said about the row, which is the half that stayed wrong ----
#
# Column V learned the difference above. The progress page and the email did
# not, and they are the half Alex reads. A production run on 5 August pulled 71
# cases and mailed:
#
#   Iowa Courts recorded "CLOSED" on 1 case, which Napier does not recognise.
#   Those rows are coded OTH and say so in the notes column: <case>.
#
# Every clause of that is false of the row it names. The wording is not missing
# from charge_code_map, the row is not coded OTH, and the notes column says
# something else entirely, so the one sentence pointing at the workbook pointed
# at a note that contradicted it.


class FakeJob(object):
    id = 'synthetic-job-id'
    kind = 'crs'

    def __init__(self):
        self.progress = []
        self.messages = []

    def log(self, message, **kwargs):
        self.messages.append(message)


@pytest.fixture(autouse=True)
def clean_alerts():
    alerts.reset()
    yield
    alerts.reset()


def _recorded(monkeypatch):
    sent = []
    monkeypatch.setattr(alerts, 'record',
                        lambda *a, **kw: sent.append((a, kw)))
    return sent


UNREADABLE_WORDING = 'SYNTHETIC WORDING NOBODY HAS SEEN'
UNCODED = {(CLOSED, False): ['00000  SRCR000000']}
CODED = {(UNREADABLE_WORDING, True): ['00000  FECR000000']}


def test_the_run_does_not_say_the_row_is_coded_oth(monkeypatch):
    """The defect, in the words that were mailed out."""
    _recorded(monkeypatch)
    job = FakeJob()
    tasks.report_unknown_dispositions(job, UNCODED)
    assert job.messages
    for message in job.messages:
        assert 'OTH' not in message, message


def test_the_run_says_what_the_sheets_will_do_with_the_row(monkeypatch):
    """Same thing column V had to say, for the same reason: an attorney is
    about to read three sheets calling a closed case a pending charge."""
    _recorded(monkeypatch)
    job = FakeJob()
    tasks.report_unknown_dispositions(job, UNCODED)
    message = " ".join(job.messages)
    assert 'open charge' in message, message
    for sheet in ('BANKRUPTCY', 'EXEMPTIONS', 'SOL'):
        assert sheet in message, message


def test_the_run_still_names_the_wording_and_the_case(monkeypatch):
    """The two facts the old line did carry. Without them the mail says a run
    went wrong somewhere and stops."""
    _recorded(monkeypatch)
    job = FakeJob()
    tasks.report_unknown_dispositions(job, UNCODED)
    message = " ".join(job.messages)
    assert CLOSED in message and '00000  SRCR000000' in message, message


def test_it_goes_out_under_its_own_subject(monkeypatch):
    """Not the missing-word subject. Nothing is missing from charge_code_map
    here, and what CLOSED deserves is a question for Iowa Legal Aid, so it is
    filed and answered separately."""
    sent = _recorded(monkeypatch)
    tasks.report_unknown_dispositions(FakeJob(), UNCODED)
    assert len(sent) == 1
    args, kwargs = sent[0]
    assert args[2] == alerts.UNCODED_CASE_STATUS
    assert kwargs['case status'] == CLOSED
    assert '00000  SRCR000000' in kwargs['cases']


def test_the_alert_carries_no_defendant(monkeypatch):
    """Article 5, same as the alert it split from. The status and the case
    number are court public record; the client is not."""
    sent = _recorded(monkeypatch)
    tasks.report_unknown_dispositions(FakeJob(), UNCODED)
    _, kwargs = sent[0]
    assert set(kwargs) <= {'progress', 'case status', 'cases'}


def test_a_run_carrying_both_reports_both(monkeypatch):
    """The cost of the shared class, which the wording hid. alerts.record
    suppresses a second email of the same class within the floor, so a run with
    one of each used to mail whichever came first and drop the other."""
    sent = _recorded(monkeypatch)
    job = FakeJob()
    both = dict(UNCODED)
    both.update(CODED)
    tasks.report_unknown_dispositions(job, both)
    assert {args[2] for args, _ in sent} == {alerts.UNCODED_CASE_STATUS,
                                            alerts.UNKNOWN_DISPOSITION}
    assert len(job.messages) == 2


def test_the_other_row_still_gets_the_line_it_always_had(monkeypatch):
    """A count Napier could not read is coded OTH, and nothing above changes
    what is said about it."""
    sent = _recorded(monkeypatch)
    job = FakeJob()
    tasks.report_unknown_dispositions(job, CODED)
    assert sent[0][0][2] == alerts.UNKNOWN_DISPOSITION
    assert 'coded OTH' in " ".join(job.messages)
    assert 'open charge' not in " ".join(job.messages)


# -- and that a real build carries the distinction that far -----------------

def test_the_build_hands_back_which_kind_each_wording_was(tmp_path, monkeypatch):
    """process_case can tell them apart all it likes if build_workbook drops
    it. Both cases in one workbook, because that is the run that was losing an
    email as well as mislabelling one."""
    monkeypatch.setattr(tasks, 'tmp_dir', str(tmp_path) + os.sep)
    _, unknown, _, _ = tasks.build_workbook(
        [_case(UNADJUDICATED, CLOSED, case_id='00000  SRCR000000'),
         _case(UNREADABLE, '', case_id='00000  FECR000000')],
        'TEST CLIENT', '01/01/1980', False)
    assert unknown == {(CLOSED, False): ['00000  SRCR000000'],
                       (UNREADABLE_WORDING, True): ['00000  FECR000000']}
