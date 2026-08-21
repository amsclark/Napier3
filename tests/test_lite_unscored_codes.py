"""CRS Lite has no formula for JWV or CIV, and Napier writes both.

The templates are not the same vocabulary. Every code in the full CRS is named
by some formula on some sheet; CRS Lite names neither JWV nor CIV anywhere at
all. A Lite workbook carrying one of those rows counted it on no sheet and said
nothing about it, so the file read as complete.

Both codes are live. CIV has come out of every extradition hold and cleared
parole violation since 19 August, and JWV since Iowa Legal Aid settled the
juvenile transfer wording on 20 August.

The workbook now says so in column V on the row itself, rather than the run
refusing to build or the code being quietly swapped for one the template
happens to know. Swapping would be worse: JUV is the nearest code Lite scores
and it means the juvenile court kept the case, which is the opposite of a
waiver up to adult court.

Synthetic case numbers throughout. The repository is public.
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import crs
from openpyxl import load_workbook

FULL = 'CRS 3.5.5.xlsx'
LITE = 'CRS Lite 3.5.5.xlsx'


class FakeSheet(dict):
    """Just enough sheet for note_unscored_codes: column G in, column V out."""

    class Cell:
        def __init__(self, value=None):
            self.value = value

    def __init__(self, codes):
        super().__init__()
        for offset, code in enumerate(codes):
            self['G' + str(crs.FIRST_CASE_ROW + offset)] = self.Cell(code)

    def __getitem__(self, key):
        return self.setdefault(key, FakeSheet.Cell())

    def note_on(self, offset):
        return self['V' + str(crs.FIRST_CASE_ROW + offset)].value


def _notes(template, codes):
    sheet = FakeSheet(codes)
    crs.note_unscored_codes(template, sheet,
                            crs.FIRST_CASE_ROW + len(codes) - 1)
    return [sheet.note_on(i) for i in range(len(codes))]


class TestWhatEachTemplateScores:
    def test_the_full_crs_scores_everything_napier_writes(self):
        """Except OTH, which is meant to be in no formula in either file."""
        scored = crs.codes_the_template_scores(FULL)
        produced = set()
        for entry in crs.charge_code_map.values():
            produced.update(entry)
        produced.update(*crs.JUVENILE_DISPOSITIONS.values())
        assert produced - scored == set(crs.DELIBERATELY_UNSCORED)

    def test_lite_is_missing_jwv_and_civ(self):
        """The premise. If a future template fixes this, delete the guard."""
        scored = crs.codes_the_template_scores(LITE)
        assert 'JWV' not in scored
        assert 'CIV' not in scored

    def test_lite_does_score_the_ordinary_ones(self):
        scored = crs.codes_the_template_scores(LITE)
        for code in ('GTR', 'GPL', 'DEF', 'JUV', 'DISM', 'ACQ', 'WTHD',
                     'NOTF', 'TNSF'):
            assert code in scored, code


class TestTheNoteOnALiteWorkbook:
    @pytest.mark.parametrize('code', ['JWV', 'CIV'])
    def test_an_unscored_code_says_so(self, code):
        note = _notes(LITE, [code])[0]
        assert note == crs.UNSCORED_CODE_NOTE % code
        assert code in note

    def test_a_scored_code_says_nothing(self):
        assert _notes(LITE, ['GTR', 'JUV', 'DISM']) == [None, None, None]

    def test_only_the_row_that_carries_it_is_marked(self):
        assert _notes(LITE, ['GTR', 'JWV', 'DISM']) == [
            None, crs.UNSCORED_CODE_NOTE % 'JWV', None]

    def test_oth_is_left_alone(self):
        """OTH is in no formula on purpose, and the row already carries a
        column V note saying the disposition was not recognised. Marking it
        again would put the note on every uncoded row in every workbook."""
        assert _notes(LITE, ['OTH']) == [None]

    def test_an_empty_code_is_left_alone(self):
        """A blank column G is an open charge, which three sheets read on
        purpose."""
        assert _notes(LITE, [None, '']) == [None, None]

    def test_it_joins_a_note_already_there(self):
        sheet = FakeSheet(['JWV'])
        crs.append_note(sheet, crs.FIRST_CASE_ROW, 'Filed under TEST NAME.')
        crs.note_unscored_codes(LITE, sheet, crs.FIRST_CASE_ROW)
        note = sheet.note_on(0)
        assert note.startswith('Filed under TEST NAME.')
        assert 'JWV' in note


class TestTheFullCrsIsSilent:
    @pytest.mark.parametrize('code', ['JWV', 'CIV', 'GTR', 'TNSF', 'OTH'])
    def test_nothing_is_marked(self, code):
        assert _notes(FULL, [code]) == [None]


# The columns on the shared sheet that name a disposition code, in each file.
# They are the same columns in a different order: CRS Lite drops the four
# columns that read BANKRUPTCY and SOL, and everything after them shifts left.
SHARED_CODE_COLUMNS = [('E', 'E'), ('F', 'F'), ('I', 'I'), ('M', 'L'),
                       ('T', 'Q')]

# The three sheets the full CRS has and CRS Lite does not, which is the whole
# of the difference between them.
SHEETS_ONLY_THE_FULL_CRS_HAS = ['BANKRUPTCY', 'EXEMPTIONS', 'SOL']


def _formula(cell):
    """A cell's text whether the template stored it as an array formula."""
    value = cell.value
    return getattr(value, 'text', value)


class TestWhatLiteIsActuallyShortOf:
    """Iowa Legal Aid asked on 21 August whether the changes on the full CRS
    could be matched on Lite, which is what Drake Legal Clinic runs.

    Almost all of them already are, and not by anyone porting them: Napier
    writes the same columns into whichever template was chosen, so the notes
    column, the charge classes in column E, the disposition codes in column G,
    the court debt and the contempt and juvenile readings land in a Lite
    workbook exactly as they land in a full one.

    What does not is anything that needs a sheet Lite has not got. These tests
    measure that boundary off the two files rather than describing it, so the
    answer stays true when a template is replaced by its author -- which is
    how these files change, since this repo does not own them.
    """

    def test_the_shared_sheet_asks_the_same_questions_in_both(self):
        """Every code-bearing formula on EXPUNGEMENT & 910.7, character for
        character. This is the test behind UNSCORED_CODE_NOTE: a CIV or JWV
        row gets the same eligibility answers on Lite as on the full CRS, so a
        note telling staff to rebuild on the full CRS to fix those answers
        would be sending them after nothing."""
        full = load_workbook(FULL)['EXPUNGEMENT & 910.7']
        lite = load_workbook(LITE)['EXPUNGEMENT & 910.7']
        for full_column, lite_column in SHARED_CODE_COLUMNS:
            row = str(crs.FIRST_CASE_ROW - 1)
            assert (_formula(full[full_column + row])
                    == _formula(lite[lite_column + row])), full_column

    def test_the_codes_lite_cannot_see_are_named_only_on_its_missing_sheets(self):
        """So the gap is those sheets, not a formula anybody forgot to update.
        Adding CIV to a Lite formula is not a change that exists to make: the
        formulas that name it are on BANKRUPTCY and EXEMPTIONS, and putting
        those sheets in is what would stop the file being Lite."""
        missing = crs.codes_the_template_scores(FULL) - \
            crs.codes_the_template_scores(LITE)
        assert missing == {'CIV', 'JWV'}
        for code in missing:
            named_on = _sheets_naming(FULL, code)
            assert named_on, code
            assert named_on <= set(SHEETS_ONLY_THE_FULL_CRS_HAS), (code,
                                                                   named_on)

    def test_that_is_the_only_difference_in_sheets(self):
        """If a later template adds one of these back, the note above and the
        answer given to Iowa Legal Aid both need revisiting."""
        full = set(load_workbook(FULL).sheetnames)
        lite = set(load_workbook(LITE).sheetnames)
        assert full - lite == set(SHEETS_ONLY_THE_FULL_CRS_HAS)
        assert lite - full == set()


def _sheets_naming(template, code):
    """Which sheets of a template mention a disposition code in a formula."""
    workbook = load_workbook(template)
    found = set()
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(min_row=1, max_row=crs.FIRST_CASE_ROW + 200)
        for row in rows:
            for cell in row:
                text = _formula(cell)
                if isinstance(text, str) and '"%s"' % code in text:
                    found.add(sheet.title)
                    break
    return found
