"""Whether the workbook Napier saves reaches its own last case.

This file used to check a prediction. crs.py carried the depths the templates
shipped at -- 97 cases in the totals, 147 rows on SOL, 297 on CASE DATA -- and
the finish page warned whenever a run was bigger than one of them, and here
those numbers were read back out of the two .xlsx files so a CRS 3.6 that moved
them would fail rather than go unnoticed.

What went unnoticed instead was the extension. grid.extend_formula_grid has
filled those sheets down to the case list on every build since, so the ceilings
stopped being true of anything Napier saves, and the warning went on quoting
them: Iowa Legal Aid ran 184 cases on 2026-08-18, got a complete workbook, and
was told on the finish page that the analysis sheets stopped at 147 and that
splitting the search in two was the way around it. Neither half was true, and
Napier has never had a way to split a search.

So the templates' depths are still checked here, because they are what the
extension has to start from, but the thing staff are told now comes from
grid.shortfalls measuring the saved workbook. The two tests that matter are
that it finds the real ceilings on a template nobody extended, and finds
nothing at all on one Napier built.
"""

import os
import re
import sys

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import crs
import grid

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')
LITE = os.path.join(ROOT, 'CRS Lite 3.5.5.xlsx')

# The three sheets whose row-1 totals are the ones staff read off.
TOTALLED = ('SOL', 'BANKRUPTCY', 'EXEMPTIONS')

# What the shipped templates reach, in cases. Not thresholds any more: the
# starting point the extension has to cover, and the answer the measurement
# below has to produce on a workbook that was never extended.
CASE_DATA_TOTALS_REACH = 297
ANALYSIS_TOTALS_REACH = 97
ANALYSIS_ROWS_REACH = 147
LITE_ANALYSIS_ROWS_REACH = 198


def _total_range_end(sheet):
    """Last row the =SUM() totals on row 1 reach, or None if there are none."""
    for cell in sheet[1]:
        if isinstance(cell.value, str) and cell.value.startswith('=SUM('):
            match = re.search(r':[A-Z]+(\d+)\)', cell.value.replace('\n', ''))
            if match:
                return int(match.group(1))
    return None


def _last_case_row_covered(sheet):
    """Last CASE DATA row this sheet has a row for.

    Each analysis sheet mirrors column A of CASE DATA, some of them offset by a
    row or two, so the reference inside the formula is the honest answer rather
    than the sheet's own row number.
    """
    last = None
    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row, column=1).value
        if not isinstance(value, str):
            continue
        match = re.search(r"'CASE DATA'!A(\d+)", value)
        if match:
            last = max(last or 0, int(match.group(1)))
    return last


def _built(count, template=FULL):
    """A template with a case list written in and the grid extended to it."""
    workbook = load_workbook(template)
    sheet = workbook['CASE DATA']
    for n in range(count):
        sheet['A%d' % (crs.FIRST_CASE_ROW + n)] = '00000  FECR%06d' % n
    grid.extend_formula_grid(workbook, count)
    return workbook


@pytest.fixture(scope='module')
def full():
    return load_workbook(FULL)


@pytest.fixture(scope='module')
def lite():
    return load_workbook(LITE)


# -- what the templates ship with --------------------------------------------

def test_case_data_totals_stop_where_we_say_they_do(full, lite):
    for workbook in (full, lite):
        end = _total_range_end(workbook['CASE DATA'])
        cases = end - crs.FIRST_CASE_ROW + 1
        assert cases == CASE_DATA_TOTALS_REACH


def test_analysis_totals_stop_where_we_say_they_do(full):
    ends = {name: _total_range_end(full[name]) for name in TOTALLED}
    assert None not in ends.values(), ends
    cases = min(ends.values()) - crs.FIRST_CASE_ROW + 1
    assert cases == ANALYSIS_TOTALS_REACH


def test_the_lite_template_has_no_totalled_analysis_sheets(lite):
    assert not [name for name in TOTALLED if name in lite.sheetnames]


def test_the_analysis_sheets_run_out_of_rows_where_we_say_they_do(full, lite):
    for workbook, reach in ((full, ANALYSIS_ROWS_REACH),
                            (lite, LITE_ANALYSIS_ROWS_REACH)):
        covered = {}
        for name in workbook.sheetnames:
            if name == 'CASE DATA':
                continue
            last = _last_case_row_covered(workbook[name])
            if last is not None:
                covered[name] = last - crs.FIRST_CASE_ROW + 1
        assert covered, 'no sheet mirrors CASE DATA any more'
        assert min(covered.values()) == reach, covered


# -- the measurement, on a workbook nobody extended --------------------------
#
# Non-vacuity. Every silent result below is only worth something while these
# fail loudly, so they run against the raw templates at case counts chosen to
# land either side of each ceiling above.

def test_a_short_list_is_short_of_nothing(full):
    assert grid.shortfalls(full, ANALYSIS_TOTALS_REACH) == {}


def test_the_first_thing_to_give_way_is_the_analysis_totals(full):
    short = grid.shortfalls(full, ANALYSIS_TOTALS_REACH + 1)
    assert set(short) == set(TOTALLED), short
    for name in TOTALLED:
        assert short[name] == ['totals'], short


def test_past_the_analysis_grid_the_rows_go_too(full):
    short = grid.shortfalls(full, ANALYSIS_ROWS_REACH + 1)
    assert 'rows' in short['SOL'], short
    assert 'CASE DATA' not in short, short


def test_past_the_case_data_totals_the_case_list_itself_is_short(full):
    short = grid.shortfalls(full, CASE_DATA_TOTALS_REACH + 1)
    assert 'CASE DATA' in short, short


def test_the_lite_template_gives_way_on_rows_and_has_no_totals_to_lose(lite):
    assert grid.shortfalls(lite, LITE_ANALYSIS_ROWS_REACH) == {}
    short = grid.shortfalls(lite, LITE_ANALYSIS_ROWS_REACH + 1)
    assert short, 'the Lite template has no ceiling at all now?'
    assert not [name for name in short if name in TOTALLED], short


# -- the measurement, on a workbook Napier built -----------------------------

@pytest.mark.parametrize('count', [1, 40, 98, 148, 298, 400])
def test_an_extended_workbook_is_short_of_nothing(count):
    assert grid.shortfalls(_built(count), count) == {}


@pytest.mark.parametrize('count', [1, 40, 199, 400])
def test_an_extended_lite_workbook_is_short_of_nothing(count):
    assert grid.shortfalls(_built(count, LITE), count) == {}


def test_an_empty_run_measures_nothing(full):
    # No cases is not a workbook that stops before its last case, and a run
    # where Iowa Courts gave up nothing already says so in its own words.
    assert grid.shortfalls(full, 0) == {}


# -- what it puts in front of staff ------------------------------------------

def test_nothing_short_says_nothing():
    assert grid.describe_shortfalls({}) == []


def test_a_sheet_short_of_rows_says_where_those_cases_are():
    lines = grid.describe_shortfalls({'SOL': ['rows']})
    assert len(lines) == 1
    assert 'CASE DATA and nowhere else' in lines[0]
    assert lines[0].startswith('SOL ')


def test_a_sheet_short_of_totals_only_says_the_rows_are_right():
    line = grid.describe_shortfalls({'BANKRUPTCY': ['totals']})[0]
    assert 'rows are right' in line
    assert 'totals are low' in line


def test_the_case_list_gets_its_own_sentence():
    # "the cases past that point are on CASE DATA and nowhere else" is nonsense
    # when the sheet in question is CASE DATA.
    line = grid.describe_shortfalls({'CASE DATA': ['rows', 'totals']})[0]
    assert 'nowhere else' not in line
    assert 'TOTAL column' in line


def test_one_line_per_sheet_in_a_settled_order():
    lines = grid.describe_shortfalls({'SOL': ['rows'], 'BANKRUPTCY': ['totals'],
                                      'CASE DATA': ['totals']})
    assert len(lines) == 3
    assert lines[0].startswith('BANKRUPTCY')
    assert lines[1].startswith('CASE DATA')
    assert lines[2].startswith('SOL')
