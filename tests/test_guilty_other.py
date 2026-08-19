"""A conviction ICOS words as GUILTY - OTHER.

Widening the captured corpus from 300 cases to 400, and from 57 counties to
68, turned up a disposition wording nothing before it had shown. One case
carries it on both of its counts, two simple misdemeanours, with $640.00 still
owed on the case.

charge_code_map had no entry for it, so it fell through to OTH, which is the
code that means Napier could not read the word. The licence sheet reads OTH as
no conviction, so LICENSE-REGIS was answering "Z - Neither license nor
registration" about a client who has a conviction and $640.00 of court debt,
which is exactly the pair of facts that puts a registration hold in play. The
row was honest about it, the note in column V said so, and the sheet still gave
the wrong answer to anyone reading the sheet.

The only thing "- OTHER" leaves genuinely open is how the guilty finding was
arrived at, and the last test here is the reason that does not matter: every
formula in both templates that names GTR also names GPL, so nothing computed
anywhere turns on which of the two a case is coded. What is not open is that
the client was found guilty.

The case here is synthetic. This repository is public and a real charges page
is one person's unredacted criminal record.
"""

import os
import re
import sys
from decimal import Decimal

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import case_parser
import crs
import tasks

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')
LITE = os.path.join(ROOT, 'CRS Lite 3.5.5.xlsx')

WORDING = 'GUILTY - OTHER'

# What the real case owed, kept because a licence answer is only reachable on a
# case with money on it and this is the amount that was actually at stake.
OWED = Decimal('640.00')

# A reference to a CASE DATA disposition code inside a formula, which is how
# LICENSE-REGIS asks whether a case is a conviction.
CODE_TEST = re.compile(r"G(\d+)\s*=\s*\"([A-Z]+)\"")


def charges_page(*adjudications):
    """A charges page with one count per adjudication, in the shape ICOS
    serves it. Parsed rather than hand-built so this exercises the same path a
    real page takes."""
    def cells(*values):
        return '<tr>%s</tr>' % ''.join(
            '<td><font size="2">%s</font></td>' % value for value in values)
    rows = ['<html><body><table>']
    for number, adjudication in enumerate(adjudications, 1):
        rows += [
            cells('Count %02d' % number, 'Original Charge'),
            cells('Charge:', '714.2(5)', 'Description:', 'SYNTHETIC OFFENCE'),
            cells('Offense Date:', '01/01/1900', 'Arrest Date:', ''),
            cells('Adjudication'),
            cells('Charge:', '714.2(5)', 'Description:', 'SYNTHETIC OFFENCE'),
            cells('Adjudication:', adjudication,
                  'Adjudication Date:', '02/02/1901'),
        ]
    rows.append('</table></body></html>')
    return ''.join(rows).encode('utf-8')


def synthetic_case(*adjudications):
    """One case owing OWED, spread so it reaches more than one of J to S."""
    case = {'id': '00000  SMSM000000', 'county': 'SYNTHETIC',
            'financials': [], 'sentences': [],
            'summary_created_date': '01/01/1900',
            'summary_disposition_date': '02/02/1901',
            'summary_dispo_status': ''}
    case_parser.parse_case_charges(charges_page(*adjudications), case)
    share = OWED / 4
    case['summary_categories'] = [
        {'label': label, 'original': share, 'paid': Decimal('0'),
         'due': share}
        for label in ('COSTS', 'FINE', 'SURCHARGE', 'RESTITUTION')]
    case['total_due'] = '$%s' % OWED
    return case


def licence_codes():
    """The codes LICENSE-REGIS will answer a licence or registration question
    about. Read off the shipped template rather than listed here, because the
    template is what a staffer is reading."""
    sheet = load_workbook(FULL)['LICENSE-REGIS']
    text = sheet['C2'].value
    assert isinstance(text, str) and text.startswith('='), text
    codes = {code for _, code in CODE_TEST.findall(text)}
    assert codes, text
    return codes


# -- the code ---------------------------------------------------------------

def test_the_wording_is_read_as_a_conviction():
    case = synthetic_case(WORDING)
    assert crs.get_dominant_charge(case['charges'])['disposition'] == 'GTR'


def test_it_is_no_longer_reported_as_a_wording_napier_cannot_read():
    """Both counts, which is how the real case carries it. While it was
    unmapped every run that met this case emailed about it."""
    case = synthetic_case(WORDING, WORDING)
    charge = crs.get_dominant_charge(case['charges'])
    assert charge['unknown_dispositions'] == []


def test_a_guilty_count_beside_it_still_reads_as_guilty():
    """It is ranked with the convictions, so it cannot demote a case the way
    an unreadable word used to."""
    for other in ('GUILTY', 'GUILTY - NEGOTIATED/VOLUN PLEA', 'DEFERRED'):
        case = synthetic_case(WORDING, other)
        code = crs.get_dominant_charge(case['charges'])['disposition']
        assert code in ('GTR', 'GPL', 'DEF'), (other, code)


def test_the_negotiated_plea_wording_is_untouched():
    """The two wordings share a prefix and ICOS writes both. Mapping one must
    not start answering for the other."""
    assert crs.get_dominant_charge(
        synthetic_case('GUILTY - NEGOTIATED/VOLUN PLEA')
        ['charges'])['disposition'] == 'GPL'


# -- what the sheet says about it -------------------------------------------

def test_the_licence_sheet_can_now_answer_about_the_case():
    """The whole point, read the way the spreadsheet reads it.

    Comparing formulas proves nothing here: LICENSE-REGIS column C carries one
    formula and it is the same before and after. What decides the answer is
    whether the code the build put in column G is in the list that formula
    clears, so this builds the workbook, takes the code off CASE DATA, and
    looks for it.
    """
    path, unknown, _, _ = tasks.build_workbook(
        [synthetic_case(WORDING, WORDING)], 'TEST CLIENT', '01/01/1980', False)
    try:
        cases = load_workbook(path)['CASE DATA']
        code = cases['G4'].value
        # The licence formula only reaches its code test on a case that owes
        # something, so a case owing nothing would pass this whatever the code
        # said. The money is checked before the code is.
        owed = [cases.cell(row=4, column=column).value or 0
                for column in range(10, 20)]              # J to S
        assert sum(owed) == OWED, owed
        assert len([amount for amount in owed if amount]) >= 2, owed
        assert code in licence_codes(), (
            'CASE DATA G4 is %r, and LICENSE-REGIS answers about %s'
            % (code, sorted(licence_codes())))
        assert unknown == {}, unknown
    finally:
        os.remove(path)


def test_the_row_carries_no_note_saying_napier_guessed():
    """Column V told the staffer the code was a guess and the licence sheet
    would read it as no conviction. Neither is true now, and a note that has
    stopped being true is worse than no note."""
    sheet = load_workbook(FULL)['CASE DATA']
    case = synthetic_case(WORDING, WORDING)
    unknown = crs.process_case(case, sheet, crs.FIRST_CASE_ROW)
    note = sheet['V%d' % crs.FIRST_CASE_ROW].value or ''
    assert unknown == []
    assert WORDING not in note, note
    assert 'OTH' not in note, note


# -- why GTR and not GPL ----------------------------------------------------

def test_no_formula_anywhere_tells_a_trial_conviction_from_a_plea():
    """The justification for picking one, made enforceable.

    "- OTHER" does not say whether the guilty finding came from a trial or a
    plea, and GTR and GPL are the two codes that would. Choosing between them
    is safe only for as long as nothing computes a different answer from them,
    so this fails if a formula is ever added that does. Then GUILTY - OTHER
    needs a real answer from Iowa Legal Aid rather than the convenient one.
    """
    lonely = []
    for template in (FULL, LITE):
        workbook = load_workbook(template)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    text = cell.value
                    if not isinstance(text, str) or not text.startswith('='):
                        continue
                    if ('GTR' in text) != ('GPL' in text):
                        lonely.append('%s %s %s' % (
                            os.path.basename(template), sheet.title,
                            cell.coordinate))
    assert not lonely, lonely[:6]
