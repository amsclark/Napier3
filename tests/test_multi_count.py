"""A case whose counts were disposed on different days, and the one date column D has.

ICOS gives every count its own adjudication and its own adjudication date. The
CRS has one row per case, so column G is chosen by rank across the counts and
column D holds a single date. Those two used to come from different counts:
column G from whichever count outranked the rest, column D from whichever count
ICOS printed first.

On three of the nine captured multi-count cases the counts really were disposed
on different days, and one of them shows the pairing coming apart. A count
dismissed, and a second count pleaded out some months later. The row reported
the plea against the date of the dismissal.

That case used to be written here with an adjudication on top, which is how
ICOS records a probation violation and which used to outrank the plea. It no
longer does, on anything but a juvenile case number, because Iowa Legal Aid
reported that the row then carried both the wrong code and the wrong date. See
test_adult_adjudication.py.

The date is the part that costs something. The SOL sheet asks

  IF('CASE DATA'!D4 + 7300 < today, ...)

on every row, which is the twenty year limit on enforcing an Iowa judgment, and
sorts the row's debt into time barred or "NO ARGUMENT" on the answer. Reaching
for the earliest date on the page makes a judgment look older than it is, so
the error ran in the direction of telling an attorney a debt had aged out when
it had not.

The pages here are synthetic. This repo is public and a real charges page is one
person's unredacted criminal record.
"""

import os
import sys
from decimal import Decimal

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import case_parser
import crs

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')


def _cells(*cells):
    return '<tr>%s</tr>' % ''.join(
        '<td><font size="2">%s</font></td>' % cell for cell in cells)


def charges_page(counts):
    """ICOS's real shape for a multi-count case.

    Each count is (statute, description, adjudication, adjudication date). An
    adjudication of None is a count no court has ruled on yet, which ICOS prints
    with the block present and its cells empty.
    """
    html = ['<html><body><table>']
    for number, (statute, description, outcome, when) in enumerate(counts, 1):
        html.append(_cells('Count %02d' % number, 'Original Charge'))
        html.append(_cells('Charge:', statute, 'Description:', description))
        html.append(_cells('Offense Date:', '01/01/1900', 'Arrest Date:', ''))
        html.append(_cells('Adjudication'))
        if outcome is None:
            html.append(_cells('Charge:', '', 'Description:', ''))
            html.append(_cells('Adjudication:', '', 'Adjudication Date:', ''))
        else:
            html.append(_cells('Charge:', statute, 'Description:', description))
            html.append(_cells('Adjudication:', outcome,
                               'Adjudication Date:', when))
    html.append('</table></body></html>')
    return ''.join(html).encode('utf-8')


def _case(counts, costs='65.75'):
    case = {'id': '00000  FECR000000', 'county': 'SYNTHETIC',
            'financials': [], 'sentences': [],
            'summary_created_date': '01/01/1900',
            'summary_disposition_date': '', 'summary_dispo_status': ''}
    case_parser.parse_case_charges(charges_page(counts), case)
    case['summary_categories'] = [
        {'label': label,
         'original': Decimal(costs) if label == 'COSTS' else Decimal('0'),
         'paid': Decimal('0'),
         'due': Decimal(costs) if label == 'COSTS' else Decimal('0')}
        for label in ('COSTS', 'FINE', 'SURCHARGE', 'RESTITUTION', 'OTHER')]
    case['total_due'] = '$' + costs
    return case


def _row(counts, costs='65.75'):
    sheet = load_workbook(FULL)['CASE DATA']
    case = _case(counts, costs)
    crs.process_case(case, sheet, crs.FIRST_CASE_ROW)
    row = str(crs.FIRST_CASE_ROW)
    return {column: sheet[column + row].value
            for column in ('D', 'F', 'G', 'V')}


# -- the parser keeps a date per count ---------------------------------------

def test_every_count_keeps_its_own_adjudication_date():
    """Nothing recorded them before, so nothing could pair them up."""
    case = {'id': '00000  FECR000000'}
    case_parser.parse_case_charges(charges_page([
        ('714.2(3)', 'SYNTHETIC THEFT', 'DISMISSED BY COURT', '01/01/1900'),
        ('124.401(5)', 'SYNTHETIC POSSESSION', 'GUILTY', '02/02/1901'),
    ]), case)
    charge = case['charges'][0]
    assert charge['disposition_dates'] == ['02/02/1901', '01/01/1900']


def test_the_dates_line_up_with_the_dispositions_they_belong_to():
    """They are read as a pair, so they have to be sorted the same way.

    Both lists are built newest first. A list sorted differently from its
    partner is a quiet way to pair the wrong count with the wrong date.
    """
    case = {'id': '00000  FECR000000'}
    case_parser.parse_case_charges(charges_page([
        ('714.2(3)', 'SYNTHETIC THEFT', 'DISMISSED BY COURT', '01/01/1900'),
        ('124.401(5)', 'SYNTHETIC POSSESSION', 'GUILTY', '02/02/1901'),
    ]), case)
    charge = case['charges'][0]
    pairs = list(zip(charge['disposition'], charge['disposition_dates']))
    assert pairs == [('GUILTY', '02/02/1901'),
                     ('DISMISSED BY COURT', '01/01/1900')]


# -- the date follows the code -----------------------------------------------

def test_column_d_is_the_date_of_the_count_column_g_names():
    """The captured shape: a plea, then a dismissal, then an adjudication.

    Column G takes the plea. The adjudication is 908.11, violation of probation,
    and on a felony case number that is the clerk writing "Adjudicated" where
    the juvenile court's word does not belong, so it no longer outranks the
    conviction it is a violation of. Column D follows the count column G names,
    which is the plea in 1900 rather than the violation three years later.

    This test asserted JUV and 03/03/1903 until Iowa Legal Aid reported both as
    wrong on 3 August 2026.
    """
    cells = _row([
        ('715A.2(2)(A)', 'SYNTHETIC FORGERY', 'GUILTY', '01/01/1900'),
        ('908.11', 'SYNTHETIC VIOLATION', 'DISMISSED BY COURT', '02/02/1901'),
        ('715A.2(2)(A)', 'SYNTHETIC FORGERY', 'ADJUDICATED', '03/03/1903'),
    ])
    assert cells['G'] == 'GTR'
    assert cells['D'] == '01/01/1900'


def test_a_dismissal_printed_first_does_not_date_the_conviction():
    """The shape that reads worst: the row says guilty as of the dismissal."""
    cells = _row([
        ('714.2(3)', 'SYNTHETIC THEFT', 'DISMISSED BY COURT', '01/01/1900'),
        ('124.401(5)', 'SYNTHETIC POSSESSION', 'GUILTY', '02/02/1901'),
    ])
    assert cells['G'] == 'GTR'
    assert cells['D'] == '02/02/1901'


def test_the_conviction_coming_first_is_left_alone():
    """The guard. Two of the three captured cases already had it right and
    a fix that moves those has broken something."""
    cells = _row([
        ('124.401(5)', 'SYNTHETIC POSSESSION', 'GUILTY', '01/01/1900'),
        ('714.2(3)', 'SYNTHETIC THEFT', 'DISMISSED BY COURT', '02/02/1901'),
    ])
    assert cells['G'] == 'GTR'
    assert cells['D'] == '01/01/1900'


def test_two_counts_sharing_the_winning_code_take_the_later_date():
    """A captured case convicted on two counts two years apart.

    The case finished reaching that disposition on the second date, and it is
    the reading that will not retire a debt early.
    """
    cells = _row([
        ('714.2(3)', 'SYNTHETIC THEFT', 'GUILTY', '01/01/1900'),
        ('714.2(3)', 'SYNTHETIC THEFT', 'GUILTY', '02/02/1901'),
    ])
    assert cells['G'] == 'GTR'
    assert cells['D'] == '02/02/1901'


def test_a_single_count_case_is_untouched():
    """81 of the 90 captured cases have exactly one count."""
    cells = _row([
        ('124.401(5)', 'SYNTHETIC POSSESSION', 'GUILTY', '02/02/1901'),
    ])
    assert cells['G'] == 'GTR'
    assert cells['D'] == '02/02/1901'


def test_an_unadjudicated_count_contributes_no_date():
    """It has no adjudication, so it cannot date one.

    Its empty date must not be picked up as the pair for a code that came from
    a count that really was decided.
    """
    cells = _row([
        ('124.401(5)', 'SYNTHETIC POSSESSION', 'GUILTY', '02/02/1901'),
        ('714.2(3)', 'SYNTHETIC THEFT', None, ''),
    ])
    assert cells['G'] == 'GTR'
    assert cells['D'] == '02/02/1901'


# -- and the row says it is compressing more than one judgment ---------------

def test_a_case_disposed_over_several_days_says_so():
    """The sheet reads one date and applies it to every dollar on the row."""
    cells = _row([
        ('714.2(3)', 'SYNTHETIC THEFT', 'DISMISSED BY COURT', '01/01/1900'),
        ('124.401(5)', 'SYNTHETIC POSSESSION', 'GUILTY', '02/02/1901'),
    ])
    note = cells['V'] or ''
    assert '01/01/1900' in note
    assert '02/02/1901' in note
    assert '20 year' in note


def test_counts_disposed_the_same_day_stay_quiet():
    """Six of the nine captured multi-count cases were disposed in one hearing.

    There is no spread to warn about and column V is where the notes that
    matter go.
    """
    cells = _row([
        ('714.2(3)', 'SYNTHETIC THEFT', 'DISMISSED BY COURT', '02/02/1901'),
        ('124.401(5)', 'SYNTHETIC POSSESSION', 'GUILTY', '02/02/1901'),
    ])
    assert 'different dates' not in (cells['V'] or '')


def test_a_case_that_owes_nothing_stays_quiet():
    """The SOL sheet sorts debt into buckets. With no debt there is none to
    sort, and the caveat describes a decision nobody is going to make."""
    cells = _row([
        ('714.2(3)', 'SYNTHETIC THEFT', 'DISMISSED BY COURT', '01/01/1900'),
        ('124.401(5)', 'SYNTHETIC POSSESSION', 'GUILTY', '02/02/1901'),
    ], costs='0')
    assert 'different dates' not in (cells['V'] or '')


def test_the_caveat_joins_what_the_money_left_in_column_v():
    """Column V belongs to process_financials first. This joins, never replaces."""
    sheet = load_workbook(FULL)['CASE DATA']
    case = _case([
        ('714.2(3)', 'SYNTHETIC THEFT', 'DISMISSED BY COURT', '01/01/1900'),
        ('124.401(5)', 'SYNTHETIC POSSESSION', 'GUILTY', '02/02/1901'),
    ])
    case['financials'] = [
        {'detail': 'SYNTHETIC UNCATEGORISED LINE', 'amount': Decimal('10.00'),
         'paid': None},
    ]
    crs.process_case(case, sheet, crs.FIRST_CASE_ROW)
    note = sheet['V' + str(crs.FIRST_CASE_ROW)].value or ''
    assert 'MISCELLANEOUS' in note, note
    assert 'different dates' in note, note
