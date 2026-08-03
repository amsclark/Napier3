"""The two expungement columns that answered NO on every case ever run.

Iowa Code 123.46(6) and 123.47(9) expunge a public intoxication or PAULA
conviction two years on, and 725.1(4) does the same for prostitution. These are
the cheapest things on the sheet to act on: no petition to draft against a
disputed record, just a wait and a filing. The EXPUNGEMENT sheet has a column
for each and both asked whether CASE DATA column F was the string "123.46",
"123.47" or "725.1" exactly.

Iowa Courts never write a section that bare. Column F is a semicolon joined list
of the adjudicated statutes and the statutes carry subsections, and a charge
brought under a city ordinance carries a two letter prefix as well. Across the
300 captured cases the two that cite one of these read "123.47(4)" and
"PO/123.47(2)". Neither is equal to "123.47", so both columns said NO, and the
one real conviction among them is a guilty plea to possession of alcohol by a
person under the legal age. That case is on a traffic citation number, which
closes the misdemeanour columns as "n/a" by design, so the PAULA column was the
only thing on the sheet that could have caught it.

Napier answers it now, the way it already answers "Vehicular?": in Python,
where the list can be taken apart, into a cell the sheet reads.

The cases here are synthetic. This repo is public and a real charges page is one
person's unredacted criminal record.
"""

import os
import re
import sys

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import crs
import statutes

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')
LITE = os.path.join(ROOT, 'CRS Lite 3.5.5.xlsx')

PAULA = crs.PUBLIC_INTOX_PAULA_SECTIONS
PROSTITUTION = crs.PROSTITUTION_SECTIONS


# -- the wordings Iowa Courts actually writes -------------------------------

# The first two are the only real ones in the corpus, and are the whole reason
# this file exists. An equality test passes none of them.

REAL = [
    ('PO/123.47(2)', PAULA,
     'the real PAULA conviction: an ordinance prefix and a subsection'),
    ('123.47(4)', PAULA, 'the other real one, a subsection and nothing else'),
]


@pytest.mark.parametrize('statute,sections,why', REAL)
def test_the_wordings_iowa_courts_actually_writes(statute, sections, why):
    assert crs.cites_section(statute, sections) == "YES", why


def test_none_of_the_captured_statutes_are_the_bare_section():
    """Why the equality test could never have fired.

    Not one of the 300 captured cases carries "123.46", "123.47" or "725.1" on
    its own, so a test written against an invented fixture that does would pass
    while the sheet stayed broken.
    """
    for statute, sections, _ in REAL:
        assert statute not in sections


CITED = [
    ('123.46', PAULA, 'public intoxication, bare'),
    ('123.46(2)', PAULA, 'and with a subsection'),
    ('123.47', PAULA, 'PAULA, bare'),
    ('PO/123.46', PAULA, 'an ordinance citation carrying the state section'),
    ('MA-123.47(2)', PAULA, 'the same with a dash, which is the other shape'),
    ('717B.9(1);123.46', PAULA, 'second count on a joined list'),
    ('123.46; 714.1', PAULA, 'first count, with a space after the semicolon'),
    ('725.1', PROSTITUTION, 'prostitution, bare'),
    ('725.1(4)', PROSTITUTION, 'and with the subsection the statute is in'),
    ('123.47(2)', PAULA, 'lower case is not a thing ICOS does, but neither is'),
    ('po/123.47', PAULA, 'assuming it never will be'),
]


@pytest.mark.parametrize('statute,sections,why', CITED)
def test_a_cited_section_says_yes(statute, sections, why):
    assert crs.cites_section(statute, sections) == "YES", why


NOT_CITED = [
    ('124.401', PAULA, 'controlled substances'),
    ('123.49', PAULA, 'chapter 123, but not one of the two sections'),
    ('123.460', PAULA, 'a longer section number that starts with 123.46'),
    ('123.46.1', PAULA, 'and one that continues past it with a dot'),
    ('725.10', PROSTITUTION, 'the reason 725.1 cannot be a prefix test'),
    ('725.1', PAULA, 'prostitution is not public intoxication'),
    ('123.46', PROSTITUTION, 'and the other way round'),
    ('1123.46', PAULA, 'a section that merely ends in the digits'),
    ('714.1;708.2', PAULA, 'a joined list with neither on it'),
]


@pytest.mark.parametrize('statute,sections,why', NOT_CITED)
def test_a_section_that_is_not_cited_says_no(statute, sections, why):
    assert crs.cites_section(statute, sections) == "NO", why


UNANSWERABLE = [
    ('', 'no adjudicated statute, because every count was dismissed'),
    (None, 'the same, before column F is written at all'),
    ('n/a', 'the civil placeholder'),
    ('N/A', 'and the way somebody else might have typed it'),
    (';', 'a joined list with nothing in it'),
    ('MA/62.01(120)-0198', 'a bare municipal section with no state section in it'),
]


@pytest.mark.parametrize('statutes_field,why', UNANSWERABLE)
def test_nothing_to_judge_from_gives_none(statutes_field, why):
    """The same three answer contract as is_vehicular, for the same reason.

    A bare municipal ordinance is a section number in one city's code. It may
    well mirror public intoxication and Napier cannot tell, so it does not say
    NO, which would be asserting something it does not know.
    """
    assert crs.cites_section(statutes_field, PAULA) is None, why


def test_any_count_carries_the_case():
    """Not just the one get_dominant_charge picked to speak for the row.

    The client has the conviction whichever count column G names.
    """
    assert crs.cites_section('321J.2;123.46;714.1', PAULA) == "YES"


# -- and that it reaches the cells ------------------------------------------

# Everything above would pass against a Napier that never wrote these columns,
# because it calls cites_section directly. These write a case into a worksheet
# the way a real run does.

PAULA_COLUMN = 'AJ'
PROSTITUTION_COLUMN = 'AK'


def _case(statutes_field, disposition='GUILTY BY PLEA'):
    return {
        'id': '00000  FECR000000',
        'county': 'SYNTHETIC',
        'charges': [{
            'charge': statutes_field,
            'description': 'SYNTHETIC OFFENSE',
            'disposition': [disposition],
            'offenseDate': '01/01/1900',
            'dispositionDate': '02/02/1901',
        }],
        'financials': [],
        'summary_categories': [],
    }


def _written(case):
    sheet = load_workbook(FULL)['CASE DATA']
    crs.process_case(case, sheet, crs.FIRST_CASE_ROW)
    row = str(crs.FIRST_CASE_ROW)
    return (sheet[PAULA_COLUMN + row].value,
            sheet[PROSTITUTION_COLUMN + row].value)


def test_the_real_paula_conviction_reaches_its_column():
    """The prostitution column is left blank rather than told NO.

    A match on an ordinance citation is evidence, because a city numbering its
    own code 123.47 and prosecuting somebody for supplying alcohol under it is
    not a coincidence worth entertaining. The absence of one is not evidence,
    because the ordinance may number prostitution anything at all. The sheet
    reads the blank as NO, which is the same answer with an honest reason.
    """
    assert _written(_case('PO/123.47(2)')) == ('YES', None)


def test_a_prostitution_conviction_reaches_its_column():
    assert _written(_case('725.1(4)')) == ('NO', 'YES')


def test_an_unrelated_conviction_answers_no_in_both():
    assert _written(_case('124.401(5)')) == ('NO', 'NO')


def test_a_case_with_no_adjudicated_statute_leaves_both_alone():
    """Every count dismissed. There is no conviction to expunge."""
    assert _written(_case('', disposition='DISMISSED')) == (None, None)


def test_a_civil_case_leaves_both_alone():
    """The civil branch writes "n/a" into column F, which is not a statute."""
    civil = {
        'id': '00000  SCSC000000',
        'county': 'SYNTHETIC',
        'charges': [],
        'summary_created_date': '01/01/1900',
        'summary_disposition_date': '02/02/1901',
        'summary_dispo_status': 'SYNTHETIC STATUS',
        'financials': [],
        'summary_categories': [],
    }
    assert _written(civil) == (None, None)


def test_these_columns_are_clear_of_the_stray_splitter_slot():
    """Row 9 of both templates carries a thirteenth splitter slot in AI.

    statutes._clear_strays blanks it, and only blanks a cell that still holds
    the array formula. A Napier output column on top of it would put a "NO" on
    row 9 of every workbook and leave the stray in place on the rest.
    """
    stray = statutes.LAST_COLUMN + 1
    for column in (PAULA_COLUMN, PROSTITUTION_COLUMN):
        index = load_workbook(FULL)['CASE DATA'][column + '4'].column
        assert index > stray, \
            '%s is inside the stray splitter slot' % column


# -- what the sheet does with them ------------------------------------------

# The full and Lite templates put these columns in different places, so both are
# found by their header rather than by a letter written down here.

HEADERS = {
    PAULA_COLUMN: 'PUB INTX',
    PROSTITUTION_COLUMN: '725.1?',
}


def _column_reading(path, needle):
    """The row 3 formula under the header naming this offence."""
    sheet = load_workbook(path)['EXPUNGEMENT & 910.7']
    for cell in sheet[2]:
        if isinstance(cell.value, str) and needle in cell.value:
            return sheet.cell(row=3, column=cell.column).value
    return None


@pytest.mark.parametrize('path,name', [(FULL, 'full'), (LITE, 'Lite')])
@pytest.mark.parametrize('column,needle', sorted(HEADERS.items()))
def test_the_sheet_reads_the_column_napier_writes(path, name, column, needle):
    formula = _column_reading(path, needle)
    assert formula is not None, \
        'the %s template lost the %s column' % (name, needle)
    assert "'CASE DATA'!%s4" % column in formula, formula


@pytest.mark.parametrize('path,name', [(FULL, 'full'), (LITE, 'Lite')])
@pytest.mark.parametrize('needle', sorted(HEADERS.values()))
def test_the_sheet_no_longer_compares_column_f_to_a_bare_section(path, name,
                                                                 needle):
    """The regression. This is the formula that was there.

    It is still a perfectly reasonable looking formula, which is why it survived
    however many releases, so what fails here is its return.
    """
    formula = _column_reading(path, needle)
    assert "'CASE DATA'!F" not in formula, \
        'the %s template is back to reading column F: %s' % (name, formula)


@pytest.mark.parametrize('path,name', [(FULL, 'full'), (LITE, 'Lite')])
@pytest.mark.parametrize('needle', sorted(HEADERS.values()))
def test_yes_is_the_string_the_sheet_is_testing_for(path, name, needle):
    """Why cites_section returns "YES" and not True or "Y"."""
    formula = _column_reading(path, needle)
    match = re.search(r"'CASE DATA'![A-Z]+\d+\s*=\s*\"([^\"]*)\"", formula)
    assert match, formula
    assert match.group(1) == "YES"


@pytest.mark.parametrize('path,name', [(FULL, 'full'), (LITE, 'Lite')])
@pytest.mark.parametrize('needle', sorted(HEADERS.values()))
def test_the_two_year_clock_still_hangs_off_the_answer(path, name, needle):
    """The column exists to start a clock, and the clock is the next column.

    A YES that nothing reads is the same as the NO it replaced.
    """
    sheet = load_workbook(path)['EXPUNGEMENT & 910.7']
    for cell in sheet[2]:
        if isinstance(cell.value, str) and needle in cell.value:
            clock = sheet.cell(row=3, column=cell.column + 1).value
            assert '730.5' in clock, clock
            return
    pytest.fail('no %s column in the %s template' % (needle, name))
