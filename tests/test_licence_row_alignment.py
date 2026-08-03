"""The licence sheet reads the case list out of order.

LICENSE-REGIS says, per case, whether the county treasurer can refuse to renew
a licence or a registration over what is still owed, and how much that is. Its
299 data rows read CASE DATA rows 4 to 302, which is every case row and none of
them twice, but not in that order. Sheet row 2 reads case row 6, row 3 reads 7,
row 4 reads 10. Case row 4, the first case in the workbook, is down at sheet
row 57.

Nothing is lost and every arithmetic answer on the sheet is right. What breaks
is reading it. A run of ten cases puts them on sheet rows 2, 3, 4, 5, 6, 22,
23, 24, 57, 58 with blank rows in between, and every blank row answers
"Z - no debt", because a blank case row owes nothing. So the top of the sheet
fills with cases that are not there saying there is no exposure, and the cases
that do carry a licence hold are somewhere below them. Only at 97 cases does
the shuffle close up and the sheet read straight through.

Both templates ship this way and no other sheet does. POLK R&B APPEAL, right
beside it, is sheet row = case row - 2 without exception. The "Z - " prefixes
on the labels in column C are what happened: they are there so that sorting the
column floats the exposed cases to the top, somebody sorted the sheet on a real
workbook, and saving it wrote the shuffled row references into the template
permanently.

The fix is a renumbering. Each of the four columns carries exactly one formula,
so putting row r back on case row r + 2 changes which case a row reads and
nothing else.

The cases here are synthetic. This repo is public and a real charges page is
one person's unredacted criminal record.
"""

import os
import re
import sys
from decimal import Decimal

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import case_parser
import tasks

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')
LITE = os.path.join(ROOT, 'CRS Lite 3.5.5.xlsx')
TEMPLATES = (FULL, LITE)

SHEET = 'LICENSE-REGIS'
FIRST_ROW, LAST_ROW = 2, 300
COLUMNS = ('A', 'B', 'C', 'D')

# The case row a sheet row is meant to read. Row 2 is the first case, which
# lands on CASE DATA row 4.
OFFSET = 2

# A relative reference to a case row. An absolute one is a count over the whole
# column and does not belong to any single row; the licence sheet has none, and
# the group is here so that one appearing is not silently read as a per-row
# lookup.
CASE_REF = re.compile(r"'CASE DATA'!(\$?)([A-Z]{1,3})(\$?)(\d+)")


def case_rows(formula):
    return {int(match.group(4)) for match in CASE_REF.finditer(formula)
            if not match.group(3)}


def formula(sheet, column, row):
    value = sheet['%s%d' % (column, row)].value
    return value if isinstance(value, str) and value.startswith('=') else None


# -- the shipped templates ---------------------------------------------------

def test_every_row_reads_the_case_that_belongs_on_it():
    for template in TEMPLATES:
        sheet = load_workbook(template)[SHEET]
        checked = 0
        wrong = []
        for row in range(FIRST_ROW, LAST_ROW + 1):
            for column in COLUMNS:
                text = formula(sheet, column, row)
                assert text, '%s %s%d has no formula' % (
                    os.path.basename(template), column, row)
                named = case_rows(text)
                checked += 1
                if named != {row + OFFSET}:
                    wrong.append('%s%d reads %s' % (column, row,
                                                    sorted(named)))
        assert checked == len(COLUMNS) * (LAST_ROW - FIRST_ROW + 1), checked
        assert not wrong, '%s: %d rows out of place, first few %s' % (
            os.path.basename(template), len(wrong), wrong[:6])


def test_no_row_reads_two_cases():
    """A renumbering that merged two rows would still pass the count."""
    for template in TEMPLATES:
        sheet = load_workbook(template)[SHEET]
        for row in range(FIRST_ROW, LAST_ROW + 1):
            for column in COLUMNS:
                named = case_rows(formula(sheet, column, row))
                assert len(named) == 1, '%s %s%d reads %s' % (
                    os.path.basename(template), column, row, sorted(named))


def test_the_same_case_rows_are_covered_as_before():
    """The sheet read case rows 4 to 302, once each, and still has to."""
    for template in TEMPLATES:
        sheet = load_workbook(template)[SHEET]
        seen = []
        for row in range(FIRST_ROW, LAST_ROW + 1):
            seen.extend(case_rows(formula(sheet, 'A', row)))
        assert sorted(seen) == list(range(4, 4 + LAST_ROW - FIRST_ROW + 1)), (
            os.path.basename(template), sorted(seen)[:8], sorted(seen)[-4:])


def test_the_sheet_beside_it_is_unchanged():
    """POLK R&B APPEAL is the model and has to stay the model."""
    for template in TEMPLATES:
        sheet = load_workbook(template)['POLK R&B APPEAL']
        for row in range(FIRST_ROW, 299 + 1):
            named = case_rows(formula(sheet, 'A', row))
            assert named == {row + OFFSET}, '%s A%d reads %s' % (
                os.path.basename(template), row, sorted(named))


def test_the_labels_that_explain_the_shuffle_are_still_there():
    """Column C's "Z - " prefixes sort the cases with no exposure to the
    bottom. Renumbering the rows does not remove the reason for sorting, so
    losing them would take away the thing the sheet is read with."""
    for template in TEMPLATES:
        text = formula(load_workbook(template)[SHEET], 'C', FIRST_ROW)
        assert '"Z - no debt"' in text, text
        assert '"Z - Neither license nor registration"' in text, text


# -- through the real build --------------------------------------------------

def charges_page(statute):
    def cells(*values):
        return '<tr>%s</tr>' % ''.join(
            '<td><font size="2">%s</font></td>' % value for value in values)
    return ''.join([
        '<html><body><table>',
        cells('Count 01', 'Original Charge'),
        cells('Charge:', statute, 'Description:', 'SYNTHETIC OFFENCE'),
        cells('Offense Date:', '01/01/1900', 'Arrest Date:', ''),
        cells('Adjudication'),
        cells('Charge:', statute, 'Description:', 'SYNTHETIC OFFENCE'),
        cells('Adjudication:', 'GUILTY', 'Adjudication Date:', '02/02/1901'),
        '</table></body></html>']).encode('utf-8')


def synthetic_cases(owed):
    """One case per amount, convicted, owing what it is given."""
    page = charges_page('321.285')
    cases = []
    for number, amount in enumerate(owed):
        case = {'id': '00000  FECR%06d' % number, 'county': 'SYNTHETIC',
                'financials': [], 'sentences': [],
                'summary_created_date': '01/01/1900',
                'summary_disposition_date': '', 'summary_dispo_status': ''}
        case_parser.parse_case_charges(page, case)
        due = Decimal(amount)
        case['summary_categories'] = [
            {'label': label,
             'original': due if label == 'FINE' else Decimal('0'),
             'paid': Decimal('0'),
             'due': due if label == 'FINE' else Decimal('0')}
            for label in ('COSTS', 'FINE', 'SURCHARGE', 'RESTITUTION', 'OTHER')]
        case['total_due'] = '$%s' % amount
        cases.append(case)
    return cases


def test_a_short_run_puts_its_cases_on_the_first_rows():
    """What a staffer sees. Four cases, the money on the first two, and the
    licence sheet has to open with them rather than with blank rows."""
    owed = ['197.43', '89.50', '0.00', '0.00']
    path, _, _ = tasks.build_workbook(
        synthetic_cases(owed), 'TEST CLIENT', '01/01/1980', False)
    try:
        workbook = load_workbook(path)
        cases = workbook['CASE DATA']
        sheet = workbook[SHEET]
        assert cases['A4'].value == '00000  FECR000000', cases['A4'].value
        for number in range(len(owed)):
            text = formula(sheet, 'D', FIRST_ROW + number)
            assert case_rows(text) == {4 + number}, (number, text)
    finally:
        os.remove(path)
