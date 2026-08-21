"""One client, more than one spelling, one workbook.

Iowa Courts matches a name exactly as it is written on the case, so a client
whose docket spells them "Alhameed" one year and "Al Hameed" the next is two
searches. Staff used to run both, get two workbooks and merge them by hand,
which is an hour and a client's whole debt riding on somebody's copy and paste.

The three things that can go wrong here and cannot be seen in a finished file:

  * a case listed under both spellings written twice, which doubles what the
    client owes on every sheet that sums a column;
  * a case asked for while the wrong spelling was the last thing Iowa Courts
    answered, which comes back a stub -- right heading, no charges, no money;
  * one spelling refusing and taking the other one's results with it.

The names are invented and every case number is 00000-shaped. The repository is
public and a real Iowa case number attached to a real name is a person.
"""

import os
import sys
import time

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ['NAPIER_DISABLE_BACKGROUND'] = '1'

import app as app_module
import jobs
import roster
import tasks
from icos import IcosError

# tasks.build_workbook is monkeypatched away for this whole file, so the
# tests that want a real CRS keep hold of it here.
real_build_workbook = tasks.build_workbook
from test_formula_grid import synthetic_cases   # noqa: E402

DOB = "01/01/1900"

# The shared case. It is on the docket under both spellings, which is the
# ordinary reason a name gets spelled two ways in the first place.
SHARED = '00000 FECR000000'
ONLY_JOINED = '00000 SRCR000000'
ONLY_SPACED = '00000 SMCR000000'


def results_for(surname, given, cases):
    rows = "".join(
        '<tr><td>%s</td><td></td><td>STATE VS %s</td><td>%s, %s</td>'
        '<td>%s</td><td>DEFENDANT</td></tr>' % (case, surname, surname, given,
                                                DOB)
        for case in cases)
    return ("<html><table>"
            "<tr><td>Case ID</td><td></td><td>Title</td><td>Name</td>"
            "<td>DOB</td><td>Role</td></tr>%s</table></html>" % rows).encode()


PEOPLE = {
    'ALHAMEED': results_for('ALHAMEED', 'ALI', [ONLY_JOINED, SHARED]),
    'AL HAMEED': results_for('AL HAMEED', 'ALI', [SHARED, ONLY_SPACED]),
    'DOE': results_for('DOE', 'JANE', ['00000 AGCR000000']),
}

JOINED_KEY = '1900-01-01 ALHAMEED, ALI'
SPACED_KEY = '1900-01-01 AL HAMEED, ALI'


class FakeClient:
    """Stands in for IcosClient and records what was asked in what order."""

    instances = []
    fail_searches = ()      # 1-based positions in the run that ICOS refuses
    fail_cases = ()

    def __init__(self, log=None, alert=None, **kwargs):
        self.log = log or (lambda m, **kw: None)
        self.alert = alert
        self.should_stop = lambda: False
        self.searched = []
        self.cases = []
        # The whole conversation in order. Which search preceded which case
        # pull is the only thing that separates a real case from a stub, and it
        # cannot be seen anywhere else.
        self.trace = []
        FakeClient.instances.append(self)

    def set_alert(self, alert):
        self.alert = alert

    def set_log(self, log):
        self.log = log or (lambda m, **kw: None)

    def set_stop_check(self, should_stop):
        self.should_stop = should_stop or (lambda: False)

    def login(self, username, password):
        self.log("Signed in to Iowa Courts Online.")

    def search(self, first, middle, last):
        self.searched.append((first, middle, last))
        self.trace.append(('search', last.upper()))
        if len(self.searched) in FakeClient.fail_searches:
            raise IcosError("Iowa Courts did not answer.")
        return PEOPLE.get(last.upper(), results_for(last.upper(), 'X', []))

    def case_bundle(self, case_id):
        self.cases.append(case_id)
        self.trace.append(('case', case_id))
        if case_id in FakeClient.fail_cases:
            raise IcosError("Iowa Courts did not answer for this case.")
        return b"<summary>", b"<charges>", b"<financials>"

    def logoff(self):
        pass


@pytest.fixture(autouse=True)
def fake_icos(monkeypatch):
    FakeClient.instances = []
    FakeClient.fail_searches = ()
    FakeClient.fail_cases = ()
    written.clear()
    attributed.clear()
    monkeypatch.setattr(tasks, 'IcosClient', FakeClient)
    monkeypatch.setattr(tasks.case_parser, 'parse_case_summary',
                        lambda html, case: case.update(county='DUBUQUE'))
    monkeypatch.setattr(tasks.case_parser, 'parse_case_charges',
                        lambda html, case: case.update(charges=[]))
    monkeypatch.setattr(tasks.case_parser, 'parse_case_financials',
                        lambda html, case: case.update(financials=[],
                                                       total_due='$0.00'))
    monkeypatch.setattr(tasks, 'build_workbook', record_workbook)
    yield


written = []
# The filed_as map each of those builds was handed.
attributed = []


def record_workbook(cases, name, dob, lite, failed=(), filed_as=None, no_dob=()):
    """What actually reached the workbook, which is the question the dedup
    test is asking. A case counted twice here is a client billed twice."""
    written.append([case['id'] for case in cases])
    attributed.append(dict(filed_as or {}))
    path = os.path.join(tasks.tmp_dir, 'test_alias_stub.xlsx')
    with open(path, 'wb') as handle:
        handle.write(b'PK\x03\x04 stub workbook')
    return path, {}, {'balance': '$0.00', 'monthly': None, 'months': 12}, []


@pytest.fixture
def client():
    app_module.app.config['TESTING'] = True
    app_module.app.secret_key = 'test'
    return app_module.app.test_client()


def await_job(client, job_id, timeout=5):
    deadline = time.time() + timeout
    while time.time() < deadline:
        state = client.get('/job/' + job_id).get_json()
        if state['done']:
            return state
        time.sleep(0.01)
    raise AssertionError("job never finished")


def search_for(client, spellings):
    """Post the search form with one name row per spelling."""
    form = {'username': 'ILATEST', 'password': 'secret',
            'firstname': [], 'middlename': [], 'lastname': []}
    for first, middle, last in spellings:
        form['firstname'].append(first)
        form['middlename'].append(middle)
        form['lastname'].append(last)
    response = client.post('/search', data=form)
    if response.status_code != 302:
        return None, response
    job_id = response.headers['Location'].rsplit('/', 1)[-1]
    return job_id, await_job(client, job_id)


BOTH_SPELLINGS = [('Ali', '', 'Alhameed'), ('Ali', '', 'Al Hameed')]


def build_from(client, spellings=BOTH_SPELLINGS, keys=None):
    """The whole errand: search every spelling, tick everyone, build."""
    search_id, _ = search_for(client, spellings)
    client.get('/results/' + search_id)
    result = jobs.get(search_id).result
    response = client.post('/crs-job',
                           json={'search_job_id': search_id,
                                 'keys': result['keys'] if keys is None
                                 else keys})
    job_id = response.get_json()['job_id']
    return search_id, job_id, await_job(client, job_id)


class TestSearchingEverySpelling:
    def test_both_spellings_are_searched_on_one_sign_in(self, client):
        """The point of the feature. Two spellings used to be two sign ins
        queueing for the same shared Iowa Courts account."""
        search_id, _ = search_for(client, BOTH_SPELLINGS)
        assert len(FakeClient.instances) == 1
        assert FakeClient.instances[0].searched == [
            ('Ali', '', 'Alhameed'), ('Ali', '', 'Al Hameed')]

    def test_the_results_are_one_list_to_pick_from(self, client):
        search_id, _ = search_for(client, BOTH_SPELLINGS)
        result = jobs.get(search_id).result
        assert sorted(result['keys']) == [SPACED_KEY, JOINED_KEY]
        assert result['found_by'] == {JOINED_KEY: 0, SPACED_KEY: 1}

    def test_the_page_says_which_spelling_found_which_person(self, client):
        search_id, _ = search_for(client, BOTH_SPELLINGS)
        page = client.get('/results/' + search_id).get_data(as_text=True)
        assert 'found as Ali Alhameed' in page
        assert 'found as Ali Al Hameed' in page

    def test_one_spelling_looks_exactly_as_it_always_did(self, client):
        """Every ordinary search is this one. It must not grow a spellings
        block, and it must not spend a second request on ICOS."""
        search_id, _ = search_for(client, [('Jane', '', 'Doe')])
        page = client.get('/results/' + search_id).get_data(as_text=True)
        assert 'found as' not in page
        assert FakeClient.instances[0].searched == [('Jane', '', 'Doe')]

    def test_the_same_spelling_typed_twice_is_searched_once(self, client):
        """A second row filled in with what is already in the first holds the
        shared account for an answer Napier already has."""
        search_for(client, [('Ali', '', 'Alhameed'), ('ali', '', 'ALHAMEED')])
        assert len(FakeClient.instances[0].searched) == 1

    def test_an_empty_extra_row_is_not_a_search(self, client):
        search_for(client, [('Ali', '', 'Alhameed'), ('', '', '')])
        assert len(FakeClient.instances[0].searched) == 1

    def test_no_name_at_all_is_refused_before_anyone_signs_in(self, client):
        _, response = search_for(client, [('Ali', '', '')])
        assert response.status_code == 200
        assert 'Enter a last name' in response.get_data(as_text=True)
        assert FakeClient.instances == []

    def test_more_spellings_than_the_cap_are_refused(self, client):
        many = [('Ali', '', 'Spelling%d' % n)
                for n in range(app_module.MAX_SPELLINGS + 1)]
        _, response = search_for(client, many)
        assert response.status_code == 200
        assert 'spellings of one name' in response.get_data(as_text=True)
        assert FakeClient.instances == []

    def test_no_name_reaches_the_progress_log(self, client):
        """The log is quoted into alert email and a client's name is
        privileged. Which of two spellings is being searched is not."""
        _, state = search_for(client, BOTH_SPELLINGS)
        joined = " ".join(state['progress']).upper()
        assert 'ALI' not in joined
        assert 'HAMEED' not in joined
        assert 'spelling 1 of 2' in " ".join(state['progress'])


class TestOneSpellingRefusing:
    def test_the_other_spelling_still_answers(self, client):
        FakeClient.fail_searches = (1,)
        search_id, state = search_for(client, BOTH_SPELLINGS)
        assert state['status'] == 'done'
        result = jobs.get(search_id).result
        assert result['keys'] == [SPACED_KEY]
        assert result['searches'][0]['error']
        assert result['searches'][1]['keys']

    def test_the_page_says_so_rather_than_showing_a_short_list(self, client):
        """A list one spelling short reads exactly like a client who only ever
        had those cases, and nothing else on the page contradicts it."""
        FakeClient.fail_searches = (1,)
        search_id, _ = search_for(client, BOTH_SPELLINGS)
        page = client.get('/results/' + search_id).get_data(as_text=True)
        assert 'would not answer this one' in page

    def test_every_spelling_refusing_is_still_a_failed_run(self, client):
        """What a single refused search has always been. A page that offers
        nobody to pick would be the same page as a client with no cases."""
        FakeClient.fail_searches = (1, 2)
        _, state = search_for(client, BOTH_SPELLINGS)
        assert state['status'] == 'failed'


class TestPullingWhatWasPicked:
    def test_a_case_under_both_spellings_is_written_once(self, client):
        """The one that doubles a client's debt. It is on the docket under both
        names, so it comes back in both searches under two defendant keys, and
        a staffer ticking both is doing exactly what the page tells them to."""
        _, job_id, state = build_from(client)
        assert state['status'] == 'done'
        assert written[-1].count(SHARED) == 1
        assert sorted(written[-1]) == sorted([SHARED, ONLY_JOINED,
                                              ONLY_SPACED])

    def test_it_is_only_asked_for_once_as_well(self, client):
        _, job_id, _ = build_from(client)
        assert FakeClient.instances[0].cases.count(SHARED) == 1

    def test_each_spelling_is_put_back_in_front_of_icos_first(self, client):
        """ICOS answers a case request out of the last result set it produced.
        A case asked for under the wrong spelling comes back a stub: right
        heading, no charges, no money, and it fails every validator for the
        full four minute case budget before anyone finds out."""
        _, job_id, _ = build_from(client)
        trace = FakeClient.instances[0].trace
        standing = None
        for kind, what in trace:
            if kind == 'search':
                standing = what
                continue
            assert standing is not None
            found_under = ('ALHAMEED' if what == ONLY_JOINED else
                           'AL HAMEED' if what == ONLY_SPACED else standing)
            assert standing == found_under, (
                "%s was asked for while ICOS was standing on %s"
                % (what, standing))

    def test_one_spelling_does_not_pay_for_a_re_search(self, client):
        """The ordinary run. Its search is already the last thing ICOS
        answered, so repeating it would spend a request and a turn with the
        shared account on every search anyone ever does."""
        build_from(client, [('Jane', '', 'Doe')])
        assert FakeClient.instances[0].trace == [
            ('search', 'DOE'), ('case', '00000 AGCR000000')]

    def test_picking_one_spelling_only_searches_that_one(self, client):
        """Two people came back and the staffer says only one of them is the
        client. The other spelling's search is then work nobody asked for."""
        _, job_id, _ = build_from(client, keys=[JOINED_KEY])
        pulls = [what for kind, what in FakeClient.instances[0].trace
                 if kind == 'case']
        assert sorted(pulls) == sorted([ONLY_JOINED, SHARED])
        assert ('search', 'AL HAMEED') not in \
            FakeClient.instances[0].trace[2:]


class TestFinishingAShortRun:
    def test_the_retry_carries_the_spelling_behind_each_missing_case(self, client):
        """A retry signs in fresh, so nothing is standing in front of ICOS at
        all. Re-selecting the first spelling and then asking for a case the
        second one found is the same stub, on the staffer's second try."""
        FakeClient.fail_cases = (ONLY_SPACED,)
        _, job_id, _ = build_from(client)
        payload = jobs.get(job_id).result['retry']
        assert payload is not None

        FakeClient.fail_cases = ()
        response = client.post('/retry/' + job_id,
                               data={'username': 'ILATEST',
                                     'password': 'secret'})
        retry_id = response.headers['Location'].rsplit('/', 1)[-1]
        await_job(client, retry_id)

        retry_client = FakeClient.instances[-1]
        assert retry_client.trace == [('search', 'AL HAMEED'),
                                      ('case', ONLY_SPACED)]
        assert written[-1].count(SHARED) == 1
        assert sorted(written[-1]) == sorted([SHARED, ONLY_JOINED,
                                              ONLY_SPACED])


class TestTheRosterReadsAnAka:
    def test_one_line_can_carry_a_second_spelling(self):
        people, rejected = roster.parse("Al Hameed, Ali aka Alhameed, Ali")
        assert rejected == []
        assert len(people) == 1
        assert roster.spellings(people[0]) == [
            {'first': 'Ali', 'middle': '', 'last': 'Al Hameed'},
            {'first': 'Ali', 'middle': '', 'last': 'Alhameed'}]

    @pytest.mark.parametrize('marker', ['aka', 'AKA', 'a/k/a', 'A.K.A.',
                                        'also known as'])
    def test_however_the_clinic_writes_it(self, marker):
        people, _ = roster.parse("Al Hameed, Ali %s Alhameed, Ali" % marker)
        assert len(roster.spellings(people[0])) == 2

    def test_a_line_with_no_aka_is_what_it_always_was(self):
        people, _ = roster.parse("Doe, Jane")
        assert roster.spellings(people[0]) == [
            {'first': 'Jane', 'middle': '', 'last': 'Doe'}]
        assert people[0].get('aliases') == []

    def test_a_spelling_that_repeats_the_primary_is_dropped(self):
        people, _ = roster.parse("Doe, Jane aka Doe, Jane")
        assert len(roster.spellings(people[0])) == 1

    def test_the_roster_page_says_what_it_will_search(self):
        people, _ = roster.parse("Al Hameed, Ali aka Alhameed, Ali")
        assert roster.describe(people[0]) == "Ali Al Hameed (also Ali Alhameed)"

    def test_the_cap_counts_searches_not_clients(self):
        people, _ = roster.parse("Doe, Jane aka Roe, Jane\nSmith, Sam")
        assert len(people) == 2
        assert roster.searches_count(people) == 3

    def test_two_clients_sharing_an_aka_are_still_two_clients(self):
        """Folding duplicates is about not searching one client twice. Two
        people who happen to have gone by the same name are not one person."""
        people, _ = roster.parse("Doe, Jane aka Roe, Pat\n"
                                 "Smith, Sam aka Roe, Pat")
        assert len(people) == 2


class TestAClinicListWithAnAka:
    def run_list(self, client, text):
        response = client.post('/batch', data={
            'username': 'ILATEST', 'password': 'secret', 'roster': text})
        search_id = response.headers['Location'].rsplit('/', 1)[-1]
        await_job(client, search_id)
        client.get('/roster/' + search_id)
        result = jobs.get(search_id).result
        picks = [{'client': index, 'keys': entry['keys']}
                 for index, entry in enumerate(result['clients'])
                 if entry['keys']]
        started = client.post('/batch-crs',
                              json={'search_job_id': search_id,
                                    'picks': picks}).get_json()
        return search_id, await_job(client, started['job_id'])

    def test_both_spellings_are_searched_for_the_one_client(self, client):
        search_id, _ = self.run_list(client,
                                     "Alhameed, Ali aka Al Hameed, Ali")
        # The first two. The build reuses the same pooled session and puts
        # each spelling back in front of ICOS, so it searches them again.
        assert FakeClient.instances[0].searched[:2] == [
            ('Ali', '', 'Alhameed'), ('Ali', '', 'Al Hameed')]
        entry = jobs.get(search_id).result['clients'][0]
        assert sorted(entry['keys']) == [SPACED_KEY, JOINED_KEY]

    def test_the_client_gets_one_workbook_with_the_case_in_it_once(self, client):
        self.run_list(client, "Alhameed, Ali aka Al Hameed, Ali")
        assert len(written) == 1
        assert sorted(written[0]) == sorted([SHARED, ONLY_JOINED, ONLY_SPACED])

    def test_each_spelling_is_re_selected_before_its_own_cases(self, client):
        self.run_list(client, "Alhameed, Ali aka Al Hameed, Ali")
        trace = FakeClient.instances[0].trace
        pulls = trace[trace.index(('search', 'ALHAMEED'), 2):]
        standing = None
        for kind, what in pulls:
            if kind == 'search':
                standing = what
            elif what == ONLY_JOINED:
                assert standing == 'ALHAMEED'
            elif what == ONLY_SPACED:
                assert standing == 'AL HAMEED'

    def test_a_spelling_that_refuses_does_not_lose_the_client(self, client):
        """The second search of the two. The client's other spelling answered
        and those cases are real, so this is not a client with no matches."""
        FakeClient.fail_searches = (2,)
        search_id, _ = self.run_list(client,
                                     "Alhameed, Ali aka Al Hameed, Ali")
        entry = jobs.get(search_id).result['clients'][0]
        assert entry['keys'] == [JOINED_KEY]
        assert entry['error'] is None


class TestWhichSpellingARowCameFrom:
    """Iowa Legal Aid asked for this on 2026-08-18.

    Two spellings merge into one workbook, and the finished rows do not say
    which of them each case came from. CASE DATA has no name column at all: the
    client is named once, on BASIC INFO, and def_name there is whichever key
    sorted first. So a staffer checking a merged summary against the person in
    front of them has no way to tell the two dockets apart.
    """

    def test_a_case_says_which_spelling_it_is_docketed_under(self, client):
        build_from(client)
        assert attributed[-1][ONLY_JOINED] == 'ALHAMEED, ALI'
        assert attributed[-1][ONLY_SPACED] == 'AL HAMEED, ALI'

    def test_a_case_on_both_dockets_is_attributed_once_to_one_of_them(self,
                                                                     client):
        """It is one row. Naming both spellings on it would be truthful and
        would also be the row nobody can act on."""
        build_from(client)
        assert attributed[-1][SHARED] in ('ALHAMEED, ALI', 'AL HAMEED, ALI')
        assert len(written[-1]) == len(attributed[-1])

    def test_one_spelling_attributes_nothing(self, client):
        """Every ordinary run. A note repeated on every row of a clinic sheet
        is a column staff learn to skip, and the coding caveats are in it."""
        build_from(client, spellings=[('Jane', '', 'Doe')])
        assert attributed[-1] == {}

    def test_picking_only_one_of_two_spellings_attributes_nothing(self,
                                                                  client):
        """Nothing merged, so there is nothing to tell apart."""
        build_from(client, keys=[JOINED_KEY])
        assert attributed[-1] == {}

    def test_a_clinic_list_attributes_the_merged_client(self, client):
        """Where it is least visible. The finish page shows one row per client
        and names them by the first key, so the second spelling appears
        nowhere else at all."""
        TestAClinicListWithAnAka().run_list(
            client, "Alhameed, Ali aka Al Hameed, Ali")
        assert attributed[-1][ONLY_SPACED] == 'AL HAMEED, ALI'

    def test_the_note_survives_a_rebuild(self, client):
        """A retry rebuilds the workbook from the retry payload and never goes
        back to the search job, which is the only thing that knew this."""
        FakeClient.fail_cases = (ONLY_SPACED,)
        search_id, job_id, _ = build_from(client)
        FakeClient.fail_cases = ()
        response = client.post('/retry/' + job_id,
                               data={'username': 'ILATEST',
                                     'password': 'secret'})
        await_job(client, response.headers['Location'].rsplit('/', 1)[-1])
        assert attributed[-1][ONLY_SPACED] == 'AL HAMEED, ALI'


class TestTheNoteInTheWorkbook:
    """The note itself, written into a real CRS rather than a stub."""

    def _built(self, filed_as):
        from openpyxl import load_workbook
        path, _, _, _ = real_build_workbook(
            synthetic_cases(2), 'ALHAMEED, ALI', DOB, False,
            filed_as=filed_as)
        return load_workbook(path)['CASE DATA']

    def test_it_lands_in_the_notes_column(self):
        sheet = self._built({'00000  FECR000000': 'ALHAMEED, ALI',
                             '00000  FECR000001': 'AL HAMEED, ALI'})
        assert 'ALHAMEED, ALI' in sheet['V4'].value
        assert 'AL HAMEED, ALI' in sheet['V5'].value

    def test_nothing_is_written_without_it(self):
        # Column V is not empty on an ordinary row -- process_financials has
        # its own say about the fee columns -- so this is about the one
        # sentence, not about the cell.
        sheet = self._built({})
        assert 'Filed under' not in (sheet['V4'].value or '')

    def test_it_does_not_displace_a_coding_caveat(self):
        """Column V already carries what process_financials and the coding
        guesses put there, and append_note joins rather than replaces."""
        import crs
        cases = synthetic_cases(1)
        path, _, _, _ = real_build_workbook(
            cases, 'ALHAMEED, ALI', DOB, False,
            filed_as={'00000  FECR000000': 'ALHAMEED, ALI'})
        from openpyxl import load_workbook
        sheet = load_workbook(path)['CASE DATA']
        crs.append_note(sheet, 4, 'A later caveat.')
        assert 'ALHAMEED, ALI' in sheet['V4'].value
        assert 'A later caveat.' in sheet['V4'].value
