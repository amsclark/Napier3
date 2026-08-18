"""An acquittal is sorted as though it were a conviction.

BANKRUPTCY splits a case's debt four ways by how likely a bankruptcy is to
discharge it, and EXEMPTIONS splits it two ways by which exemptions reach it.
Both start from one question: was there a conviction here? Both ask it the same
way, OR(B="DISM", B="JWV", B="JUV", B="CIV"), and treat everything else as a
conviction.

Napier writes four more codes that are not convictions. ACQ is an acquittal,
WTHD a withdrawn charge, NOTF a charge never filed, TNSF a case transferred
away. The expungement sheet in the same workbook knows all four: its DISMISSED?
column asks OR(G="DISM", G="ACQ", G="NOTF", G="WTHD", G="TNSF"). The two money
sheets were given the shorter list, so a client acquitted at trial has whatever
the case still carries moved out of the fully dischargeable column into the
narrower ones, and out of "all exemptions apply" into "federal only".

Nothing is lost either way. C+D+E+F is the whole of J:S on bankruptcy and C+D
is the whole of it on exemptions, so the case total is right in both readings.
What changes is the answer a staffer reads off the sheet and repeats to a
client about debt they may not have to pay.

Across 300 real cases the money that moves is $0.00, and that is the reason
this shipped. Iowa taxes court costs on conviction, so an acquitted case
usually owes nothing and the shorter list costs nothing. Usually is not
always: costs can survive an acquittal, a withdrawn charge can leave a filing
fee, and 11 U.S.C. 523(a)(7) excepts a fine from discharge because it is a
penalty, which is a thing an acquittal does not produce. There is no reading
under which a dismissal is fully dischargeable and an acquittal is not.

Two codes are deliberately left out of the fix. OTH is the code Napier gives a
disposition it could not read, so it means unknown, and treating unknown as no
conviction would tell a client debt is dischargeable on the strength of a word
nobody understood. An undisposed case has no code at all and column B calls it
"open charge"; it carries no conviction yet, but it is not cleared either, and
which column an open case belongs in is a question for the clinic rather than
for this file. Both are in OPEN_QUESTIONS.md.

The cases here are synthetic. This repo is public and a real charges page is
one person's unredacted criminal record.
"""

import os
import re
import sys
from decimal import Decimal

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import case_parser
import crs
import tasks

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')
LITE = os.path.join(ROOT, 'CRS Lite 3.5.5.xlsx')

# Sheet, its columns that sort debt by the disposition, and its case rows.
# Neither sheet exists in the Lite template, which is checked below rather
# than assumed.
SORTERS = (
    ('BANKRUPTCY', ('C', 'D', 'E', 'F'), 4, 201),
    ('EXEMPTIONS', ('C', 'D'), 4, 200),
)

# Every code Napier can put in column G that does not mean an adult criminal
# conviction, and so every code these two sheets have to clear. JUV is a
# juvenile adjudication, which is not a criminal conviction; CIV is a civil
# matter; JWV was waived to adult court, so the disposition is on another case.
NO_CONVICTION = frozenset({'DISM', 'JWV', 'JUV', 'CIV',
                           'ACQ', 'WTHD', 'NOTF', 'TNSF'})

# What must never be cleared. GTR, GPL and DEF are convictions. OTH means the
# disposition could not be read, and guessing in the client's favour is the one
# guess this sheet must not make.
CONVICTIONS = ('GTR', 'GPL', 'DEF', 'OTH')

TESTED_CODE = re.compile(r'B(\d+)\s*=\s*"([A-Z]+)"')


def formula(sheet, column, row):
    value = sheet['%s%d' % (column, row)].value
    return value if isinstance(value, str) and value.startswith('=') else None


def cleared_by(text):
    """The codes a row's formula treats as no conviction, and the rows it
    asked about. A test against a neighbouring row would still look like a
    list of codes, so the rows come back too."""
    found = TESTED_CODE.findall(text)
    return {code for _, code in found}, {int(row) for row, _ in found}


# -- the shipped template -----------------------------------------------------

def test_every_row_clears_every_code_that_is_not_a_conviction():
    workbook = load_workbook(FULL)
    for title, columns, first, last in SORTERS:
        sheet = workbook[title]
        wrong = []
        for row in range(first, last + 1):
            for column in columns:
                text = formula(sheet, column, row)
                assert text, '%s %s%d has no formula' % (title, column, row)
                codes, _ = cleared_by(text)
                if codes != NO_CONVICTION:
                    wrong.append('%s%d misses %s' % (
                        column, row, sorted(NO_CONVICTION - codes) or
                        'nothing, but adds %s' % sorted(codes - NO_CONVICTION)))
        assert not wrong, '%s: %d cells, first few %s' % (
            title, len(wrong), wrong[:4])


def test_a_conviction_is_never_cleared():
    workbook = load_workbook(FULL)
    for title, columns, first, last in SORTERS:
        sheet = workbook[title]
        for row in range(first, last + 1):
            for column in columns:
                codes, _ = cleared_by(formula(sheet, column, row))
                for code in CONVICTIONS:
                    assert code not in codes, '%s %s%d clears %s' % (
                        title, column, row, code)


def test_a_row_only_asks_about_its_own_case():
    """The list is rewritten a cell at a time, and a substitution that dropped
    a row number would still leave a plausible looking list of codes."""
    workbook = load_workbook(FULL)
    for title, columns, first, last in SORTERS:
        sheet = workbook[title]
        for row in range(first, last + 1):
            for column in columns:
                _, rows = cleared_by(formula(sheet, column, row))
                assert rows == {row}, '%s %s%d asks about %s' % (
                    title, column, row, sorted(rows))


def test_the_two_sheets_agree_with_each_other():
    """They answer different questions off the same one. A fix applied to one
    and not the other is worse than neither."""
    workbook = load_workbook(FULL)
    seen = set()
    for title, columns, first, _ in SORTERS:
        for column in columns:
            codes, _ = cleared_by(formula(workbook[title], column, first))
            seen.add(frozenset(codes))
    assert len(seen) == 1, sorted(sorted(group) for group in seen)


def test_the_money_still_lands_somewhere():
    """Whatever the sorting says, every column of CASE DATA money has to come
    out on this sheet exactly once. A widened test that also changed which
    columns are summed would move debt off the workbook."""
    workbook = load_workbook(FULL)
    expected = {
        'BANKRUPTCY': ["SUM('CASE DATA'!J4:S4)", "'CASE DATA'!J4+'CASE DATA'!K4",
                       "SUM('CASE DATA'!L4:P4)", "'CASE DATA'!Q4",
                       "'CASE DATA'!R4+'CASE DATA'!S4"],
        'EXEMPTIONS': ["SUM('CASE DATA'!J4:S4)",
                       "SUM('CASE DATA'!J4:P4)+'CASE DATA'!S4",
                       "'CASE DATA'!Q4+'CASE DATA'!R4"],
    }
    for title, columns, first, _ in SORTERS:
        sheet = workbook[title]
        joined = ' '.join(formula(sheet, column, first) for column in columns)
        for piece in expected[title]:
            assert piece in joined, '%s lost %s' % (title, piece)


def test_the_lite_template_has_neither_sheet():
    """The reason this fix touches one file. If a Lite workbook ever grows a
    bankruptcy sheet it has to be fixed too, and this is what will say so."""
    names = load_workbook(LITE).sheetnames
    for title, _, _, _ in SORTERS:
        assert title not in names, names


# -- the note the workbook carries -------------------------------------------

def test_the_note_does_not_promise_what_the_sheets_do_not_do():
    """Column V tells the staffer what an unreadable disposition costs them.
    It used to say the bankruptcy and exemption sheets treat OTH as no
    conviction, and they do the opposite: they sort its debt as a conviction's.
    Somebody reading that note and trusting it tells a client the wrong thing
    about money."""
    note = crs.UNKNOWN_DISPOSITION_NOTE
    workbook = load_workbook(FULL)
    for title, columns, first, _ in SORTERS:
        codes, _ = cleared_by(formula(workbook[title], columns[0], first))
        assert 'OTH' not in codes, title
    lowered = note.lower()
    assert 'bankruptcy' in lowered, note
    assert "as a conviction's" in lowered, note
    assert ('bankruptcy, exemption and licence sheets treat oth as no '
            'conviction') not in lowered, note


# -- through the real build --------------------------------------------------

def charges_page(adjudication):
    def cells(*values):
        return '<tr>%s</tr>' % ''.join(
            '<td><font size="2">%s</font></td>' % value for value in values)
    return ''.join([
        '<html><body><table>',
        cells('Count 01', 'Original Charge'),
        cells('Charge:', '321.285', 'Description:', 'SYNTHETIC OFFENCE'),
        cells('Offense Date:', '01/01/1900', 'Arrest Date:', ''),
        cells('Adjudication'),
        cells('Charge:', '321.285', 'Description:', 'SYNTHETIC OFFENCE'),
        cells('Adjudication:', adjudication, 'Adjudication Date:', '02/02/1901'),
        '</table></body></html>']).encode('utf-8')


def synthetic_cases(adjudications, amount='197.43'):
    """One case per disposition, every one of them owing the same money."""
    cases = []
    for number, adjudication in enumerate(adjudications):
        case = {'id': '00000  FECR%06d' % number, 'county': 'SYNTHETIC',
                'financials': [], 'sentences': [],
                'summary_created_date': '01/01/1900',
                'summary_disposition_date': '02/02/1901',
                'summary_dispo_status': ''}
        case_parser.parse_case_charges(charges_page(adjudication), case)
        # Money in every bucket, because the four bankruptcy columns read
        # different stretches of J to S and a case owing one kind of debt
        # only exercises one of them.
        due = Decimal(amount)
        case['summary_categories'] = [
            {'label': label, 'original': due, 'paid': Decimal('0'),
             'due': due}
            for label in ('COSTS', 'FINE', 'SURCHARGE', 'RESTITUTION', 'OTHER')]
        case['total_due'] = '$%s' % (due * 5)
        cases.append(case)
    return cases


def test_an_acquittal_is_sorted_like_a_dismissal():
    """What a staffer sees, read the way the sheet reads it.

    Comparing the two rows' formulas proves nothing: a column carries one
    formula and the two rows differ only by row number whether the fix is in or
    not. What decides the answer is whether the code the build actually put in
    column G is in the list that formula clears. So this takes the code off
    CASE DATA and looks for it, which is the test the spreadsheet runs.
    """
    dispositions = ['DISMISSED', 'ACQUITTED', 'WITHDRAWN', 'NOT FILED',
                    'GUILTY']
    path, _, _, _ = tasks.build_workbook(
        synthetic_cases(dispositions), 'TEST CLIENT', '01/01/1980', False)
    try:
        workbook = load_workbook(path)
        cases = workbook['CASE DATA']
        written = [cases['G%d' % (4 + n)].value
                   for n in range(len(dispositions))]
        assert written == ['DISM', 'ACQ', 'WTHD', 'NOTF', 'GTR'], written
        # A case owing nothing comes out of every column as zero and would
        # pass this whatever the sorting says, so the money is checked before
        # the sorting is.
        for number in range(len(dispositions)):
            owed = [cases.cell(row=4 + number, column=column).value or 0
                    for column in range(10, 20)]          # J to S
            assert len([amount for amount in owed if amount]) >= 3, owed

        for title, columns, first, _ in SORTERS:
            sheet = workbook[title]
            for number, code in enumerate(written):
                row = first + number
                for column in columns:
                    codes, _ = cleared_by(formula(sheet, column, row))
                    cleared = code in codes
                    assert cleared == (code != 'GTR'), (
                        '%s %s%d: %s is %s' % (title, column, row, code,
                                               'cleared' if cleared
                                               else 'sorted as a conviction'))
    finally:
        os.remove(path)


def test_the_codes_the_sheets_clear_are_codes_napier_writes():
    """A list that clears a code Napier never produces is a list that has
    stopped describing this program.

    TNSF was one of those until a real run on 3 August 2026 met a charge
    disposed CHANGE OF VENUE, so every code these sheets clear is now a code
    something can actually put in column G.
    """
    written = set()
    for mapping in crs.charge_code_map.values():
        written.update(mapping)
    unbacked = NO_CONVICTION - written
    assert not unbacked, sorted(unbacked)
    assert 'TNSF' in crs.charge_code_map['CHANGE OF VENUE']
