"""Column I, and the ICOS table Napier downloaded for years and never opened.

CASE DATA column I is "Under supervison?" [sic]. The expungement sheet reads it
to decide whether to fill in "Amount of debt subject to 910.7?", and reads a
blank as no, so that column has been "n/a" on every row of every workbook Napier
has ever produced.

The answer was on a page Napier already fetches. Under each count the charges
page carries a sentence table: the sentence type, the date it was imposed and
how long it runs. Across 180 real captures it appears on every case that carries
a conviction. case_parser marked the section boundary and read nothing out of it.

What Napier can say from that is narrow, and these tests pin the narrowness as
hard as they pin the answer. A term that ends after the clinic date is a YES. No
term, an expired term, or a term with no end date is left blank rather than
called a NO, because ICOS does not record an early discharge, records extensions
inconsistently, and never carries parole at all.
"""

import os
import sys
from datetime import date

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import case_parser
import crs

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')

CLINIC = date(2026, 7, 31)


def _sentence(kind, day, duration):
    return {'type': kind, 'date': day, 'duration': duration}


# -- the term arithmetic ----------------------------------------------------

def test_a_year_term_is_a_calendar_year_not_365_days():
    """A leap year in the middle of the term is the case that separates them."""
    assert crs._add_term(date(2024, 1, 15), 1, 'Year') == date(2025, 1, 15)


def test_a_five_year_term_lands_on_the_same_day_five_years_later():
    assert crs._add_term(date(2001, 7, 2), 5, 'Year') == date(2006, 7, 2)


def test_a_term_starting_on_a_leap_day_falls_back_to_the_28th():
    assert crs._add_term(date(2024, 2, 29), 1, 'Year') == date(2025, 2, 28)


def test_a_term_starting_on_the_31st_falls_back_to_a_short_month():
    assert crs._add_term(date(2026, 1, 31), 1, 'Month') == date(2026, 2, 28)


def test_day_terms_are_counted_in_days():
    assert crs._add_term(date(1995, 3, 31), 365, 'Day') == date(1996, 3, 30)


def test_week_terms_are_counted_in_weeks():
    assert crs._add_term(date(2026, 1, 1), 2, 'Week') == date(2026, 1, 15)


# -- reading a date the way ICOS writes it ----------------------------------

@pytest.mark.parametrize('text', ['07/02/2001', ' 07/02/2001 '])
def test_an_icos_date_is_read(text):
    assert crs.parse_us_date(text) == date(2001, 7, 2)


@pytest.mark.parametrize('text', ['', None, 'n/a', '2001-07-02', '13/45/2001'])
def test_anything_else_is_not_guessed_at(text):
    assert crs.parse_us_date(text) is None


# -- which sentences count --------------------------------------------------

def test_probation_still_running_is_a_yes():
    answer, term = crs.is_under_supervision(
        [_sentence('PROBATION', '01/01/2025', '3 Year(s)')], CLINIC)
    assert answer == "YES"
    assert term[3] == date(2028, 1, 1)


def test_probation_that_has_run_out_is_left_blank():
    answer, _ = crs.is_under_supervision(
        [_sentence('PROBATION', '01/01/1995', '2 Year(s)')], CLINIC)
    assert answer is None


def test_a_term_ending_on_the_clinic_date_still_counts():
    """The last day of a term is a day under supervision."""
    answer, _ = crs.is_under_supervision(
        [_sentence('PROBATION', '07/31/2024', '2 Year(s)')], CLINIC)
    assert answer == "YES"


def test_prison_is_not_supervision():
    """Somebody in custody is not what the 910.7 column is asking about, and
    the day they get out is not on this page."""
    answer, _ = crs.is_under_supervision(
        [_sentence('PRISON', '01/01/2025', '5 Year(s)')], CLINIC)
    assert answer is None


@pytest.mark.parametrize('kind', ['FINE', 'JAIL', 'DISMISSED', 'TIME SERVED',
                                  'SUSPENDED PRISON', 'COMMUNITY SERVICE'])
def test_the_other_real_sentence_types_are_not_supervision(kind):
    answer, _ = crs.is_under_supervision(
        [_sentence(kind, '01/01/2025', '5 Year(s)')], CLINIC)
    assert answer is None


@pytest.mark.parametrize('kind', ['PROBATION', 'PROBATION - OTHER THAN DCS',
                                  'PROBATION EXTENDED', 'DRUG COURT',
                                  'RESIDENTIAL FACILITY'])
def test_every_community_supervision_type_counts(kind):
    answer, _ = crs.is_under_supervision(
        [_sentence(kind, '01/01/2025', '3 Year(s)')], CLINIC)
    assert answer == "YES"


def test_probation_someone_other_than_dcs_holds_is_still_probation():
    """Two of the 300 captured cases are people on probation right now, and
    both of them are this wording. 910.7 is petitionable during the period of
    probation, and it does not ask who administers it."""
    answer, term = crs.is_under_supervision(
        [_sentence('PROBATION - OTHER THAN DCS', '12/22/2025', '2 Year(s)')],
        CLINIC)
    assert answer == "YES"
    assert term[3] == date(2027, 12, 22)


def test_an_extension_can_be_the_row_that_keeps_the_term_running():
    """The one place the wording matters on its own. The original term is spent
    and the extension is not, so reading only PROBATION says the person is off
    supervision on the day they are not."""
    sentences = [_sentence('PROBATION', '01/01/2023', '2 Year(s)'),
                 _sentence('PROBATION EXTENDED', '01/01/2025', '3 Year(s)')]
    answer, term = crs.is_under_supervision(sentences, CLINIC)
    assert answer == "YES"
    assert term[0] == 'PROBATION EXTENDED'
    assert term[3] == date(2028, 1, 1)


def test_icos_saying_there_is_no_supervision_is_not_supervision():
    """A real wording, and the one type in the table that means the opposite."""
    answer, _ = crs.is_under_supervision(
        [_sentence('NO SUPERVISION', '01/01/2025', '12 Month(s)')], CLINIC)
    assert answer is None


def test_the_longest_running_term_is_the_one_reported():
    """ICOS repeats a term under every count, and revocation adds another."""
    answer, term = crs.is_under_supervision([
        _sentence('PROBATION', '01/01/2024', '1 Year(s)'),
        _sentence('PROBATION', '01/01/2025', '5 Year(s)'),
        _sentence('PROBATION', '01/01/2024', '2 Year(s)'),
    ], CLINIC)
    assert answer == "YES"
    assert term[3] == date(2030, 1, 1)


def test_an_expired_term_does_not_hide_a_running_one():
    answer, _ = crs.is_under_supervision([
        _sentence('PROBATION', '01/01/1995', '1 Year(s)'),
        _sentence('PROBATION', '01/01/2025', '5 Year(s)'),
    ], CLINIC)
    assert answer == "YES"


# -- and where it refuses to answer -----------------------------------------

@pytest.mark.parametrize('sentences', [None, [], [_sentence('FINE', '', '')]])
def test_nothing_to_go_on_is_left_blank(sentences):
    assert crs.is_under_supervision(sentences, CLINIC) == (None, None)


def test_a_term_with_no_duration_is_left_blank():
    """An open-ended row has no end date, so it cannot answer the question."""
    answer, _ = crs.is_under_supervision(
        [_sentence('PROBATION', '01/01/2025', '')], CLINIC)
    assert answer is None


def test_a_term_with_no_date_is_left_blank():
    answer, _ = crs.is_under_supervision(
        [_sentence('PROBATION', '', '3 Year(s)')], CLINIC)
    assert answer is None


def test_no_is_never_written():
    """The whole point of leaving it blank. A NO would assert that nobody is on
    parole and no term was extended, neither of which ICOS can tell Napier."""
    for sentences in ([], [_sentence('PROBATION', '01/01/1990', '1 Year(s)')],
                      [_sentence('PRISON', '01/01/2025', '5 Year(s)')]):
        assert crs.is_under_supervision(sentences, CLINIC)[0] != "NO"


# -- the parser -------------------------------------------------------------

CHARGES_PAGE = b"""
<html><body><table>
<tr><td><font>Sentence</font></td><td><font>Sentence Date</font></td>
<td><font>Duration</font></td><td><font>Fine</font></td>
<td><font>Appeal</font></td><td><font>Judge</font></td>
<td><font>Facility Type</font></td><td><font>Attorney</font></td>
<td><font>Restitution</font></td><td><font>Drug</font></td>
<td><font>Extradition</font></td><td><font>Lic. Revoked</font></td>
<td><font>DDS</font></td><td><font>Batterer</font></td></tr>
<tr><td><font>PROBATION</font></td><td><font>07/02/2001</font></td>
<td><font>5 Year(s)</font></td><td><font></font></td>
<td><font></font></td><td><font>SYNTHETIC, JUDGE</font></td>
<td><font></font></td><td><font>N</font></td>
<td><font>N</font></td><td><font>N</font></td>
<td><font>N</font></td><td><font>N</font></td>
<td><font>N</font></td><td><font></font></td></tr>
</table></body></html>
"""


def test_the_sentence_row_is_read_off_the_page():
    case = {'id': '00000  FECR000000'}
    case_parser.parse_case_charges(CHARGES_PAGE, case)
    assert case['sentences'] == [
        {'type': 'PROBATION', 'date': '07/02/2001', 'duration': '5 Year(s)'}]


def test_the_header_row_is_not_mistaken_for_a_sentence():
    case = {'id': '00000  FECR000000'}
    case_parser.parse_case_charges(CHARGES_PAGE, case)
    assert all(s['type'] != 'Sentence' for s in case['sentences'])


def test_a_page_with_no_sentence_table_yields_none():
    case = {'id': '00000  FECR000000'}
    case_parser.parse_case_charges(
        b"<html><body><table><tr><td><font>Nothing here</font></td></tr>"
        b"</table></body></html>", case)
    assert case['sentences'] == []


def test_a_renamed_column_stops_the_read_rather_than_shifting_it():
    """The header is matched exactly so that a changed page reads as no data.

    Reading the wrong cell as a date is the failure worth designing against:
    it would put a confident YES in column I off a number that is not a term.
    """
    page = CHARGES_PAGE.replace(b'Lic. Revoked', b'License Revoked')
    case = {'id': '00000  FECR000000'}
    case_parser.parse_case_charges(page, case)
    assert case['sentences'] == []


# -- end to end into the workbook -------------------------------------------

def _case(sentences, disposition='GUILTY'):
    return {
        'id': '00000  FECR000000', 'county': 'SYNTHETIC',
        'charges': [{'charge': '124.401', 'description': 'SYNTHETIC OFFENSE',
                     'disposition': [disposition], 'offenseDate': '01/01/1900',
                     'dispositionDate': '02/02/1901'}],
        'financials': [], 'summary_categories': [],
        'sentences': sentences,
    }


def _row(case):
    sheet = load_workbook(FULL)['CASE DATA']
    crs.process_case(case, sheet, crs.FIRST_CASE_ROW, CLINIC)
    row = str(crs.FIRST_CASE_ROW)
    return sheet['I' + row].value, sheet['V' + row].value


def test_a_running_term_reaches_column_i():
    value, _ = _row(_case([_sentence('PROBATION', '01/01/2025', '3 Year(s)')]))
    assert value == "YES"


def test_an_expired_term_leaves_column_i_alone():
    value, _ = _row(_case([_sentence('PROBATION', '01/01/1995', '1 Year(s)')]))
    assert value is None


def test_the_row_says_what_the_yes_rests_on():
    """A YES that staff cannot audit is worse than a blank, because this one
    feeds a dollar figure on the expungement sheet."""
    _, note = _row(_case([_sentence('PROBATION', '01/01/2025', '3 Year(s)')]))
    assert 'probation' in note
    assert '3 Year(s)' in note
    assert '01/01/2025' in note
    assert '01/01/2028' in note


def test_the_agency_keeps_its_capitals_in_the_note():
    """The note lowercases the ICOS wording so it reads as a sentence, which
    turned the Department of Correctional Services into "dcs" on a workbook
    somebody files off."""
    _, note = _row(_case([
        _sentence('PROBATION - OTHER THAN DCS', '01/01/2025', '3 Year(s)')]))
    assert 'probation - other than DCS term' in note


def test_term_wording_only_shouts_where_icos_meant_to():
    assert crs.term_wording('PROBATION') == 'probation'
    assert crs.term_wording('DRUG COURT') == 'drug court'
    assert crs.term_wording('PROBATION - OTHER THAN DCS') == \
        'probation - other than DCS'
    assert crs.term_wording('') == ''


def test_the_note_admits_what_napier_cannot_see():
    _, note = _row(_case([_sentence('PROBATION', '01/01/2025', '3 Year(s)')]))
    assert 'parole' in note


def test_the_supervision_note_does_not_displace_the_money_note():
    _, money_only = _row(_case([]))
    _, both = _row(_case([_sentence('PROBATION', '01/01/2025', '3 Year(s)')]))
    assert money_only, 'this case is supposed to carry a fee note'
    assert money_only in both


def test_a_case_with_no_sentences_writes_nothing_in_column_i():
    """Today's behaviour, which must survive: a blank reads as n/a and that is
    the honest answer when Napier has nothing."""
    value, _ = _row(_case([]))
    assert value is None


def test_the_clinic_date_decides_not_today():
    """Reopening a workbook next year must not change what column I said."""
    case = _case([_sentence('PROBATION', '01/01/2025', '3 Year(s)')])
    sheet = load_workbook(FULL)['CASE DATA']
    crs.process_case(case, sheet, crs.FIRST_CASE_ROW, date(2040, 1, 1))
    assert sheet['I' + str(crs.FIRST_CASE_ROW)].value is None


def test_a_civil_case_can_still_report_supervision():
    """The sentence table is read whatever get_dominant_charge made of the
    adjudications, so the write is outside that branch."""
    case = {
        'id': '00000  SCSC000000', 'county': 'SYNTHETIC',
        'charges': [], 'financials': [], 'summary_categories': [],
        'summary_created_date': '01/01/1900',
        'summary_disposition_date': '02/02/1901',
        'summary_dispo_status': 'SYNTHETIC STATUS',
        'sentences': [_sentence('PROBATION', '01/01/2025', '3 Year(s)')],
    }
    value, note = _row(case)
    assert value == "YES"
    assert 'probation' in note
