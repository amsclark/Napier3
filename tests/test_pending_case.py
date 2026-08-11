"""A case no court has ruled on yet, and what the SOL sheet makes of it.

ICOS prints the Adjudication block on every count whether or not there is an
adjudication in it. On an open case the block is there and its cells are empty,
so Napier has no disposition and no disposition date, and column D goes out
blank.

That blank is deliberate and load bearing. The EXPUNGEMENT sheet counts rows
with no disposition date as the pending charges that block expungement under
901C.2, so filling column D with a guess would hide an open case from the one
sheet that has to see it.

The SOL sheet reads the same cell as a number. Its three columns are

  C  =IFERROR(IF('CASE DATA'!D4+7300<'BASIC INFO'!$B$3,'CASE DATA'!J4+'CASE DATA'!K4,0),0)
  D  =IFERROR(IF('CASE DATA'!D4+7300<'BASIC INFO'!$B$3,'CASE DATA'!L4,0),0)
  E  =IFERROR(IF('CASE DATA'!D4+7300>'BASIC INFO'!$B$3,SUM('CASE DATA'!J4:S4),0),0)

and a blank cell in that arithmetic is zero, which is a date in 1900. Measured
against the real workbook with LibreOffice doing the arithmetic: a pending row
carrying 300 indigent defense, 50 collection costs and 120 jail and room and
board reports 350 barred and 120 twenty years old, and the fines and surcharges
on the same row land in none of the three columns. No court has decided the
case, so there is no judgment for the twenty years to have run against.

Napier cannot fix the sheet and must not fill column D, so the row says so.

The pages here are synthetic. This repo is public and a real charges page is one
person's unredacted criminal record.
"""

import os
import sys
from decimal import Decimal

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import case_parser
import crs

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')


def _cells(*cells):
    return '<tr>%s</tr>' % ''.join(
        '<td><font size="2">%s</font></td>' % cell for cell in cells)


def charges_page(outcome, when):
    """One count, adjudicated or not.

    outcome of None is the open case: ICOS prints the Adjudication block with
    every cell in it empty, which is what an undecided charge looks like.
    """
    html = ['<html><body><table>',
            _cells('Count 01', 'Original Charge'),
            _cells('Charge:', '321.285', 'Description:', 'SYNTHETIC SPEEDING'),
            _cells('Offense Date:', '01/01/1900', 'Arrest Date:', ''),
            _cells('Adjudication')]
    if outcome is None:
        html.append(_cells('Charge:', '', 'Description:', ''))
        html.append(_cells('Adjudication:', '', 'Adjudication Date:', ''))
    else:
        html.append(_cells('Charge:', '321.285',
                           'Description:', 'SYNTHETIC SPEEDING'))
        html.append(_cells('Adjudication:', outcome,
                           'Adjudication Date:', when))
    html.append('</table></body></html>')
    return ''.join(html).encode('utf-8')


def _case(outcome=None, when='', costs='89.50', case_date='', status=''):
    case = {'id': '00000  STA0000000', 'county': 'SYNTHETIC',
            'financials': [], 'sentences': [],
            'summary_created_date': '01/01/1900',
            'summary_disposition_date': case_date,
            'summary_dispo_status': status}
    case_parser.parse_case_charges(charges_page(outcome, when), case)
    case['summary_categories'] = [
        {'label': label,
         'original': Decimal(costs) if label == 'FINE' else Decimal('0'),
         'paid': Decimal('0'),
         'due': Decimal(costs) if label == 'FINE' else Decimal('0')}
        for label in ('COSTS', 'FINE', 'SURCHARGE', 'RESTITUTION', 'OTHER')]
    case['total_due'] = '$' + costs
    return case


def _row(**kwargs):
    sheet = load_workbook(FULL)['CASE DATA']
    crs.process_case(_case(**kwargs), sheet, crs.FIRST_CASE_ROW)
    row = str(crs.FIRST_CASE_ROW)
    return {column: sheet[column + row].value for column in 'DGV'}


# -- the blank stays blank ----------------------------------------------------

def test_column_d_is_still_blank_on_an_open_case():
    """The guard on the fix, and it matters more than the caveat does.

    EXPUNGEMENT counts rows with no disposition date as pending charges. A fix
    that put a date in column D to keep the SOL sheet happy would take an open
    case off the one sheet that has to show it.
    """
    cells = _row()
    assert cells['D'] in (None, '')
    assert cells['G'] in (None, '')


# -- and the row says what that costs -----------------------------------------

def test_an_open_case_that_owes_money_says_the_sol_figures_are_unanswered():
    cells = _row()
    note = cells['V'] or ''
    assert 'no adjudication' in note, note
    assert '20 year' in note, note
    assert 'unanswered' in note.lower(), note


def test_the_caveat_names_both_sheets_that_read_column_d():
    """Somebody reading this has to know why the blank cannot just be filled."""
    note = _row()['V'] or ''
    assert 'EXPUNGEMENT' in note, note
    assert 'SOL' in note, note


def test_an_open_case_that_owes_nothing_stays_quiet():
    """All three SOL columns are zero on a row with no money, whatever column D
    says, so there is no wrong figure to warn about."""
    assert 'no adjudication' not in (_row(costs='0')['V'] or '')


def test_a_decided_case_is_left_alone():
    """The common case: 208 of the 210 captured cases have a disposition."""
    cells = _row(outcome='GUILTY', when='02/02/1901')
    assert cells['D'] == '02/02/1901'
    assert 'no adjudication' not in (cells['V'] or '')


def test_a_case_dated_off_the_summary_is_left_alone():
    """No count was adjudicated but the case itself was disposed.

    Column D gets the case level date, so the SOL sheet has a real date to work
    from and there is nothing to caveat.
    """
    cells = _row(case_date='02/02/1901', status='DISMISSED')
    assert cells['D'] == '02/02/1901'
    assert 'no adjudication' not in (cells['V'] or '')


def test_the_caveat_joins_what_the_money_left_in_column_v():
    """Column V belongs to process_financials first. This joins, never replaces.

    The money sentence used to be the whole-row fallback, which named
    MISCELLANEOUS. Since 6 August a fine the itemization omits is recovered into
    the fine column instead of stranding the row, so the sentence here is the
    unreconciled one. What this test is about is that a financial sentence and
    the adjudication caveat both survive in one cell.
    """
    sheet = load_workbook(FULL)['CASE DATA']
    case = _case()
    case['financials'] = [
        {'detail': 'SYNTHETIC UNCATEGORISED LINE', 'amount': Decimal('10.00'),
         'paid': None},
    ]
    crs.process_case(case, sheet, crs.FIRST_CASE_ROW)
    note = sheet['V' + str(crs.FIRST_CASE_ROW)].value or ''
    assert 'did not add up' in note, note
    assert 'no adjudication' in note, note
