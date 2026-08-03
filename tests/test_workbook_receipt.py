"""What the workbook says about what is not in it.

A criminal record summary does not stay on the screen it was built on. It gets
emailed to the attorney taking the case, dropped on a shared drive, and opened
three weeks later by somebody who never watched it run. Everything Napier used
to say about a run coming back short lived on the finish page and in the
progress log, and both of those are gone two hours after the run.

So a workbook that Iowa Courts had refused two cases out of looked exactly like
a workbook of a client who only ever had the other cases. Whoever opened it had
no way to tell, and the way that goes wrong is somebody advising on a record
they think is complete.

The file says it now, on the sheet that opens, above the fold.

Every case number is the synthetic 00000 FECR000000 family and no real person
appears anywhere in here, because this repository is public.
"""

import os
import sys
from datetime import date

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

os.environ['NAPIER_DISABLE_BACKGROUND'] = '1'

import actions
import crs

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')

CLINIC = date(2026, 7, 31)
FRESH = '01/01/2020'

MISSING = ['00000  FECR000001', '00000  SRCR000002']

# Ask the sheet where its own lines are. Hard-coding the rows here meant that
# adding one line to the header broke nine tests that were not about the header.
MISSING_LABEL = 'A%d' % actions.SUMMARY_ROWS['missing']
MISSING_LINE = 'B%d' % actions.SUMMARY_ROWS['missing']
CAVEAT_LINE = 'A%d' % actions.SUMMARY_ROWS['caveat']


def built(failed=(), rows=({'J': 250},)):
    """A real CRS workbook, built knowing what would not come off ICOS."""
    workbook = load_workbook(FULL)
    sheet = workbook['CASE DATA']
    for offset, values in enumerate(rows):
        row = crs.FIRST_CASE_ROW + offset
        cells = dict({'A': '00000  FECR000000', 'B': 'SYNTHETIC',
                      'D': FRESH, 'G': 'GTR'}, **values)
        for column, value in cells.items():
            sheet[column + str(row)] = value
    actions.build_action_sheet(workbook, [], len(rows), CLINIC,
                               'SYNTHETIC CLIENT', failed)
    return workbook['ACTION LIST']


class TestWhatTheLineSays:
    def test_a_run_that_got_everything_says_so_outright(self):
        """Silence would be ambiguous. A blank cell reads as a sheet that was
        never filled in, and the whole point is a reader who can tell the
        difference without asking whoever ran it."""
        sheet = built()
        assert sheet[MISSING_LABEL].value == 'Not in this file'
        assert sheet[MISSING_LINE].value == "none. Every case the search turned up is here."

    def test_a_short_run_names_every_case_it_could_not_get(self):
        sheet = built(failed=MISSING)
        line = sheet[MISSING_LINE].value
        for case_id in MISSING:
            assert case_id in line
        assert line.startswith("2 cases")

    def test_one_missing_case_is_not_described_in_the_plural(self):
        line = built(failed=MISSING[:1])[MISSING_LINE].value
        assert line.startswith("1 case Iowa Courts")
        assert "It is on no sheet" in line
        assert "Look it up" in line

    def test_it_says_what_to_do_about_it(self):
        """Naming a case number without saying it has to be looked up leaves
        the reader to work out whether it matters, and the answer is that it
        always matters."""
        line = built(failed=MISSING)[MISSING_LINE].value
        assert "on no sheet in this workbook" in line
        assert "Look them up on Iowa Courts" in line
        assert "the whole record" in line

    def test_the_missing_line_is_not_set_in_the_same_type_as_the_rest(self):
        """A reader skimming a summary sheet reads the values, not the labels.
        This one has to stop them."""
        quiet = built()[MISSING_LINE].font
        loud = built(failed=MISSING)[MISSING_LINE].font
        assert loud.bold and not quiet.bold


class TestWhereItSits:
    def test_it_is_on_the_sheet_that_opens(self):
        workbook = load_workbook(FULL)
        actions.build_action_sheet(workbook, [], 0, CLINIC, 'SYNTHETIC CLIENT',
                                   MISSING)
        assert workbook.sheetnames[0] == 'ACTION LIST'

    def test_and_above_the_actions_rather_than_under_them(self):
        """A list of forty ranked actions is a lot of scrolling, and a warning
        underneath it is a warning nobody reaches."""
        assert 7 < actions.FIRST_ACTION_ROW

    def test_the_caveat_and_the_action_rows_came_down_with_it(self):
        """The row it took was the caveat's. Everything below has to have moved
        or the sheet writes the header over the caveat and the first action
        over the header."""
        sheet = built(failed=MISSING)
        assert 'has been read by a lawyer' in sheet[CAVEAT_LINE].value
        header = sheet.cell(actions.FIRST_ACTION_ROW - 1, 1)
        assert header.value == actions.ACTION_HEADERS[0][0]
        assert header.font.bold
        assert sheet.freeze_panes == 'A' + str(actions.FIRST_ACTION_ROW)

    def test_the_summary_rows_above_it_are_untouched(self):
        sheet = built(failed=MISSING)
        assert sheet['A2'].value == 'Clinic date'
        assert sheet['A3'].value == 'Cases read'
        assert sheet['A4'].value == 'Total owed'
        assert sheet['A5'].value == 'Payments on record'
        assert sheet['A6'].value == 'Registration hold'


class TestItSurvivesTheBuild:
    def test_the_function_the_jobs_call_writes_it_to_the_file(self, tmp_path):
        """Everything above works on the sheet in memory. This is the function
        crs_task calls, and a line that is written and then not saved is worse
        than none, because the file looks like it was checked."""
        import tasks

        monkeypatched = tasks.tmp_dir
        tasks.tmp_dir = str(tmp_path) + os.sep
        try:
            path, _, _ = tasks.build_workbook(
                [], 'SYNTHETIC CLIENT', '01/01/1900', False, MISSING)
        finally:
            tasks.tmp_dir = monkeypatched

        try:
            sheet = load_workbook(path)['ACTION LIST']
        finally:
            os.unlink(path)
        assert MISSING[0] in sheet[MISSING_LINE].value
        assert MISSING[1] in sheet[MISSING_LINE].value

    def test_the_lite_workbook_gets_it_too(self, tmp_path):
        """Lite is what most clinics actually hand out, and it is the one where
        a reader has the least else to go on."""
        import tasks

        monkeypatched = tasks.tmp_dir
        tasks.tmp_dir = str(tmp_path) + os.sep
        try:
            path, _, _ = tasks.build_workbook(
                [], 'SYNTHETIC CLIENT', '01/01/1900', True, MISSING)
        finally:
            tasks.tmp_dir = monkeypatched

        try:
            sheet = load_workbook(path)['ACTION LIST']
        finally:
            os.unlink(path)
        assert MISSING[0] in sheet[MISSING_LINE].value


class TestWhatItMustNotCarry:
    def test_no_credential_reaches_the_file(self):
        """Workbooks get emailed by staff, which puts them outside the ESA
        agreement's reach. Case numbers are public record and may travel. The
        account they were pulled with is half a credential and may not."""
        sheet = built(failed=MISSING)
        line = " ".join(str(sheet[cell].value) for cell in
                        ('B2', 'B3', 'B5', 'B6', MISSING_LINE, CAVEAT_LINE))
        for smell in ('ILA', 'drakelegalclinic', 'password', 'user ID'):
            assert smell not in line
