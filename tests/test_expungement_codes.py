"""The 910.7 columns computing an error on the rows they exist to answer.

CASE DATA columns W to AH take column F apart, one statute per column, so the
expungement sheet can test the codes one at a time. The template did it with an
array formula that pads its unused slots with #VALUE!.

Column M reads that range with COUNTIF, which steps over errors, and answers.
Column N reads it with LEFT, which does not, and returns #VALUE!. N is gated on
M being "YES", so it is only ever evaluated on a case that cleared the first
screen, and columns O to S all read N.

Built from the 210 captured cases and computed by LibreOffice: the 17 rows
where M says YES are exactly the 17 rows where N says #VALUE!, no exceptions
either way, and O POSSIBLE PENDING CHARGES, P DISCHARGEABLE DEBT, Q TIME BARRED
DEBT, R TOTAL DEBT and S TIME ELAPSED are errors on all 17. The offences are
forgery, third degree theft, drug possession, assault, harassment and fugitive
from justice. With the split written as text instead, all 17 read "eligible",
the five columns compute, and column M does not change its mind about anybody.

The cases here are synthetic. This repo is public and a real charges page is
one person's unredacted criminal record.
"""

import os
import sys
from decimal import Decimal

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import case_parser
import statutes
import tasks

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')
LITE = os.path.join(ROOT, 'CRS Lite 3.5.5.xlsx')

# The depth both templates dragged the array formula down to, so the row past
# which a long case list had no split at all.
TEMPLATE_DEPTH = 300

COLUMNS = list(range(statutes.FIRST_COLUMN, statutes.LAST_COLUMN + 1))


def _cells(*cells):
    return '<tr>%s</tr>' % ''.join(
        '<td><font size="2">%s</font></td>' % cell for cell in cells)


def charges_page(charge='321.285'):
    """One count, convicted, so the row carries a code and a date."""
    return ''.join([
        '<html><body><table>',
        _cells('Count 01', 'Original Charge'),
        _cells('Charge:', charge, 'Description:', 'SYNTHETIC OFFENSE'),
        _cells('Offense Date:', '01/01/1900', 'Arrest Date:', ''),
        _cells('Adjudication'),
        _cells('Charge:', charge, 'Description:', 'SYNTHETIC OFFENSE'),
        _cells('Adjudication:', 'GUILTY', 'Adjudication Date:', '02/02/1901'),
        '</table></body></html>']).encode('utf-8')


def synthetic_cases(count, charge='321.285'):
    cases = []
    page = charges_page(charge)
    for n in range(count):
        case = {'id': '00000  FECR%06d' % n, 'county': 'SYNTHETIC',
                'financials': [], 'sentences': [],
                'summary_created_date': '01/01/1900',
                'summary_disposition_date': '', 'summary_dispo_status': ''}
        case_parser.parse_case_charges(page, case)
        case['summary_categories'] = [
            {'label': label,
             'original': Decimal('10.00' if label == 'FINE' else '0'),
             'paid': Decimal('0'),
             'due': Decimal('10.00' if label == 'FINE' else '0')}
            for label in ('COSTS', 'FINE', 'SURCHARGE', 'RESTITUTION', 'OTHER')]
        case['total_due'] = '$10.00'
        cases.append(case)
    return cases


def split_row(sheet, row):
    return [sheet.cell(row=row, column=column).value for column in COLUMNS]


def is_empty(value):
    """Blank as far as LEFT() is concerned.

    An empty string is not saved as anything: openpyxl writes the cell with no
    value at all and it reads back as None. Either way LEFT() of it is "" and
    the comparison in column N is simply false, which is the entire fix.
    """
    return value is None or value == ''


def with_statutes(references, template=FULL):
    """A workbook with column F written by hand and the split run over it."""
    workbook = load_workbook(template)
    sheet = workbook['CASE DATA']
    for offset, reference in enumerate(references):
        sheet.cell(row=4 + offset, column=1).value = '00000  FECR%06d' % offset
        sheet.cell(row=4 + offset, column=statutes.STATUTE_COLUMN).value = \
            reference
    overflow = statutes.write_statute_split(workbook, len(references))
    return workbook, sheet, overflow


# -- the split itself --------------------------------------------------------

def test_one_statute_lands_in_the_first_column():
    _, sheet, _ = with_statutes(['321.285'])
    assert split_row(sheet, 4)[0] == '321.285'


def test_the_slots_it_did_not_fill_are_blank_and_not_errors():
    """The whole defect. LEFT() on a #VALUE! is a #VALUE!, and column N is
    nothing but LEFT() over this range."""
    _, sheet, _ = with_statutes(['321.285'])
    rest = split_row(sheet, 4)[1:]
    assert rest == [''] * (statutes.SLOTS - 1), rest


def test_every_slot_is_text_the_sheet_can_take_the_left_of():
    """Not a formula, not None, on any case row. A formula is what the
    template left and what LEFT() choked on."""
    _, sheet, _ = with_statutes(['715A.2(A);', '124.401(5);124.401(5)',
                                 None, 'n/a', '321.285'])
    for row in range(4, 9):
        for value in split_row(sheet, row):
            assert isinstance(value, str), (row, value)


def test_two_statutes_both_get_a_column():
    _, sheet, _ = with_statutes(['714.2(3);714.2(3)'])
    assert split_row(sheet, 4)[:2] == ['714.2(3)', '714.2(3)']


def test_a_trailing_semicolon_keeps_its_empty_slot():
    """'715A.2(A);' is what ICOS gave for a real case, and the array formula
    read it as one code and one empty field. Position is what the sheet reads,
    so the shape has to survive."""
    assert statutes.split_codes('715A.2(A);') == ['715A.2(A)', '']


def test_a_row_with_no_statute_is_blank_all_the_way_across():
    _, sheet, _ = with_statutes([None])
    assert split_row(sheet, 4) == [''] * statutes.SLOTS


def test_the_civil_placeholder_is_left_as_it_is():
    """Column F says "n/a" on a civil case. It is not a statute and nothing on
    CODE SECTIONS matches it, so it may sit in the first slot unharmed."""
    _, sheet, _ = with_statutes(['n/a'])
    assert split_row(sheet, 4)[0] == 'n/a'


def test_spaces_around_a_code_are_taken_off():
    """'AL/ 3.03' is a real ordinance citation and LEFT() compares against
    unpadded chapter numbers."""
    _, sheet, _ = with_statutes(['321.285 ; 719.1'])
    assert split_row(sheet, 4)[:2] == ['321.285', '719.1']


# -- the rows the template had already written -------------------------------

def test_the_template_leaves_no_array_formula_behind_it():
    """Rows past the case list keep the formula and its #VALUE! padding, which
    reads as a broken workbook to anyone who scrolls right."""
    plain = load_workbook(FULL)['CASE DATA']
    assert plain.cell(row=TEMPLATE_DEPTH, column=statutes.FIRST_COLUMN).value \
        is not None, 'the template splitter did not reach that row'

    _, sheet, _ = with_statutes(['321.285'])
    for row in (5, 100, TEMPLATE_DEPTH):
        assert split_row(sheet, row) == [''] * statutes.SLOTS, row


def test_the_slot_dragged_one_column_too_far_goes_too():
    """Row 9 of both templates carries a thirteenth splitter cell in AI. The
    expungement sheet asks for W9:AH9 and never sees it, so it does nothing but
    print #VALUE! on a case row for the life of the workbook."""
    stray = statutes.LAST_COLUMN + 1
    plain = load_workbook(FULL)['CASE DATA']
    assert statutes._is_splitter(plain.cell(row=9, column=stray).value), \
        'the stray slot is not there, so nothing is being tested'

    _, sheet, _ = with_statutes(['321.285'] * 10)
    assert sheet.cell(row=9, column=stray).value == ''


def test_a_column_out_there_that_is_not_the_splitter_is_left_alone():
    """Only the array formula. Anything else off the right hand edge is
    somebody's working note and none of Napier's business."""
    workbook = load_workbook(FULL)
    sheet = workbook['CASE DATA']
    sheet.cell(row=9, column=statutes.LAST_COLUMN + 4).value = 'ari was here'
    sheet.cell(row=4, column=1).value = '00000  FECR000000'
    statutes.write_statute_split(workbook, 1)
    assert sheet.cell(row=9, column=statutes.LAST_COLUMN + 4).value == \
        'ari was here'


def test_a_case_list_past_the_template_splitter_is_still_split():
    """Both templates dragged the array formula to row 300, so case 298 had no
    split at all and the expungement sheet had nothing to screen."""
    count = TEMPLATE_DEPTH + 20
    _, sheet, _ = with_statutes(['719.1'] * count)
    last = 3 + count
    assert sheet.cell(row=last, column=1).value, 'no case on that row'
    assert split_row(sheet, last)[0] == '719.1'


def test_the_lite_template_is_split_too():
    """Lite drops the SOL and bankruptcy sheets but keeps the expungement one,
    and it carries the same array formula."""
    _, sheet, _ = with_statutes(['715A.2'], template=LITE)
    assert split_row(sheet, 4)[0] == '715A.2'
    assert split_row(sheet, 4)[1:] == [''] * (statutes.SLOTS - 1)


def test_an_empty_case_list_still_comes_out_clean():
    """No cases means no answers, but it may not mean a sheet of #VALUE!."""
    workbook = load_workbook(FULL)
    statutes.write_statute_split(workbook, 0)
    sheet = workbook['CASE DATA']
    for row in (4, TEMPLATE_DEPTH):
        assert split_row(sheet, row) == [''] * statutes.SLOTS, row


# -- more statutes than there are columns ------------------------------------

def test_a_case_with_more_statutes_than_slots_is_reported():
    """Twelve slots against a worst case of two across 210 captured cases, so
    this is a backstop. It still may not be silent."""
    codes = ['719.%d' % n for n in range(statutes.SLOTS + 3)]
    _, sheet, overflow = with_statutes([';'.join(codes)])
    assert list(overflow) == [4], overflow
    assert overflow[4] == codes[statutes.SLOTS:]
    assert split_row(sheet, 4) == codes[:statutes.SLOTS]


def test_a_case_that_fits_is_reported_as_nothing():
    _, _, overflow = with_statutes(['321.285', '714.2(3);714.2(3)'])
    assert overflow == {}


# -- through the real build --------------------------------------------------

def test_a_built_workbook_has_no_errors_waiting_in_the_split():
    """The whole point, through the path a clinic actually takes."""
    path, _, _ = tasks.build_workbook(
        synthetic_cases(3, '715A.2'), 'TEST CLIENT', '01/01/1980', False)
    try:
        sheet = load_workbook(path)['CASE DATA']
        assert sheet['F4'].value == '715A.2', \
            'the statute never reached column F, so nothing is proved'
        for row in range(4, 7):
            values = split_row(sheet, row)
            assert values[0] == '715A.2', (row, values)
            assert all(is_empty(value) for value in values[1:]), (row, values)
        # And nothing anywhere in the range is still a formula, which is what
        # the padding came out of.
        for row in range(4, TEMPLATE_DEPTH + 1):
            for value in split_row(sheet, row):
                assert not (isinstance(value, str) and value.startswith('=')), \
                    (row, value)
    finally:
        os.remove(path)


def test_the_overflow_note_reaches_the_staffer():
    """Column V is the only place the workbook talks to whoever opens it."""
    charge = ';'.join('719.%d' % n for n in range(statutes.SLOTS + 2))
    path, _, _ = tasks.build_workbook(
        synthetic_cases(1, charge), 'TEST CLIENT', '01/01/1980', False)
    try:
        sheet = load_workbook(path)['CASE DATA']
        note = sheet['V4'].value or ''
        assert 'by hand' in note, note
        assert '719.13' in note, note
    finally:
        os.remove(path)
