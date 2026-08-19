"""Three changes Iowa Legal Aid asked for on 5 August, and what each may not do.

All three are meant to change a label and no money. The tests that matter here
are the ones checking the second half of that, because a code on CASE DATA is
read by formulas on five other sheets and moving one quietly moves a client's
debt between dischargeable, exempt and not.

  OLD CASE CHARGE CODE reads OTH instead of leaving column G empty, which three
  sheets were rendering as "open charge" on a case closed in 1993.

  Fugitive from justice and violation of parole read CIV, keyed on the statute
  rather than on whatever the clerk typed in the disposition.

  COLLECTION BY CO ATTY moves from column P to column K.

The trap in the middle one is 908.11. Violation of probation is one section
away from violation of parole and there are 20 of them in the same captured
corpus carrying about $17,600, mostly real guilty pleas. A rule written on
chapter 908 rather than on the section would flip all of them to civil, which
is the opposite of what Iowa Legal Aid asked for. That is the test to keep.

Synthetic pages throughout. The repository is public and a real charges page is
one person's unredacted criminal record.
"""

import os
import sys
from datetime import date

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import case_parser
import crs

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FULL = os.path.join(ROOT, 'CRS 3.5.5.xlsx')

CLINIC = date(2026, 7, 31)


def _cells(*cells):
    return '<tr>%s</tr>' % ''.join(
        '<td><font size="2">%s</font></td>' % cell for cell in cells)


def charges_page(counts):
    """ICOS's real shape: each count charged, then adjudicated.

    counts is (statute, description, outcome). outcome of None is the page ICOS
    serves before a court has ruled, with the adjudication block present and
    its cells empty.
    """
    html = ['<html><body><table>']
    for number, (statute, description, outcome) in enumerate(counts, start=1):
        html.append(_cells('Count %02d' % number, 'Original Charge'))
        html.append(_cells('Charge:', statute, 'Description:', description))
        html.append(_cells('Offense Date:', '01/01/1900', 'Arrest Date:', ''))
        html.append(_cells('Adjudication'))
        if outcome is None:
            html.append(_cells('Charge:', '', 'Description:', ''))
            html.append(_cells('Adjudication:', '', 'Adjudication Date:', ''))
        else:
            html.append(_cells('Charge:', statute, 'Description:', description))
            html.append(_cells('Adjudication:', outcome,
                               'Adjudication Date:', '02/02/1901'))
    html.append('</table></body></html>')
    return ''.join(html).encode('utf-8')


def _case(counts, status='', dispo_date='', financials=None):
    case = {'id': '00000  SMSM000000', 'county': 'SYNTHETIC',
            'financials': financials or [], 'summary_categories': [],
            'sentences': [], 'summary_created_date': '01/01/1900',
            'summary_disposition_date': dispo_date,
            'summary_dispo_status': status}
    case_parser.parse_case_charges(charges_page(counts), case)
    return case


def _row(case):
    sheet = load_workbook(FULL)['CASE DATA']
    unknown = crs.process_case(case, sheet, crs.FIRST_CASE_ROW, CLINIC)
    i = str(crs.FIRST_CASE_ROW)
    cells = {column: sheet[column + i].value
             for column in ('A', 'D', 'F', 'G', 'J', 'K', 'P', 'V')}
    return cells, unknown


# -- OLD CASE CHARGE CODE ---------------------------------------------------

OLD_CASE = [('CR/OLDCASE', 'OLD CASE CHARGE CODE', None)]


class TestTheOldCaseCode:
    def test_it_stops_reading_as_an_open_charge(self):
        """Empty is 0 to Excel and SOL, BANKRUPTCY and EXEMPTIONS all render
        IF(G=0, "open charge", G). A case closed in 1993 was printing as an
        open charge on three sheets."""
        cells, _ = _row(_case(OLD_CASE, status='CLOSED',
                              dispo_date='09/23/1901'))
        assert cells['G'] == 'OTH'

    def test_it_is_in_no_cleared_set_so_no_money_moves(self):
        """The whole reason this one is safe. OTH appears in no formula in
        either template, so nothing that sorts a balance can see it."""
        workbook = load_workbook(FULL)
        for name in ('BANKRUPTCY', 'SOL', 'EXEMPTIONS',
                     'EXPUNGEMENT & 910.7'):
            sheet = workbook[name]
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        assert '"OTH"' not in cell.value, (
                            "%s!%s tests for OTH" % (name, cell.coordinate))

    def test_the_date_icos_gave_is_still_kept(self):
        """The guard on the fix that came before this one. A blank column D is
        how the expungement sheet counts pending charges, so losing the date
        here would put the case back to blocking expungement."""
        cells, _ = _row(_case(OLD_CASE, status='CLOSED',
                              dispo_date='09/23/1901'))
        assert cells['D'] == '09/23/1901'

    def test_the_status_it_could_not_read_is_still_reported(self):
        """Coding the row does not mean Napier understood CLOSED. The note
        changes to the one that describes a row carrying a code, but the
        wording still travels."""
        cells, unknown = _row(_case(OLD_CASE, status='CLOSED',
                                    dispo_date='09/23/1901'))
        assert unknown == [('CLOSED', True)]
        assert 'CLOSED' in (cells['V'] or '')

    def test_a_closed_case_that_is_not_an_old_case_is_left_alone(self):
        """Keyed on the charge code, not on the status. CLOSED turns up on
        cases with a real adjudication and this must not speak for them."""
        cells, _ = _row(_case([('321.285', 'SYNTHETIC SPEED', None)],
                              status='CLOSED', dispo_date='09/23/1901'))
        assert cells['G'] in (None, '')

    def test_a_genuinely_pending_case_still_reads_as_open(self):
        """No status, no date, nothing adjudicated. This is the row the "open
        charge" rendering exists for and it has to keep working."""
        cells, _ = _row(_case([('321.285', 'SYNTHETIC SPEED', None)]))
        assert cells['G'] in (None, '')
        assert cells['D'] in (None, '')

    def test_an_old_case_that_did_get_adjudicated_codes_off_that(self):
        """If ICOS ever does print an adjudication on one of these, the
        adjudication is the better answer and OTH must not overwrite it."""
        cells, _ = _row(_case([('CR/OLDCASE', 'OLD CASE CHARGE CODE',
                                'GUILTY')]))
        assert cells['G'] == 'GTR'


# -- CIV on the statute -----------------------------------------------------

class TestFugitiveAndParoleReadCivil:
    @pytest.mark.parametrize('outcome', ['GUILTY', 'GUILTY BY COURT',
                                         'DEFERRED', 'SYNTHETIC NONSENSE'])
    def test_fugitive_reads_civil_whatever_the_clerk_typed(self, outcome):
        """Her actual complaint. The four captured fugitive cases came out as
        four different codes because Napier reads the disposition wording.
        SYNTHETIC NONSENSE is the one that was coming out OTH."""
        cells, _ = _row(_case([('820.2', 'FUGITIVE FROM JUSTICE - 1989',
                                outcome)]))
        assert cells['G'] == 'CIV'

    @pytest.mark.parametrize('outcome', ['WITHDRAWN', 'DISMISSED',
                                         'ACQUITTED', 'NOT FILED',
                                         'CHANGE OF VENUE'])
    def test_a_cleared_word_does_not_hide_the_statute(self, outcome):
        """Every wording a clerk can use to clear a count, one answer.

        None of these leaves an adjudicated statute behind, so 820.2 never
        reaches column F and the statute rule reads the pre-filter statutes
        instead. It used to stand down here, to keep the case in the
        EXPUNGEMENT & 910.7 cleared set that CIV is not in. That is what made
        the same hold come out five different ways, and Iowa Legal Aid found
        it: 8/07 asked for CIV on a transferred parole violation, 8/19 asked
        for it again on three disposed NOT FILED. A civil-in-nature case is
        not eligible for dismissed-or-acquitted expungement anyway, so the
        cleared code was buying a YES in that column that nobody could act
        on.
        """
        cells, _ = _row(_case([('820.2', 'FUGITIVE FROM JUSTICE - 1989',
                                outcome)]))
        assert cells['G'] == 'CIV'

    @pytest.mark.parametrize('outcome', ['CHANGE OF VENUE', 'NOT FILED',
                                         'DISMISSED', 'WITHDRAWN',
                                         'ACQUITTED'])
    def test_a_cleared_parole_violation_reads_civil_too(self, outcome):
        """Iowa Legal Aid's 8/19 report, and the 8/07 one it repeats.

        Three real Polk parole violations disposed NOT FILED, 00000
        AMCR000000, AMCR000000 and AMCR000000, came out NOTF while the ones
        disposed CHANGE OF VENUE came out CIV. Same statute, same kind of
        case, two answers.
        """
        cells, _ = _row(_case([('908.1', 'VIOLATION OF PAROLE - 1985',
                                outcome)]))
        assert cells['G'] == 'CIV'

    def test_violation_of_parole_reads_civil(self):
        cells, _ = _row(_case([('908.1', 'VIOLATION OF PAROLE - 1985',
                                'GUILTY')]))
        assert cells['G'] == 'CIV'

    def test_a_subsection_of_either_still_counts(self):
        cells, _ = _row(_case([('820.2(1)', 'SYNTHETIC FUGITIVE', 'GUILTY')]))
        assert cells['G'] == 'CIV'


class TestProbationViolationStaysPut:
    """The 20 cases and about $17,600 that a chapter rule would have taken."""

    @pytest.mark.parametrize('outcome,expected', [
        ('GUILTY', 'GTR'),
        ('GUILTY - NEGOTIATED/VOLUN PLEA', 'GPL'),
        ('DEFERRED', 'DEF'),
        ('ADJUDICATED', 'JUV'),
    ])
    def test_violation_of_probation_keeps_its_own_code(self, outcome, expected):
        """908.11 is one section from 908.1 and it is a real conviction. Iowa
        Legal Aid said the conviction wins."""
        cells, _ = _row(_case([('908.11', 'VIOLATION OF PROBATION - 1985',
                                outcome)]))
        assert cells['G'] == expected

    def test_a_subsection_of_908_11_is_not_908_1_either(self):
        cells, _ = _row(_case([('908.11(1)', 'SYNTHETIC PROBATION',
                                'GUILTY')]))
        assert cells['G'] == 'GTR'

    def test_the_rest_of_chapter_908_is_untouched(self):
        cells, _ = _row(_case([('908.3', 'SYNTHETIC CHAPTER 908', 'GUILTY')]))
        assert cells['G'] == 'GTR'

    def test_820_21_is_not_820_2(self):
        cells, _ = _row(_case([('820.21', 'SYNTHETIC CHAPTER 820', 'GUILTY')]))
        assert cells['G'] == 'GTR'


class TestAConvictionAlongsideOneOfThem:
    def test_the_conviction_wins(self):
        """A case holding someone as a fugitive alongside a real conviction is
        a case with a real conviction. Coding the row CIV would move the
        conviction's whole balance into dischargeable and exempt."""
        cells, _ = _row(_case([('820.2', 'SYNTHETIC FUGITIVE', 'GUILTY'),
                               ('124.401', 'SYNTHETIC CONTROLLED', 'GUILTY')]))
        assert cells['G'] == 'GTR'

    def test_two_of_them_together_are_still_civil(self):
        cells, _ = _row(_case([('820.2', 'SYNTHETIC FUGITIVE', 'GUILTY'),
                               ('908.1', 'SYNTHETIC PAROLE', 'GUILTY')]))
        assert cells['G'] == 'CIV'

    def test_an_unadjudicated_count_does_not_block_it(self):
        """Only adjudicated statutes reach column F, so a pending count next
        to a fugitive hold leaves the fugitive hold speaking alone."""
        cells, _ = _row(_case([('820.2', 'SYNTHETIC FUGITIVE', 'GUILTY'),
                               ('321.285', 'SYNTHETIC SPEED', None)]))
        assert cells['G'] == 'CIV'


class TestWhatCivilDoesDownstream:
    def test_it_is_in_the_dischargeable_and_exempt_sets(self):
        """Worth pinning, because this is the part Iowa Legal Aid is agreeing
        to when they ask for CIV. It moves the whole balance."""
        workbook = load_workbook(FULL)
        for name in ('BANKRUPTCY', 'EXEMPTIONS'):
            formulas = [cell.value for row in workbook[name].iter_rows()
                        for cell in row if isinstance(cell.value, str)]
            assert any('"CIV"' in formula for formula in formulas), name

    def test_it_does_not_make_a_case_expungeable(self):
        """The other half of what they were told. CIV is not in the 901C.2
        column's cleared set, so this does not turn a fugitive hold into an
        expungement candidate."""
        sheet = load_workbook(FULL)['EXPUNGEMENT & 910.7']
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and '"DISM"' in cell.value:
                    assert '"CIV"' not in cell.value, (
                        "%s puts CIV in the dischargeable-or-acquitted test"
                        % cell.coordinate)


# -- COLLECTION BY CO ATTY --------------------------------------------------

class TestCollectionByCountyAttorney:
    def test_it_lands_in_the_collection_column(self):
        assert crs.get_finance_column('COLLECTION BY CO ATTY') == 'K'

    def test_it_sits_with_the_other_collection_fees(self):
        """K is where the Linebarger and Department of Revenue fees go, which
        is what makes it collection debt rather than unknown money."""
        assert crs.get_finance_column('THIRD PARTY COLLECTION FEE') == 'K'
        assert crs.get_finance_column('IOWA REVENUE COLLECTION FEE') == 'K'
        assert crs.get_finance_column('COLLECTION BY CO ATTY') == \
            crs.get_finance_column('THIRD PARTY COLLECTION FEE')

    def test_the_other_unknown_fee_did_not_move_with_it(self):
        assert crs.get_finance_column('DELINQUENT REVOLVING FUND') == 'P'

    def test_an_unpaid_one_reaches_the_dischargeable_bucket(self):
        """The rare case this is for. 47 of the 49 captured rows are marked
        paid, so on almost every real case this changes nothing at all."""
        case = _case([('321.285', 'SYNTHETIC SPEED', 'GUILTY')],
                     financials=[{'detail': 'COLLECTION BY CO ATTY',
                                  'amount': 25.00, 'paid': None}])
        cells, _ = _row(case)
        assert cells['K'] == 25.00
        assert not cells['P']
